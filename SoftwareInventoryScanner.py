#!/usr/bin/env python3
"""
Network Software Inventory Scanner
----------------------------------
Standalone Tkinter desktop application for scanning Windows computers,
inventorying installed software, storing results in SQLite, and exporting data.
"""

# ------------------ Imports ------------------ #
from __future__ import annotations

import csv
import ctypes
import ipaddress
import json
import logging
import os
import platform
import queue
import shutil
import socket
import sqlite3
import subprocess
import threading
import time
import atexit
import signal
import stat
import tempfile
import uuid
import tkinter as tk
import statistics
from collections import deque

from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from datetime import datetime
from logging.handlers import RotatingFileHandler
from pathlib import Path
from tkinter import filedialog, ttk
from typing import Any, Dict, List, Optional, Sequence, Tuple
from contextlib import contextmanager

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl.utils
from openpyxl.utils import get_column_letter

# ------------------ Application Configuration ------------------ #
class AppConfig:
    """
    Centralized constants and default application settings.

    This class intentionally stores only static configuration values.
    Runtime state should be managed by service classes, data models, or UI
    controllers.
    """

    APP_NAME = "Network Software Inventory Scanner"
    APP_VERSION = "0.4.0"

    DEFAULT_WINDOW_GEOMETRY = "1200x800"
    DEFAULT_BACKGROUND = "#ECECEC"

    DEFAULT_PING_TIMEOUT = 1.0
    DEFAULT_DNS_TIMEOUT = 3.0
    DEFAULT_REGISTRY_TIMEOUT = 60.0

    DEFAULT_MAX_PING_WORKERS = 64
    DEFAULT_MAX_INVENTORY_WORKERS = 8

    MIN_PING_WORKERS = 1
    MAX_PING_WORKERS = 256
    MIN_INVENTORY_WORKERS = 1
    MAX_INVENTORY_WORKERS = 32

    LOG_FILE = "inventory_scanner.log"
    DATABASE_FILE = "inventory.db"
    PREFERENCES_FILE = "app_preferences.json"

    EXPORT_FORMATS = ("xml", "csv", "json")

    INVENTORY_STATUS_PENDING = "pending"
    INVENTORY_STATUS_SUCCESS = "success"
    INVENTORY_STATUS_PARTIAL = "partial"
    INVENTORY_STATUS_FAILED = "failed"
    INVENTORY_STATUS_SKIPPED = "skipped"

    SCAN_STATUS_SUCCESS = "success"
    SCAN_STATUS_WARNING = "warning"
    SCAN_STATUS_ERROR = "error"
    SCAN_STATUS_SKIPPED = "skipped"
    SCAN_STATUS_CANCELLED = "cancelled"

    REGISTRY_UNINSTALL_PATHS = (
        r"HKLM:\Software\Microsoft\Windows\CurrentVersion\Uninstall\*",
        r"HKLM:\Software\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall\*",
    )

    POWERSHELL_EXECUTABLE = "powershell.exe"

    LOG_MAX_BYTES = 5 * 1024 * 1024
    LOG_BACKUP_COUNT = 5

    SQLITE_TIMEOUT = 30.0

    LIVE_FEED_TIMESTAMP_FORMAT = "%H:%M:%S"
    STORAGE_TIMESTAMP_FORMAT = "%Y-%m-%dT%H:%M:%S"

    SAFE_HOSTNAME_MAX_LENGTH = 255
    SOFTWARE_GROUP_FIELDS = ("display_name", "display_version")

# ------------------ Logging Configuration ------------------ #
class AppLogger:
    """
    Standardized logging wrapper with rotating file logging and an in-memory
    queue for live GUI log display.
    """

    _logger = logging.getLogger("InventoryScanner")
    _log_queue: queue.Queue[str] = queue.Queue()
    _configured = False
    _lock = threading.Lock()

    @classmethod
    def configure_logging(cls) -> None:
        """
        Configure the application logger once.

        Logging output includes timestamp, severity, thread name, and message.
        Existing handlers are cleared to prevent duplicate log entries during
        reloads or repeated initialization.
        """
        with cls._lock:
            if cls._configured:
                return

            cls._logger.setLevel(logging.DEBUG)
            cls._logger.handlers.clear()
            cls._logger.propagate = False

            formatter = logging.Formatter(
                "[%(asctime)s] %(levelname)s "
                "[%(threadName)s] - %(message)s"
            )

            file_handler = RotatingFileHandler(
                AppConfig.LOG_FILE,
                maxBytes=AppConfig.LOG_MAX_BYTES,
                backupCount=AppConfig.LOG_BACKUP_COUNT,
                encoding="utf-8",
            )
            file_handler.setLevel(logging.DEBUG)
            file_handler.setFormatter(formatter)

            cls._logger.addHandler(file_handler)
            cls._configured = True

            cls.log_message("info", "Logging initialized.")

    @classmethod
    def log_message(cls, level: str, message: str) -> None:
        """
        Log a message and enqueue it for the GUI Logs tab.

        Args:
            level: Logging level name.
            message: Message to record.
        """
        normalized_level = str(level).strip().lower()
        safe_message = str(message).strip()

        if not safe_message:
            safe_message = "No log message provided."

        if not cls._configured:
            cls.configure_logging()

        log_methods = {
            "debug": cls._logger.debug,
            "info": cls._logger.info,
            "warning": cls._logger.warning,
            "error": cls._logger.error,
            "critical": cls._logger.critical,
        }

        log_method = log_methods.get(normalized_level, cls._logger.info)
        log_method(safe_message)

        timestamp = time.strftime(AppConfig.LIVE_FEED_TIMESTAMP_FORMAT)
        queue_record = (
            f"[{timestamp}] {normalized_level.upper()} "
            f"[{threading.current_thread().name}] - {safe_message}"
        )
        cls._log_queue.put(queue_record)

    @classmethod
    def get_log_queue(cls) -> queue.Queue[str]:
        """
        Return the thread-safe queue used by the GUI Logs tab.

        Returns:
            Queue containing formatted log messages.
        """
        return cls._log_queue

# ------------------ Preferences Management ------------------ #
class PreferencesManager:
    """
    Load, save, and update lightweight user preferences stored as JSON.

    Preference failures are handled gracefully so the application never crashes
    due to a missing, malformed, or inaccessible preferences file.
    """

    def __init__(self, preferences_file: str = AppConfig.PREFERENCES_FILE):
        """
        Initialize the preferences manager.

        Args:
            preferences_file: JSON file path used for persisted preferences.
        """
        self.preferences_file = Path(preferences_file)
        self.preferences: Dict[str, Any] = {}
        self.load_preferences()

    def load_preferences(self) -> Dict[str, Any]:
        """
        Load preferences from disk.

        Returns:
            Dictionary of loaded preferences, or defaults when unavailable.
        """
        if not self.preferences_file.exists():
            self.preferences = self._default_preferences()
            return self.preferences

        try:
            with self.preferences_file.open("r", encoding="utf-8") as file_obj:
                loaded_preferences = json.load(file_obj)

            if not isinstance(loaded_preferences, dict):
                raise ValueError("Preference file root must be a JSON object.")

            self.preferences = {
                **self._default_preferences(),
                **loaded_preferences,
            }
            AppLogger.log_message("info", "Preferences loaded.")
        except (OSError, json.JSONDecodeError, ValueError) as exc:
            AppLogger.log_message(
                "error",
                f"Failed to load preferences: {exc}",
            )
            self.backup_corrupt_preferences()
            self.preferences = self._default_preferences()
            self.save_preferences(self.preferences)

        return self.preferences

    def save_preferences(self, preferences: Dict[str, Any]) -> bool:
        """
        Save preferences to disk.

        Args:
            preferences: Preference dictionary to persist.

        Returns:
            True when saved successfully; otherwise False.
        """
        try:
            sanitized_preferences = self._sanitize_preferences(preferences)
            temp_file = self.preferences_file.with_suffix(".json.tmp")

            with temp_file.open("w", encoding="utf-8") as file_obj:
                json.dump(sanitized_preferences, file_obj, indent=4)

            temp_file.replace(self.preferences_file)
            self.preferences = sanitized_preferences
            AppLogger.log_message("debug", "Preferences saved.")
            return True
        except OSError as exc:
            AppLogger.log_message(
                "error",
                f"Failed to save preferences: {exc}",
            )
            return False

    def get(self, key: str, default: Any = None) -> Any:
        """
        Retrieve a preference value.

        Args:
            key: Preference key.
            default: Value returned when key is missing.

        Returns:
            Stored preference value or default.
        """
        return self.preferences.get(key, default)

    def set(self, key: str, value: Any) -> bool:
        """
        Set and immediately save a preference value.

        Args:
            key: Preference key.
            value: Preference value.

        Returns:
            True when saved successfully; otherwise False.
        """
        if self._is_credential_key(key):
            AppLogger.log_message(
                "warning",
                f"Blocked attempt to save credential preference: {key}",
            )
            return False

        self.preferences[key] = value
        return self.save_preferences(self.preferences)

    def backup_corrupt_preferences(self) -> Optional[Path]:
        """
        Back up a corrupt or unreadable preferences file.

        Returns:
            Backup path when created; otherwise None.
        """
        if not self.preferences_file.exists():
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.preferences_file.with_name(
            f"{self.preferences_file.stem}_corrupt_{timestamp}"
            f"{self.preferences_file.suffix}"
        )

        try:
            shutil.copy2(self.preferences_file, backup_path)
            AppLogger.log_message(
                "warning",
                f"Corrupt preferences backed up to {backup_path}",
            )
            return backup_path
        except OSError as exc:
            AppLogger.log_message(
                "error",
                f"Failed to back up corrupt preferences: {exc}",
            )
            return None

    def _default_preferences(self) -> Dict[str, Any]:
        """
        Build default preferences.

        Returns:
            Default preference dictionary.
        """
        return {
            "last_ip_range": "",
            "last_device_filter": "",
            "last_selected_scan_mode": "standard",
            "max_ping_workers": AppConfig.DEFAULT_MAX_PING_WORKERS,
            "max_inventory_workers": AppConfig.DEFAULT_MAX_INVENTORY_WORKERS,
            "last_export_directory": str(Path.cwd()),
            "window_geometry": AppConfig.DEFAULT_WINDOW_GEOMETRY,
            "column_widths": {},
            "selected_tab": "Dashboard",
            "checkbox_states": {},
            "credential_mode": "auto",
            "auto_export_enabled": False,
        }

    def _sanitize_preferences(
        self,
        preferences: Dict[str, Any],
    ) -> Dict[str, Any]:
        """
        Remove unsafe or unsupported preference values.

        Args:
            preferences: Raw preferences.

        Returns:
            Sanitized preference dictionary.
        """
        sanitized = {
            **self._default_preferences(),
            **preferences,
        }

        for key in list(sanitized):
            if self._is_credential_key(key):
                sanitized.pop(key, None)

        sanitized["max_ping_workers"] = self._clamp_int(
            sanitized.get("max_ping_workers"),
            AppConfig.MIN_PING_WORKERS,
            AppConfig.MAX_PING_WORKERS,
            AppConfig.DEFAULT_MAX_PING_WORKERS,
        )
        sanitized["max_inventory_workers"] = self._clamp_int(
            sanitized.get("max_inventory_workers"),
            AppConfig.MIN_INVENTORY_WORKERS,
            AppConfig.MAX_INVENTORY_WORKERS,
            AppConfig.DEFAULT_MAX_INVENTORY_WORKERS,
        )

        return sanitized

    def _clamp_int(
        self,
        value: Any,
        minimum: int,
        maximum: int,
        default: int,
    ) -> int:
        """
        Convert a value to an integer within a safe range.

        Args:
            value: Raw value.
            minimum: Minimum allowed integer.
            maximum: Maximum allowed integer.
            default: Fallback integer.

        Returns:
            Clamped integer.
        """
        try:
            return max(minimum, min(maximum, int(value)))
        except (TypeError, ValueError):
            return default

    def _is_credential_key(self, key: str) -> bool:
        """
        Detect keys that may contain credential material.

        Args:
            key: Preference key.

        Returns:
            True when the key should not be persisted.
        """
        lowered_key = str(key).lower()
        blocked_tokens = ("password", "secret", "token", "credential")
        return any(token in lowered_key for token in blocked_tokens)

# ------------------ Credential Management ------------------ #
class CredentialManager:
    """
    Manage privilege detection and optional PowerShell credential usage.

    Credentials are gathered with Get-Credential and cached only for the
    current app session as a DPAPI-encrypted CLIXML file. The cache is stored in
    a per-session temp file, removed during normal shutdown, and stale caches
    are cleaned on startup.
    """

    CACHE_PREFIX = ".inventory_scanner_credential_"
    CACHE_SUFFIX = ".xml"

    def __init__(self) -> None:
        """Initialize credential manager state."""
        self._is_admin_cached: Optional[bool] = None
        self._credential_lock = threading.RLock()
        self._cleanup_registered = False

        self._cleanup_stale_credential_caches()

        session_id = uuid.uuid4().hex
        cache_name = (
            f"{self.CACHE_PREFIX}{os.getpid()}_{session_id}"
            f"{self.CACHE_SUFFIX}"
        )
        self._credential_cache_path = Path(tempfile.gettempdir()) / cache_name

        self._register_cleanup_handlers()

        if not self.is_admin():
            AppLogger.log_message(
                "warning",
                "Application is not running with administrative privileges.",
            )

    def __del__(self) -> None:
        """Best-effort credential cleanup during object destruction."""
        try:
            self.clear_cached_credentials(log_when_missing=False)
        except Exception:
            pass

    def is_admin(self) -> bool:
        """Return True when running with administrative privileges."""
        if self._is_admin_cached is not None:
            return self._is_admin_cached

        try:
            if os.name == "nt":
                self._is_admin_cached = bool(ctypes.windll.shell32.IsUserAnAdmin())
            else:
                self._is_admin_cached = os.geteuid() == 0
        except Exception as exc:
            AppLogger.log_message("error", f"Failed to determine admin status: {exc}")
            self._is_admin_cached = False

        return self._is_admin_cached

    def credentials_required(self) -> bool:
        """Return True when delegated credentials may be needed."""
        return not self.is_admin()

    def has_cached_credentials(self) -> bool:
        """Return True when session credential cache exists."""
        return self._credential_cache_path.exists()

    def prompt_for_credentials(self) -> bool:
        """
        Prompt for credentials using PowerShell Get-Credential.

        Returns:
            True when credentials were captured successfully; otherwise False.
        """
        if self.is_admin():
            AppLogger.log_message(
                "info",
                "Credential prompt skipped because app is running as admin.",
            )
            return True

        with self._credential_lock:
            try:
                self.clear_cached_credentials(log_when_missing=False)
                self._credential_cache_path.parent.mkdir(
                    parents=True,
                    exist_ok=True,
                )

                cache_path = str(self._credential_cache_path)
                escaped_cache_path = cache_path.replace("'", "''")

                command = (
                    "$credential = Get-Credential "
                    "-Message 'Enter delegated credentials for remote inventory'; "
                    "if ($null -eq $credential) { exit 1 }; "
                    f"$credential | Export-Clixml -Path '{escaped_cache_path}'; "
                    "exit 0"
                )

                completed = subprocess.run(
                    [
                        AppConfig.POWERSHELL_EXECUTABLE,
                        "-NoProfile",
                        "-Sta",
                        "-ExecutionPolicy",
                        "Bypass",
                        "-Command",
                        command,
                    ],
                    capture_output=True,
                    text=True,
                    check=False,
                )

                if completed.returncode == 0 and self.has_cached_credentials():
                    self._harden_credential_cache_file()
                    AppLogger.log_message(
                        "info",
                        "PowerShell credentials cached for this session.",
                    )
                    return True

                error_text = completed.stderr.strip() or completed.stdout.strip()
                AppLogger.log_message(
                    "warning",
                    f"Credential prompt cancelled or failed: {error_text}",
                )
                self.clear_cached_credentials(log_when_missing=False)
                return False

            except Exception as exc:
                AppLogger.log_message(
                    "error",
                    f"Failed to prompt for PowerShell credentials: {exc}",
                )
                self.clear_cached_credentials(log_when_missing=False)
                return False

    def build_credential_argument(self) -> str:
        """
        Build PowerShell credential argument for Invoke-Command.

        Returns:
            PowerShell credential argument or an empty string.
        """
        if self.is_admin() or not self.has_cached_credentials():
            return ""

        cache_path = str(self._credential_cache_path)
        escaped_cache_path = cache_path.replace("'", "''")
        return f"-Credential (Import-Clixml -Path '{escaped_cache_path}')"

    def apply_credentials_to_command(self, command: str) -> str:
        """
        Validate and return PowerShell command text.

        Args:
            command: Command text.

        Returns:
            Validated command text.
        """
        if not isinstance(command, str) or not command.strip():
            AppLogger.log_message(
                "error",
                "Invalid command passed to credential application.",
            )
            return command

        return command

    def clear_cached_credentials(self, log_when_missing: bool = True) -> None:
        """
        Delete the session credential cache file.

        Args:
            log_when_missing: Log a debug message when no cache exists.
        """
        with self._credential_lock:
            try:
                if self._credential_cache_path.exists():
                    self._credential_cache_path.unlink()
                    AppLogger.log_message("info", "Session credential cache deleted.")
                    return

                if log_when_missing:
                    AppLogger.log_message(
                        "debug",
                        "No session credential cache found to delete.",
                    )
            except OSError as exc:
                AppLogger.log_message(
                    "error",
                    f"Failed to delete session credential cache: {exc}",
                )

    def _register_cleanup_handlers(self) -> None:
        """Register best-effort cleanup handlers for normal process shutdown."""
        if self._cleanup_registered:
            return

        atexit.register(self.clear_cached_credentials, False)

        for signal_name in ("SIGINT", "SIGTERM"):
            if not hasattr(signal, signal_name):
                continue

            current_signal = getattr(signal, signal_name)
            previous_handler = signal.getsignal(current_signal)

            def handler(
                signum: int,
                frame: Any,
                prior_handler: Any = previous_handler,
            ) -> None:
                self._handle_shutdown_signal(signum, frame, prior_handler)

            try:
                signal.signal(current_signal, handler)
            except (ValueError, OSError) as exc:
                AppLogger.log_message(
                    "debug",
                    f"Unable to register credential cleanup signal handler: {exc}",
                )

        self._cleanup_registered = True

    def _handle_shutdown_signal(
        self,
        signum: int,
        frame: Any,
        prior_handler: Any,
    ) -> None:
        """
        Clear credentials before forwarding a shutdown signal.

        Args:
            signum: Signal number.
            frame: Current frame.
            prior_handler: Previously registered handler.
        """
        self.clear_cached_credentials(log_when_missing=False)

        if callable(prior_handler):
            prior_handler(signum, frame)
            return

        if prior_handler == signal.SIG_DFL:
            raise SystemExit(128 + signum)

    def _cleanup_stale_credential_caches(self) -> None:
        """
        Delete stale credential cache files from previous abnormal exits.

        A hard crash, power loss, or task kill can bypass normal cleanup. This
        startup sweep removes leftover scanner credential caches from the temp
        directory before a new session begins.
        """
        temp_dir = Path(tempfile.gettempdir())
        pattern = f"{self.CACHE_PREFIX}*{self.CACHE_SUFFIX}"

        for cache_file in temp_dir.glob(pattern):
            try:
                cache_file.unlink()
                AppLogger.log_message(
                    "warning",
                    f"Removed stale credential cache: {cache_file.name}",
                )
            except OSError as exc:
                AppLogger.log_message(
                    "error",
                    f"Failed to remove stale credential cache {cache_file}: {exc}",
                )

    def _harden_credential_cache_file(self) -> None:
        """
        Restrict and hide the credential cache file where supported.

        Export-Clixml is DPAPI-protected for the current Windows user, but this
        adds an extra filesystem-level hardening step.
        """
        try:
            if not self._credential_cache_path.exists():
                return

            os.chmod(
                self._credential_cache_path,
                stat.S_IRUSR | stat.S_IWUSR,
            )

            if os.name == "nt":
                subprocess.run(
                    [
                        "attrib",
                        "+H",
                        str(self._credential_cache_path),
                    ],
                    capture_output=True,
                    text=True,
                    check=False,
                )

        except Exception as exc:
            AppLogger.log_message(
                "warning",
                f"Credential cache hardening failed: {exc}",
            )

# ------------------ Input Validation ------------------ #
class ScanInputValidator:
    """
    Validate and normalize user-provided IP ranges and hostname filters.

    All methods are designed to be safe, rejecting malformed or potentially
    dangerous input before it reaches subprocess or PowerShell execution.
    """

    @staticmethod
    def validate_ip_range(value: str) -> Tuple[bool, str]:
        """
        Validate an IP range string.

        Args:
            value: Raw user input.

        Returns:
            Tuple of (is_valid, message).
        """
        if not value or not value.strip():
            return False, "IP range cannot be empty."

        value = value.strip()

        try:
            if "-" in value:
                start_ip, end_ip = [v.strip() for v in value.split("-", 1)]
                ipaddress.ip_address(start_ip)
                ipaddress.ip_address(end_ip)
                return True, "Valid IP range."

            if "/" in value:
                ipaddress.ip_network(value, strict=False)
                return True, "Valid CIDR range."

            ipaddress.ip_address(value)
            return True, "Valid single IP."

        except ValueError as exc:
            return False, f"Invalid IP range: {exc}"

    @staticmethod
    def expand_ip_range(value: str) -> List[str]:
        """
        Expand an IP range into a list of IP strings.

        Args:
            value: Validated IP range string.

        Returns:
            List of IP addresses.
        """
        try:
            if "-" in value:
                start_ip, end_ip = [v.strip() for v in value.split("-", 1)]
                start = int(ipaddress.ip_address(start_ip))
                end = int(ipaddress.ip_address(end_ip))

                if start > end:
                    start, end = end, start

                return [
                    str(ipaddress.ip_address(ip))
                    for ip in range(start, end + 1)
                ]

            if "/" in value:
                network = ipaddress.ip_network(value, strict=False)
                return [str(ip) for ip in network.hosts()]

            return [str(ipaddress.ip_address(value))]

        except Exception as exc:
            AppLogger.log_message(
                "error",
                f"Failed to expand IP range: {exc}"
            )
            return [], {}

    @staticmethod
    def parse_device_filters(value: str) -> List[str]:
        """
        Parse and normalize device filter string.

        Args:
            value: Raw filter string.

        Returns:
            List of normalized filter prefixes.
        """
        if not value:
            return [], {}

        separators = [",", ";", "|"]
        normalized = value

        for sep in separators:
            normalized = normalized.replace(sep, " ")

        tokens = [
            token.strip().lower()
            for token in normalized.split()
            if token.strip()
        ]

        return list(dict.fromkeys(tokens))  # remove duplicates

    @staticmethod
    def hostname_matches_filters(
        hostname: str,
        filters: List[str]
    ) -> bool:
        """
        Check if a hostname matches any filter prefix.

        Args:
            hostname: Hostname to test.
            filters: List of filter prefixes.

        Returns:
            True if match or no filters provided.
        """
        if not hostname:
            return False if filters else True

        if not filters:
            return True

        hostname_lower = hostname.lower()
        return any(
            hostname_lower.startswith(prefix)
            for prefix in filters
        )

    @staticmethod
    def sanitize_hostname(value: str) -> str:
        """
        Sanitize hostname input to safe format.

        Args:
            value: Raw hostname.

        Returns:
            Sanitized hostname.
        """
        if not value:
            return ""

        safe = "".join(
            char for char in value
            if char.isalnum() or char in ("-", ".", "_")
        )

        return safe[:AppConfig.SAFE_HOSTNAME_MAX_LENGTH]

    @staticmethod
    def sanitize_powershell_argument(value: str) -> str:
        """
        Sanitize string for safe PowerShell argument usage.

        Args:
            value: Raw string.

        Returns:
            Escaped string safe for PowerShell.
        """
        if not value:
            return ""

        # Escape dangerous characters
        replacements = {
            "`": "``",
            '"': '`"',
            "$": "`$",
        }

        sanitized = value
        for key, replacement in replacements.items():
            sanitized = sanitized.replace(key, replacement)

        return sanitized.strip()

# ------------------ Data Model: ScanResult ------------------ #
@dataclass
class ScanResult:
    """
    Structured output from threaded scan tasks.

    ScanResult objects should be passed through thread-safe queues so the GUI
    can update safely from the main thread.
    """

    scan_id: Optional[int] = None
    task_name: str = ""
    target: str = ""
    status: str = AppConfig.SCAN_STATUS_SUCCESS
    message: str = ""
    data: Dict[str, Any] = field(default_factory=dict)
    error: str = ""
    started_at: str = ""
    completed_at: str = ""
    duration_seconds: float = 0.0

    def to_display_text(self) -> str:
        """
        Convert the result to a human-readable live-feed message.

        Returns:
            Display-ready status text.
        """
        timestamp = time.strftime(AppConfig.LIVE_FEED_TIMESTAMP_FORMAT)
        target_text = f" {self.target}" if self.target else ""
        message_text = self.message or self.error or "No details provided."

        return (
            f"[{timestamp}] {self.task_name}{target_text}: "
            f"{self.status.upper()} - {message_text}"
        )

    def to_json_dict(self) -> Dict[str, Any]:
        """
        Convert the result to a JSON-serializable dictionary.

        Returns:
            Dictionary representation of the scan result.
        """
        return asdict(self)

# ------------------ Data Model: ScanProgress ------------------ #
@dataclass
class ScanProgress:
    """
    Structured progress update model for UI consumption.

    This object is passed through the result queue to inform the UI about
    current scan phase, progress metrics, and live dashboard counters.
    """

    phase: str = "validating"
    total: int = 0
    completed: int = 0
    success_count: int = 0
    warning_count: int = 0
    error_count: int = 0
    message: str = ""
    data: Dict[str, Any] = field(default_factory=dict)

    def to_display_text(self) -> str:
        """
        Convert progress into a human-readable message.

        Returns:
            Formatted progress string.
        """
        timestamp = time.strftime(AppConfig.LIVE_FEED_TIMESTAMP_FORMAT)

        return (
            f"[{timestamp}] PHASE: {self.phase.upper()} | "
            f"{self.completed}/{self.total} | "
            f"Success: {self.success_count} | "
            f"Warnings: {self.warning_count} | "
            f"Errors: {self.error_count} | "
            f"{self.message}"
        )

    def to_json_dict(self) -> Dict[str, Any]:
        """
        Convert progress object into a JSON-serializable dictionary.

        Returns:
            Dictionary representation of progress.
        """
        return asdict(self)

# ------------------ Data Model: ComputerRecord ------------------ #
@dataclass
class ComputerRecord:
    """
    Structured representation of a scanned computer.

    This model is used across scanning, inventory, and database layers.
    """

    computer_id: Optional[int] = None
    hostname: str = ""
    ip_address: str = ""
    fqdn: str = ""
    domain: str = ""
    pingable: bool = False
    matched_filter: bool = False
    inventory_status: str = AppConfig.INVENTORY_STATUS_PENDING
    inventory_error: str = ""
    operating_system: str = ""
    manufacturer: str = ""
    model: str = ""
    serial_number: str = ""
    last_seen: str = ""
    scan_id: Optional[int] = None

    def to_dict(self) -> Dict[str, Any]:
        """
        Convert the record to a dictionary.

        Returns:
            Dictionary representation of the computer record.
        """
        return asdict(self)

    def mark_inventory_success(self) -> None:
        """
        Mark the computer as successfully inventoried.
        """
        self.inventory_status = AppConfig.INVENTORY_STATUS_SUCCESS
        self.inventory_error = ""

    def mark_inventory_failure(self, error: str) -> None:
        """
        Mark the computer as failed during inventory.

        Args:
            error: Error message describing the failure.
        """
        self.inventory_status = AppConfig.INVENTORY_STATUS_FAILED
        self.inventory_error = error

    def mark_inventory_partial(self, error: str = "") -> None:
        """
        Mark the computer as partially inventoried.

        Args:
            error: Optional warning or partial failure message.
        """
        self.inventory_status = AppConfig.INVENTORY_STATUS_PARTIAL
        self.inventory_error = error

    def update_last_seen(self) -> None:
        """
        Update the last_seen timestamp to current time.
        """
        self.last_seen = datetime.now().strftime(
            AppConfig.STORAGE_TIMESTAMP_FORMAT
        )
    
# ------------------ Data Model: SoftwareRecord ------------------ #
@dataclass
class SoftwareRecord:
    """
    Structured representation of installed software on a computer.

    This model is used for inventory parsing, normalization, and database storage.
    """

    software_id: Optional[int] = None
    computer_id: Optional[int] = None
    display_name: str = ""
    display_version: str = ""
    publisher: str = ""
    install_date: str = ""
    uninstall_string: str = ""
    install_location: str = ""
    estimated_size: str = ""
    registry_key: str = ""
    registry_hive: str = ""
    architecture: str = ""
    raw_json: str = ""
    scan_id: Optional[int] = None

    def to_dict(self) -> Dict[str, Any]:
        """
        Convert the software record to a dictionary.

        Returns:
            Dictionary representation of the software record.
        """
        return asdict(self)

    def is_valid(self) -> bool:
        """
        Validate that the software record has required fields.

        Returns:
            True if valid, otherwise False.
        """
        return bool(self.display_name and self.display_name.strip())

    def normalized_key(self) -> Tuple[str, str]:
        """
        Generate a normalized identity key for grouping.

        Returns:
            Tuple of (display_name, display_version).
        """
        return (
            (self.display_name or "").strip().lower(),
            (self.display_version or "").strip().lower(),
        )

    def to_display_name(self) -> str:
        """
        Generate a user-friendly display string.

        Returns:
            Combined display name and version.
        """
        if self.display_version:
            return f"{self.display_name} ({self.display_version})"
        return self.display_name

# ------------------ Data Model: ExportResult ------------------ #
@dataclass
class ExportResult:
    """
    Structured result from a background export operation.

    ExportResult objects are created on the export worker thread and applied to
    the GUI only through the Tkinter main thread.
    """

    success: bool = False
    output_path: str = ""
    message: str = ""
    error: str = ""
    duration_seconds: float = 0.0

    def to_display_text(self) -> str:
        """
        Convert export result to readable status text.

        Returns:
            Display-ready export message.
        """
        if self.success:
            return f"Export complete: {self.output_path}"

        return self.error or self.message or "Export failed."

# ------------------ Database Management ------------------ #
class DatabaseManager:
    """
    SQLite persistence layer for scan results, computers, and software inventory.
    """

    def __init__(self, db_file: str = AppConfig.DATABASE_FILE):
        """Initialize database manager."""
        self.db_path = Path(db_file)
        self.connection: Optional[sqlite3.Connection] = None
        self._connect()
        self.initialize_database()

    def _connect(self) -> None:
        """Establish SQLite connection."""
        try:
            self.connection = sqlite3.connect(
                self.db_path,
                timeout=AppConfig.SQLITE_TIMEOUT,
                check_same_thread=False,
            )
            self.connection.row_factory = sqlite3.Row
            AppLogger.log_message("info", "Database connected.")
        except sqlite3.Error as exc:
            AppLogger.log_message("critical", f"Database connection failed: {exc}")

    def initialize_database(self) -> None:
        """Create required tables and indexes if they do not exist."""
        if not self.connection:
            AppLogger.log_message("critical", "Database initialization skipped.")
            return

        try:
            cursor = self.connection.cursor()
            cursor.executescript("""
            CREATE TABLE IF NOT EXISTS scans (
                scan_id INTEGER PRIMARY KEY,
                started_at TEXT,
                completed_at TEXT,
                ip_range TEXT,
                filters TEXT,
                status TEXT,
                pingable_count INTEGER,
                matched_count INTEGER,
                successful_inventory_count INTEGER,
                failed_inventory_count INTEGER
            );

            CREATE TABLE IF NOT EXISTS computers (
                computer_id INTEGER PRIMARY KEY,
                scan_id INTEGER,
                hostname TEXT,
                ip_address TEXT,
                ip_sort_key TEXT,
                fqdn TEXT,
                domain TEXT,
                pingable INTEGER,
                matched_filter INTEGER,
                inventory_status TEXT,
                inventory_error TEXT,
                operating_system TEXT,
                manufacturer TEXT,
                model TEXT,
                serial_number TEXT,
                last_seen TEXT
            );

            CREATE TABLE IF NOT EXISTS software (
                software_id INTEGER PRIMARY KEY,
                scan_id INTEGER,
                computer_id INTEGER,
                display_name TEXT,
                display_version TEXT,
                publisher TEXT,
                display_name_norm TEXT,
                display_version_norm TEXT,
                publisher_norm TEXT,
                install_date TEXT,
                uninstall_string TEXT,
                install_location TEXT,
                estimated_size TEXT,
                registry_key TEXT,
                registry_hive TEXT,
                architecture TEXT,
                raw_json TEXT
            );

            CREATE TABLE IF NOT EXISTS scan_events (
                event_id INTEGER PRIMARY KEY,
                scan_id INTEGER,
                timestamp TEXT,
                level TEXT,
                message TEXT,
                target TEXT,
                task_name TEXT,
                duration_seconds REAL
            );

            CREATE INDEX IF NOT EXISTS idx_computers_scan_id
                ON computers(scan_id);

            CREATE INDEX IF NOT EXISTS idx_computers_scan_status
                ON computers(scan_id, inventory_status, computer_id);

            CREATE INDEX IF NOT EXISTS idx_computers_scan_ip_sort
                ON computers(scan_id, ip_sort_key, hostname);

            CREATE INDEX IF NOT EXISTS idx_software_scan_identity
                ON software(scan_id, display_name, display_version, publisher);

            CREATE INDEX IF NOT EXISTS idx_software_scan_computer
                ON software(scan_id, computer_id);

            CREATE UNIQUE INDEX IF NOT EXISTS uq_software_computer_identity_norm
                ON software(
                    computer_id,
                    display_name_norm,
                    display_version_norm,
                    publisher_norm
                );

            CREATE INDEX IF NOT EXISTS idx_software_lookup_exact_norm
                ON software(
                    scan_id,
                    computer_id,
                    display_name_norm,
                    display_version_norm,
                    publisher_norm
                );

            CREATE INDEX IF NOT EXISTS idx_software_lookup_group_norm
                ON software(
                    scan_id,
                    computer_id,
                    display_name_norm,
                    publisher_norm
                );
            """)
            self.connection.commit()
            cursor.execute("PRAGMA table_info(computers)")
            computer_columns = {row[1] for row in cursor.fetchall()}

            if "ip_sort_key" not in computer_columns:
                cursor.execute("ALTER TABLE computers ADD COLUMN ip_sort_key TEXT")

            cursor.execute("PRAGMA table_info(software)")
            software_columns = {row[1] for row in cursor.fetchall()}
            if "display_name_norm" not in software_columns:
                cursor.execute(
                    "ALTER TABLE software ADD COLUMN display_name_norm TEXT"
                )
            if "display_version_norm" not in software_columns:
                cursor.execute(
                    "ALTER TABLE software ADD COLUMN display_version_norm TEXT"
                )
            if "publisher_norm" not in software_columns:
                cursor.execute(
                    "ALTER TABLE software ADD COLUMN publisher_norm TEXT"
                )

            cursor.execute("""
                UPDATE software
                SET
                    display_name_norm = LOWER(TRIM(COALESCE(display_name, ''))),
                    display_version_norm = LOWER(TRIM(COALESCE(display_version, ''))),
                    publisher_norm = LOWER(TRIM(COALESCE(publisher, '')))
                WHERE display_name_norm IS NULL
                    OR display_version_norm IS NULL
                    OR publisher_norm IS NULL
            """)

            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_computers_scan_ip_sort
                    ON computers(scan_id, ip_sort_key, hostname)
            """)
            cursor.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS uq_software_computer_identity_norm
                    ON software(
                        computer_id,
                        display_name_norm,
                        display_version_norm,
                        publisher_norm
                    )
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_software_lookup_exact_norm
                    ON software(
                        scan_id,
                        computer_id,
                        display_name_norm,
                        display_version_norm,
                        publisher_norm
                    )
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_software_lookup_group_norm
                    ON software(
                        scan_id,
                        computer_id,
                        display_name_norm,
                        publisher_norm
                    )
            """)
            self.connection.commit()
            AppLogger.log_message("info", "Database initialized.")
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Database initialization failed: {exc}")

    def begin_transaction(self) -> None:
        """Begin a caller-managed transaction if one is not already active."""
        if not self.connection:
            return

        if self.connection.in_transaction:
            return

        self.connection.execute("BEGIN")

    def commit_transaction(self) -> None:
        """Commit the active caller-managed transaction."""
        if self.connection and self.connection.in_transaction:
            self.connection.commit()

    def rollback_transaction(self) -> None:
        """Rollback the active caller-managed transaction."""
        if self.connection and self.connection.in_transaction:
            self.connection.rollback()

    @contextmanager
    def transaction(self):
        """Context manager for atomic write batches."""
        self.begin_transaction()
        try:
            yield
            self.commit_transaction()
        except Exception:
            self.rollback_transaction()
            raise

    def create_scan(self, ip_range: str, filters: str) -> Optional[int]:
        """Insert new scan record."""
        try:
            cursor = self.connection.cursor()
            started_at = datetime.now().strftime(AppConfig.STORAGE_TIMESTAMP_FORMAT)

            cursor.execute("""
                INSERT INTO scans (started_at, ip_range, filters, status)
                VALUES (?, ?, ?, ?)
            """, (started_at, ip_range, filters, "running"))

            self.connection.commit()
            return cursor.lastrowid
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Create scan failed: {exc}")
            return None

    def complete_scan(self, scan_id: int, summary: Dict[str, Any]) -> None:
        """Update scan completion data."""
        try:
            cursor = self.connection.cursor()
            completed_at = datetime.now().strftime(AppConfig.STORAGE_TIMESTAMP_FORMAT)

            cursor.execute("""
                UPDATE scans
                SET completed_at=?, status=?, pingable_count=?, matched_count=?,
                    successful_inventory_count=?, failed_inventory_count=?
                WHERE scan_id=?
            """, (
                completed_at,
                summary.get("status"),
                summary.get("pingable_count"),
                summary.get("matched_count"),
                summary.get("successful_inventory_count"),
                summary.get("failed_inventory_count"),
                scan_id,
            ))
            self.connection.commit()
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Complete scan failed: {exc}")

    def get_software_occurrences(self, scan_id: int) -> List[sqlite3.Row]:
        """
        Return software occurrence counts.

        Occurrence count is version-specific for dropdown rows.
        Grouped version occurrence count treats different versions of the same
        software as one software title for publisher-level rollups.
        """
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                SELECT
                    s.display_name,
                    s.display_version,
                    s.publisher,
                    COUNT(DISTINCT s.computer_id) AS occurrence_count,
                    (
                        SELECT COUNT(DISTINCT grouped_s.computer_id)
                        FROM software grouped_s
                        WHERE grouped_s.scan_id = s.scan_id
                            AND LOWER(TRIM(grouped_s.display_name))
                                = LOWER(TRIM(s.display_name))
                            AND LOWER(TRIM(grouped_s.publisher))
                                = LOWER(TRIM(s.publisher))
                    ) AS grouped_version_occurrence_count
                FROM software s
                WHERE s.scan_id=?
                    AND s.display_name IS NOT NULL
                    AND TRIM(s.display_name) <> ''
                GROUP BY
                    LOWER(TRIM(s.display_name)),
                    LOWER(TRIM(s.display_version)),
                    LOWER(TRIM(s.publisher))
                ORDER BY
                    LOWER(TRIM(s.display_name)),
                    LOWER(TRIM(s.display_version)),
                    LOWER(TRIM(s.publisher))
            """, (scan_id,))
            return cursor.fetchall()
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Software occurrence query failed: {exc}")
            return [], {}


    def get_computer_breakout_for_software(
        self,
        scan_id: int,
        display_name: str,
        display_version: str,
        publisher: str = "",
        group_versions: bool = False,
    ) -> List[Dict[str, Any]]:
        """
        Return all successfully inventoried computers with an Installed flag.

        This version avoids a correlated EXISTS check per computer by first building
        selected-computer and software-count CTEs, then joining them once.
        """
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                WITH selected_computers AS (
                    SELECT DISTINCT selected_s.computer_id
                    FROM software selected_s
                    WHERE selected_s.scan_id = ?
                        AND LOWER(TRIM(selected_s.display_name)) = LOWER(TRIM(?))
                        AND (
                            ? = ''
                            OR LOWER(TRIM(selected_s.publisher)) = LOWER(TRIM(?))
                        )
                        AND (
                            ? = 1
                            OR LOWER(TRIM(selected_s.display_version)) = LOWER(TRIM(?))
                        )
                ),
                software_counts AS (
                    SELECT
                        count_s.computer_id,
                        COUNT(DISTINCT count_s.software_id) AS software_count
                    FROM software count_s
                    WHERE count_s.scan_id = ?
                    GROUP BY count_s.computer_id
                )
                SELECT
                    c.*,
                    COALESCE(sc.software_count, 0) AS software_count,
                    CASE
                        WHEN selected_computers.computer_id IS NULL THEN 'No'
                        ELSE 'Yes'
                    END AS installed
                FROM computers c
                LEFT JOIN software_counts sc
                    ON sc.computer_id = c.computer_id
                LEFT JOIN selected_computers
                    ON selected_computers.computer_id = c.computer_id
                WHERE c.scan_id = ?
                    AND LOWER(TRIM(c.inventory_status)) IN (?, ?)
                ORDER BY
                    CASE
                        WHEN selected_computers.computer_id IS NULL THEN 1
                        ELSE 0
                    END,
                    c.ip_sort_key,
                    LOWER(TRIM(c.hostname))
            """, (
                scan_id,
                display_name,
                publisher,
                publisher,
                int(bool(group_versions)),
                display_version,
                scan_id,
                scan_id,
                AppConfig.INVENTORY_STATUS_SUCCESS,
                AppConfig.INVENTORY_STATUS_PARTIAL,
            ))

            rows = [dict(row) for row in cursor.fetchall()]
            if any(not row.get("ip_sort_key") for row in rows):
                rows.sort(
                    key=lambda row: (
                        0 if row.get("installed") == "Yes" else 1,
                        self._ip_sort_key(row.get("ip_address", "")),
                        str(row.get("hostname", "")).lower(),
                    )
                )
            return rows

        except sqlite3.Error as exc:
            AppLogger.log_message(
                "error",
                f"Computer breakout query failed: {exc}",
            )
            return [], {}

    def insert_computer(
        self,
        record: ComputerRecord,
        commit: bool = True,
    ) -> Optional[int]:
        """Insert computer record."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                INSERT INTO computers (
                    scan_id, hostname, ip_address, ip_sort_key, fqdn, domain,
                    pingable, matched_filter, inventory_status,
                    inventory_error, operating_system, manufacturer,
                    model, serial_number, last_seen
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                record.scan_id,
                record.hostname,
                record.ip_address,
                self._ip_sort_db_key(record.ip_address),
                record.fqdn,
                record.domain,
                int(record.pingable),
                int(record.matched_filter),
                record.inventory_status,
                record.inventory_error,
                record.operating_system,
                record.manufacturer,
                record.model,
                record.serial_number,
                record.last_seen,
            ))

            self.connection.commit()
            return cursor.lastrowid
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Insert computer failed: {exc}")
            return None

    def update_computer_inventory_status(
        self,
        computer_id: int,
        status: str,
        error: str = "",
        commit: bool = True,
    ) -> None:
        """Update inventory status for a computer."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                UPDATE computers
                SET inventory_status=?, inventory_error=?
                WHERE computer_id=?
            """, (status, error, computer_id))
            if commit:
                self.connection.commit()
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Update computer failed: {exc}")

    def insert_software_records(
        self,
        computer_id: int,
        software_records: List[SoftwareRecord],
        commit: bool = True,
    ) -> None:
        """
        Batch insert software records.

        Records are de-duplicated per computer by display name, display version,
        and publisher to prevent duplicate registry entries from inflating
        occurrence percentages.
        """
        try:
            cursor = self.connection.cursor()
            rows = []

            for record in software_records:
                if not record.is_valid():
                    continue

                display_name_norm = self._normalize_identity(record.display_name)
                display_version_norm = self._normalize_identity(
                    record.display_version
                )
                publisher_norm = self._normalize_identity(record.publisher)
                rows.append((
                    record.scan_id,
                    computer_id,
                    record.display_name,
                    record.display_version,
                    record.publisher,
                    display_name_norm,
                    display_version_norm,
                    publisher_norm,
                    record.install_date,
                    record.uninstall_string,
                    record.install_location,
                    record.estimated_size,
                    record.registry_key,
                    record.registry_hive,
                    record.architecture,
                    record.raw_json,
                ))

            cursor.executemany("""
                INSERT OR IGNORE INTO software (
                    scan_id, computer_id, display_name, display_version,
                    publisher, display_name_norm, display_version_norm,
                    publisher_norm, install_date, uninstall_string,
                    install_location, estimated_size, registry_key,
                    registry_hive, architecture, raw_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, rows)

            self.connection.commit()
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Insert software failed: {exc}")

    def get_all_computers_with_counts(self, scan_id: int) -> List[Dict[str, Any]]:
        """
        Return all computers with software counts.

        SQL ordering uses precomputed indexed keys; Python fallback is only
        used for legacy rows missing the key.
        """
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                SELECT c.*, COUNT(DISTINCT s.software_id) AS software_count
                FROM computers c
                LEFT JOIN software s ON c.computer_id = s.computer_id
                WHERE c.scan_id = ?
                GROUP BY c.computer_id
                ORDER BY c.ip_sort_key, LOWER(TRIM(c.hostname))
            """, (scan_id,))

            rows = [dict(row) for row in cursor.fetchall()]
            if any(not row.get("ip_sort_key") for row in rows):
                rows.sort(
                    key=lambda row: (
                        self._ip_sort_key(row.get("ip_address", "")),
                        str(row.get("hostname", "")).lower(),
                    )
                )
            return rows
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Computer count query failed: {exc}")
            return [], {}

    def get_computers_for_software(
        self,
        scan_id: int,
        display_name: str,
        display_version: str,
        publisher: str = "",
    ) -> List[Dict[str, Any]]:
        """Return computers containing selected software with software counts."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                SELECT c.*, COUNT(DISTINCT all_s.software_id) AS software_count
                FROM computers c
                JOIN software selected_s
                    ON c.computer_id = selected_s.computer_id
                LEFT JOIN software all_s
                    ON c.computer_id = all_s.computer_id
                WHERE selected_s.scan_id=?
                    AND LOWER(TRIM(selected_s.display_name))=LOWER(TRIM(?))
                    AND LOWER(TRIM(selected_s.display_version))=LOWER(TRIM(?))
                    AND (
                        ? = ''
                        OR LOWER(TRIM(selected_s.publisher))=LOWER(TRIM(?))
                    )
                GROUP BY c.computer_id
                ORDER BY c.ip_sort_key, LOWER(TRIM(c.hostname))
            """, (
                scan_id,
                display_name,
                display_version,
                publisher,
                publisher,
            ))

            rows = [dict(row) for row in cursor.fetchall()]
            if any(not row.get("ip_sort_key") for row in rows):
                rows.sort(
                    key=lambda row: (
                        self._ip_sort_key(row.get("ip_address", "")),
                        str(row.get("hostname", "")).lower(),
                    )
                )
            return rows
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Computers for software query failed: {exc}")
            return [], {}

    def get_software_for_computer(
        self,
        scan_id: int,
        computer_id: int,
    ) -> List[sqlite3.Row]:
        """Return software for a specific computer."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                SELECT *
                FROM software
                WHERE scan_id=? AND computer_id=?
                ORDER BY LOWER(display_name), LOWER(display_version), LOWER(publisher)
            """, (scan_id, computer_id))
            return cursor.fetchall()
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Software for computer query failed: {exc}")
            return [], {}

    def get_computer_properties(self, computer_id: int) -> Dict[str, Any]:
        """Return full computer properties."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT * FROM computers WHERE computer_id=?", (computer_id,))
            row = cursor.fetchone()
            return dict(row) if row else {}
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Computer property query failed: {exc}")
            return {}

    def get_software_properties(self, software_id: int) -> Dict[str, Any]:
        """Return full software properties."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT * FROM software WHERE software_id=?", (software_id,))
            row = cursor.fetchone()
            return dict(row) if row else {}
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Software property query failed: {exc}")
            return {}

    def get_scan_summary(self, scan_id: int) -> Dict[str, Any]:
        """Return scan summary."""
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT * FROM scans WHERE scan_id=?", (scan_id,))
            row = cursor.fetchone()
            return dict(row) if row else {}
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Scan summary query failed: {exc}")
            return {}

    def vacuum_database(self) -> None:
        """Optimize database size."""
        try:
            self.connection.execute("VACUUM")
            AppLogger.log_message("info", "Database vacuum completed.")
        except sqlite3.Error as exc:
            AppLogger.log_message("error", f"Vacuum failed: {exc}")

    def _normalize_identity(self, value: Any) -> str:
        """
        Normalize identity values for duplicate detection.

        Args:
            value: Raw identity value.

        Returns:
            Lowercase, whitespace-normalized string.
        """
        return " ".join(str(value or "").strip().lower().split())

    def _ip_sort_key(self, value: Any) -> Tuple[int, int, str]:
        """
        Build a stable IP-aware sort key.

        Args:
            value: Raw IP address.

        Returns:
            Sort tuple that places valid IPs before invalid or blank values.
        """
        text = str(value or "").strip()

        try:
            return 0, int(ipaddress.ip_address(text)), text
        except ValueError:
            return 1, 0, text.lower()

    def _ip_sort_db_key(self, value: Any) -> str:
        """
        Build indexed SQL sort key matching _ip_sort_key ordering semantics.

        Args:
            value: Raw IP address.

        Returns:
            Prefixed sortable key for SQLite ORDER BY operations.
        """
        text = str(value or "").strip()

        try:
            numeric_value = int(ipaddress.ip_address(text))
            return f"0:{numeric_value:039d}:{text}"
        except ValueError:
            return f"1:{text.lower()}"

# ------------------ IP Range Scanner ------------------ #
class IpRangeScanner:
    """
    Discover pingable IP addresses using concurrent subprocess ping checks.

    This scanner performs no UI updates directly. Results are returned through
    the provided queue as ScanResult objects.
    """

    def __init__(self, timeout: float = AppConfig.DEFAULT_PING_TIMEOUT):
        """
        Initialize the IP scanner.

        Args:
            timeout: Ping subprocess timeout in seconds.
        """
        self.timeout = max(0.1, float(timeout))

    def scan(
        self,
        ip_addresses: Sequence[str],
        max_workers: int,
        stop_event: threading.Event,
        result_queue: queue.Queue,
    ) -> List[str]:
        """
        Ping IP addresses concurrently.

        Args:
            ip_addresses: Expanded IP address list.
            max_workers: Maximum ping worker threads.
            stop_event: Cancellation event.
            result_queue: Queue receiving ScanResult objects.

        Returns:
            List of reachable IP addresses.
        """
        reachable_ips: List[str] = []
        worker_count = max(
            AppConfig.MIN_PING_WORKERS,
            min(AppConfig.MAX_PING_WORKERS, int(max_workers)),
        )

        with ThreadPoolExecutor(
            max_workers=worker_count,
            thread_name_prefix="PingWorker",
        ) as executor:
            future_map = {
                executor.submit(self.ping_host, ip_address): ip_address
                for ip_address in ip_addresses
                if not stop_event.is_set()
            }

            for future in as_completed(future_map):
                if stop_event.is_set():
                    break

                ip_address = future_map[future]

                try:
                    result = future.result()
                    result_queue.put(result)

                    if result.status == AppConfig.SCAN_STATUS_SUCCESS:
                        reachable_ips.append(ip_address)
                except Exception as exc:
                    AppLogger.log_message(
                        "error",
                        f"Unhandled ping worker error for {ip_address}: {exc}",
                    )
                    result_queue.put(
                        ScanResult(
                            task_name="Ping",
                            target=ip_address,
                            status=AppConfig.SCAN_STATUS_ERROR,
                            message="Unhandled ping worker error.",
                            error=str(exc),
                        )
                    )

        return reachable_ips

    def ping_host(self, ip_address: str) -> ScanResult:
        """
        Ping a single host.

        Args:
            ip_address: IP address to ping.

        Returns:
            ScanResult describing ping outcome.
        """
        started = time.perf_counter()
        started_at = datetime.now().strftime(AppConfig.STORAGE_TIMESTAMP_FORMAT)

        try:
            command = self.build_ping_command(ip_address)
            completed = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=self.timeout + 0.5,
                check=False,
            )

            duration = time.perf_counter() - started
            completed_at = datetime.now().strftime(
                AppConfig.STORAGE_TIMESTAMP_FORMAT
            )

            if completed.returncode == 0:
                return ScanResult(
                    task_name="Ping",
                    target=ip_address,
                    status=AppConfig.SCAN_STATUS_SUCCESS,
                    message="Host reachable.",
                    started_at=started_at,
                    completed_at=completed_at,
                    duration_seconds=duration,
                )

            return ScanResult(
                task_name="Ping",
                target=ip_address,
                status=AppConfig.SCAN_STATUS_WARNING,
                message="Host unreachable.",
                error=(completed.stderr or completed.stdout).strip(),
                started_at=started_at,
                completed_at=completed_at,
                duration_seconds=duration,
            )

        except subprocess.TimeoutExpired as exc:
            duration = time.perf_counter() - started
            return ScanResult(
                task_name="Ping",
                target=ip_address,
                status=AppConfig.SCAN_STATUS_WARNING,
                message="Ping timed out.",
                error=str(exc),
                started_at=started_at,
                completed_at=datetime.now().strftime(
                    AppConfig.STORAGE_TIMESTAMP_FORMAT
                ),
                duration_seconds=duration,
            )
        except Exception as exc:
            duration = time.perf_counter() - started
            AppLogger.log_message(
                "error",
                f"Ping failed for {ip_address}: {exc}",
            )
            return ScanResult(
                task_name="Ping",
                target=ip_address,
                status=AppConfig.SCAN_STATUS_ERROR,
                message="Ping command failed.",
                error=str(exc),
                started_at=started_at,
                completed_at=datetime.now().strftime(
                    AppConfig.STORAGE_TIMESTAMP_FORMAT
                ),
                duration_seconds=duration,
            )

    def build_ping_command(self, ip_address: str) -> List[str]:
        """
        Build a platform-safe ping command.

        Args:
            ip_address: IP address to ping.

        Returns:
            Subprocess command list.
        """
        safe_ip = str(ipaddress.ip_address(ip_address))

        if platform.system().lower() == "windows":
            timeout_ms = int(self.timeout * 1000)
            return ["ping", "-n", "1", "-w", str(timeout_ms), safe_ip]

        return ["ping", "-c", "1", "-W", str(int(self.timeout)), safe_ip]

# ------------------ Hostname Resolver ------------------ #
class HostnameResolver:
    """
    Resolve IP addresses to hostnames and apply optional filtering.

    Expected no-PTR DNS responses for common subnet edge addresses such as .1
    and .255 are logged at debug level instead of warning level.
    """

    QUIET_REVERSE_DNS_SUFFIXES = {"0", "1", "255"}

    def __init__(self, dns_timeout: float = AppConfig.DEFAULT_DNS_TIMEOUT):
        """
        Initialize resolver.

        Args:
            dns_timeout: Timeout for DNS resolution attempts.
        """
        self.dns_timeout = max(0.5, float(dns_timeout))

    def resolve(self, ip_address: str) -> Tuple[Optional[str], Optional[str]]:
        """
        Resolve a single IP address to hostname and FQDN.

        Args:
            ip_address: Target IP.

        Returns:
            Tuple of (hostname, fqdn).
        """
        try:
            socket.setdefaulttimeout(self.dns_timeout)
            fqdn, _, _ = socket.gethostbyaddr(ip_address)

            hostname = self.normalize_hostname(fqdn.split(".")[0])
            return hostname, fqdn

        except socket.herror as exc:
            self._log_dns_failure(ip_address, exc)
            return None, None
        except socket.gaierror as exc:
            self._log_dns_failure(ip_address, exc)
            return None, None
        except Exception as exc:
            self._log_dns_failure(ip_address, exc)
            return None, None

    def resolve_many(
        self,
        ip_addresses: Sequence[str],
        filters: List[str],
        max_workers: int,
        stop_event: threading.Event,
        result_queue: queue.Queue,
    ) -> List[ComputerRecord]:
        """
        Resolve many IPs concurrently.

        Args:
            ip_addresses: List of pingable IPs.
            filters: Device name filters.
            max_workers: Thread count.
            stop_event: Cancellation event.
            result_queue: Output queue.

        Returns:
            List of ComputerRecord objects.
        """
        results: List[ComputerRecord] = []

        worker_count = max(
            1,
            min(AppConfig.MAX_PING_WORKERS, int(max_workers)),
        )

        with ThreadPoolExecutor(
            max_workers=worker_count,
            thread_name_prefix="DNSWorker",
        ) as executor:

            future_map = {
                executor.submit(self._resolve_worker, ip, filters): ip
                for ip in ip_addresses
                if not stop_event.is_set()
            }

            for future in as_completed(future_map):
                if stop_event.is_set():
                    break

                try:
                    record, scan_result = future.result()
                    result_queue.put(scan_result)

                    if record:
                        results.append(record)

                except Exception as exc:
                    ip_address = future_map[future]
                    AppLogger.log_message(
                        "error",
                        f"Unhandled resolver error for {ip_address}: {exc}",
                    )
                    result_queue.put(
                        ScanResult(
                            task_name="Resolve",
                            target=ip_address,
                            status=AppConfig.SCAN_STATUS_ERROR,
                            message="Unhandled resolver error.",
                            error=str(exc),
                        )
                    )

        return results

    def _resolve_worker(
        self,
        ip_address: str,
        filters: List[str],
    ) -> Tuple[Optional[ComputerRecord], ScanResult]:
        """
        Worker method for resolving a single IP.

        Args:
            ip_address: Target IP.
            filters: Hostname filters.

        Returns:
            Tuple of ComputerRecord or None and ScanResult.
        """
        start = time.perf_counter()
        started_at = datetime.now().strftime(AppConfig.STORAGE_TIMESTAMP_FORMAT)

        hostname, fqdn = self.resolve(ip_address)

        duration = time.perf_counter() - start
        completed_at = datetime.now().strftime(AppConfig.STORAGE_TIMESTAMP_FORMAT)

        if hostname:
            matches = self.matches_filters(hostname, filters)

            if filters and not matches:
                return None, ScanResult(
                    task_name="Resolve",
                    target=ip_address,
                    status=AppConfig.SCAN_STATUS_SKIPPED,
                    message=f"Filtered out: {hostname}",
                    started_at=started_at,
                    completed_at=completed_at,
                    duration_seconds=duration,
                )

            record = ComputerRecord(
                hostname=hostname,
                fqdn=fqdn or "",
                ip_address=ip_address,
                pingable=True,
                matched_filter=matches or not filters,
            )
            record.update_last_seen()

            return record, ScanResult(
                task_name="Resolve",
                target=ip_address,
                status=AppConfig.SCAN_STATUS_SUCCESS,
                message=f"Resolved: {hostname}",
                started_at=started_at,
                completed_at=completed_at,
                duration_seconds=duration,
            )

        if self._is_quiet_edge_address(ip_address):
            return None, ScanResult(
                task_name="Resolve",
                target=ip_address,
                status=AppConfig.SCAN_STATUS_SKIPPED,
                message="Reverse DNS skipped for subnet edge address.",
                started_at=started_at,
                completed_at=completed_at,
                duration_seconds=duration,
            )

        if filters:
            return None, ScanResult(
                task_name="Resolve",
                target=ip_address,
                status=AppConfig.SCAN_STATUS_WARNING,
                message="Hostname unresolved, excluded by filter.",
                started_at=started_at,
                completed_at=completed_at,
                duration_seconds=duration,
            )

        record = ComputerRecord(
            hostname="",
            fqdn="",
            ip_address=ip_address,
            pingable=True,
            matched_filter=True,
        )
        record.update_last_seen()

        return record, ScanResult(
            task_name="Resolve",
            target=ip_address,
            status=AppConfig.SCAN_STATUS_WARNING,
            message="Hostname unresolved.",
            started_at=started_at,
            completed_at=completed_at,
            duration_seconds=duration,
        )

    def normalize_hostname(self, hostname: str) -> str:
        """
        Normalize hostname string.

        Args:
            hostname: Raw hostname.

        Returns:
            Sanitized hostname.
        """
        return ScanInputValidator.sanitize_hostname(hostname)

    def matches_filters(self, hostname: str, filters: List[str]) -> bool:
        """
        Apply filter matching.

        Args:
            hostname: Hostname.
            filters: Filter list.

        Returns:
            True if matches.
        """
        return ScanInputValidator.hostname_matches_filters(hostname, filters)

    def _log_dns_failure(self, ip_address: str, exc: Exception) -> None:
        """
        Log DNS failure at the appropriate severity.

        Args:
            ip_address: Target IP.
            exc: DNS exception.
        """
        if self._is_quiet_edge_address(ip_address) and self._is_no_host_error(exc):
            AppLogger.log_message(
                "debug",
                f"Reverse DNS not expected for subnet edge address {ip_address}: {exc}",
            )
            return

        AppLogger.log_message(
            "debug",
            f"DNS resolution failed for {ip_address}: {exc}",
        )

    def _is_quiet_edge_address(self, ip_address: str) -> bool:
        """
        Return True for subnet edge addresses where missing PTR is expected.

        Args:
            ip_address: Target IP.

        Returns:
            True for .0, .1, or .255 IPv4 addresses.
        """
        try:
            parsed_ip = ipaddress.ip_address(ip_address)

            if parsed_ip.version != 4:
                return False

            last_octet = str(parsed_ip).split(".")[-1]
            return last_octet in self.QUIET_REVERSE_DNS_SUFFIXES
        except ValueError:
            return False

    def _is_no_host_error(self, exc: Exception) -> bool:
        """
        Detect expected no-host reverse DNS errors.

        Args:
            exc: DNS exception.

        Returns:
            True when the error means no DNS host/PTR record exists.
        """
        text = str(exc).lower()
        return (
            "11004" in text
            or "host not found" in text
            or "unknown host" in text
            or "no host" in text
            or "not known" in text
        )

# ------------------ Registry Inventory Scanner ------------------ #
class RegistryInventoryScanner:
    """
    Gather installed software records from remote Windows computers.

    Inventory is performed with PowerShell registry queries. Win32_Product is
    intentionally avoided because it is slow and can trigger MSI repair checks.
    """

    def __init__(
        self,
        credential_manager: Optional[CredentialManager] = None,
        timeout: float = AppConfig.DEFAULT_REGISTRY_TIMEOUT,
    ):
        """
        Initialize registry inventory scanner.
        """
        self.credential_manager = credential_manager or CredentialManager()
        self.timeout = max(5.0, float(timeout))

    def inventory_computer(
        self,
        computer_record: ComputerRecord,
    ) -> Tuple[ComputerRecord, List[SoftwareRecord], ScanResult]:
        """
        Inventory one computer.
        """
        started = time.perf_counter()
        started_at = datetime.now().strftime(AppConfig.STORAGE_TIMESTAMP_FORMAT)
        target = computer_record.hostname or computer_record.ip_address

        if not target:
            error = "Computer record has no hostname or IP address."
            computer_record.mark_inventory_failure(error)
            return computer_record, [], ScanResult(
                task_name="Inventory",
                target="Unknown",
                status=AppConfig.SCAN_STATUS_ERROR,
                message="Inventory skipped.",
                error=error,
                started_at=started_at,
                completed_at=datetime.now().strftime(
                    AppConfig.STORAGE_TIMESTAMP_FORMAT
                ),
                duration_seconds=0.0,
            )

        try:
            command = self.build_registry_command(target)
            completed = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=self.timeout,
                check=False,
            )

            duration = time.perf_counter() - started
            completed_at = datetime.now().strftime(
                AppConfig.STORAGE_TIMESTAMP_FORMAT
            )

            if completed.returncode != 0:
                error_text = completed.stderr.strip() or completed.stdout.strip()
                friendly_error = self.classify_error(error_text)
                computer_record.mark_inventory_failure(friendly_error)

                return computer_record, [], ScanResult(
                    task_name="Inventory",
                    target=target,
                    status=AppConfig.SCAN_STATUS_ERROR,
                    message="Inventory failed.",
                    error=friendly_error,
                    started_at=started_at,
                    completed_at=completed_at,
                    duration_seconds=duration,
                )

            software_records, telemetry = self.parse_software_json(
                completed.stdout,
                computer_record,
            )
            if telemetry:
                AppLogger.log_message(
                    "debug",
                    (
                        f"Inventory timing for {target}: "
                        f"session_create_ms={telemetry.get('session_create_ms', 0.0):.2f}, "
                        f"query_ms={telemetry.get('query_ms', 0.0):.2f}"
                    ),
                )
            computer_record.mark_inventory_success()

            return computer_record, software_records, ScanResult(
                task_name="Inventory",
                target=target,
                status=AppConfig.SCAN_STATUS_SUCCESS,
                message=(
                    f"Inventory success: {len(software_records)} "
                    "software entries."
                ),
                data={"software_count": len(software_records)},
                started_at=started_at,
                completed_at=completed_at,
                duration_seconds=duration,
            )

        except subprocess.TimeoutExpired as exc:
            duration = time.perf_counter() - started
            error = "PowerShell registry query timed out."
            computer_record.mark_inventory_failure(error)

            return computer_record, [], ScanResult(
                task_name="Inventory",
                target=target,
                status=AppConfig.SCAN_STATUS_ERROR,
                message="Inventory timeout.",
                error=str(exc),
                started_at=started_at,
                completed_at=datetime.now().strftime(
                    AppConfig.STORAGE_TIMESTAMP_FORMAT
                ),
                duration_seconds=duration,
            )
        except Exception as exc:
            duration = time.perf_counter() - started
            error = self.classify_error(str(exc))
            computer_record.mark_inventory_failure(error)
            AppLogger.log_message(
                "error",
                f"Inventory failed for {target}: {exc}",
            )

            return computer_record, [], ScanResult(
                task_name="Inventory",
                target=target,
                status=AppConfig.SCAN_STATUS_ERROR,
                message="Inventory failed.",
                error=error,
                started_at=started_at,
                completed_at=datetime.now().strftime(
                    AppConfig.STORAGE_TIMESTAMP_FORMAT
                ),
                duration_seconds=duration,
            )

    def inventory_many(
        self,
        computer_records: Sequence[ComputerRecord],
        max_workers: int,
        stop_event: threading.Event,
        result_queue: queue.Queue,
    ) -> List[Tuple[ComputerRecord, List[SoftwareRecord]]]:
        """
        Inventory multiple computers concurrently.
        """
        inventory_results: List[Tuple[ComputerRecord, List[SoftwareRecord]]] = []
        max_worker_count = max(
            AppConfig.MIN_INVENTORY_WORKERS,
            min(AppConfig.MAX_INVENTORY_WORKERS, int(max_workers)),
        )
        target_workers = max(AppConfig.MIN_INVENTORY_WORKERS, min(4, max_worker_count))
        pending = deque(record for record in computer_records)
        in_flight: Dict[Any, ComputerRecord] = {}
        durations: deque = deque(maxlen=20)
        failures: deque = deque(maxlen=20)
        winrm_failures: deque = deque(maxlen=20)

        with ThreadPoolExecutor(max_workers=max_worker_count, thread_name_prefix="InventoryWorker") as executor:
            while (pending or in_flight) and not stop_event.is_set():
                while pending and len(in_flight) < target_workers and not stop_event.is_set():
                    record = pending.popleft()
                    in_flight[executor.submit(self.inventory_computer, record)] = record
                if not in_flight:
                    continue
                try:
                    completed_futures = as_completed(list(in_flight.keys()), timeout=0.2)
                    future = next(completed_futures)
                except Exception:
                    continue
                for future in [future]:
                    record = in_flight.pop(future)
                    target = record.hostname or record.ip_address
                    try:
                        computer, software_records, scan_result = future.result()
                        result_queue.put(scan_result)
                        inventory_results.append((computer, software_records))
                        durations.append(max(0.0, float(scan_result.duration_seconds or 0.0)))
                        failed = computer.inventory_status == AppConfig.INVENTORY_STATUS_FAILED
                        failures.append(1 if failed else 0)
                        reason = (computer.inventory_error or scan_result.error or "").lower()
                        winrm_failures.append(1 if any(t in reason for t in ("winrm", "cannot connect", "host offline", "rpc")) else 0)
                    except Exception as exc:
                        AppLogger.log_message("error", f"Unhandled inventory worker error for {target}: {exc}")
                        record.mark_inventory_failure(str(exc))
                        failures.append(1)
                        winrm_failures.append(0)
                        result_queue.put(ScanResult(task_name="Inventory", target=target, status=AppConfig.SCAN_STATUS_ERROR, message="Unhandled inventory worker error.", error=str(exc)))
                        inventory_results.append((record, []))
                    if len(failures) >= 8:
                        error_rate = sum(failures) / len(failures)
                        winrm_rate = sum(winrm_failures) / max(1, len(winrm_failures))
                        if (error_rate >= 0.25 or winrm_rate >= 0.2) and target_workers > AppConfig.MIN_INVENTORY_WORKERS:
                            target_workers -= 1
                        elif error_rate <= 0.12 and winrm_rate < 0.2 and pending and target_workers < max_worker_count:
                            target_workers += 1
                    break

        return inventory_results

    def build_registry_command(self, hostname: str) -> List[str]:
        """
        Build a PowerShell registry inventory command.
        """
        safe_hostname = ScanInputValidator.sanitize_powershell_argument(hostname)
        registry_paths = ", ".join(
            f"'{path}'" for path in AppConfig.REGISTRY_UNINSTALL_PATHS
        )

        query_script = (
            "$paths = @("
            f"{registry_paths}"
            "); "
            "$items = foreach ($path in $paths) { "
            "Get-ItemProperty -Path $path -ErrorAction SilentlyContinue "
            "}; "
            "$items | Where-Object { $_.DisplayName } | "
            "Select-Object DisplayName, DisplayVersion, Publisher, "
            "InstallDate, UninstallString, InstallLocation, EstimatedSize, "
            "PSChildName, PSPath"
        )

        credential_argument = self.credential_manager.build_credential_argument()
        command = (
            "$sessionCreateStart = Get-Date; "
            f"$session = New-PSSession -ComputerName \"{safe_hostname}\" "
            f"{credential_argument} -ErrorAction Stop; "
            "$sessionCreateMs = ((Get-Date) - $sessionCreateStart).TotalMilliseconds; "
            "try { "
            "$queryStart = Get-Date; "
            "$attempt = 0; "
            "$maxAttempts = 2; "
            "$result = $null; "
            "while ($attempt -lt $maxAttempts -and -not $result) { "
            "$attempt++; "
            "try { "
            f"$result = Invoke-Command -Session $session -ScriptBlock {{ {query_script} }} -ErrorAction Stop; "
            "} catch { if ($attempt -ge $maxAttempts) { throw } } "
            "}; "
            "$queryMs = ((Get-Date) - $queryStart).TotalMilliseconds; "
            "[PSCustomObject]@{ "
            "SessionCreateMs = [Math]::Round($sessionCreateMs, 2); "
            "QueryMs = [Math]::Round($queryMs, 2); "
            "Data = $result "
            "} | ConvertTo-Json -Compress -Depth 6 "
            "} finally { if ($session) { Remove-PSSession -Session $session -ErrorAction SilentlyContinue } }"
        )

        command = self.credential_manager.apply_credentials_to_command(command)

        return [
            AppConfig.POWERSHELL_EXECUTABLE,
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            command,
        ]

    def parse_software_json(
        self,
        raw_json: str,
        computer_record: ComputerRecord,
    ) -> Tuple[List[SoftwareRecord], Dict[str, float]]:
        """
        Parse PowerShell JSON output into SoftwareRecord objects.
        """
        if not raw_json or not raw_json.strip():
            return [], {}

        try:
            parsed = json.loads(raw_json)
            telemetry: Dict[str, float] = {}
            if isinstance(parsed, dict) and "Data" in parsed:
                telemetry = {
                    "session_create_ms": float(parsed.get("SessionCreateMs") or 0.0),
                    "query_ms": float(parsed.get("QueryMs") or 0.0),
                }
                data = parsed.get("Data")
                items = data if isinstance(data, list) else ([data] if data else [])
            else:
                items = parsed if isinstance(parsed, list) else [parsed]
            records: List[SoftwareRecord] = []

            for item in items:
                if not isinstance(item, dict):
                    continue

                display_name = str(item.get("DisplayName") or "").strip()
                if not display_name:
                    continue

                ps_path = str(item.get("PSPath") or "")
                architecture = "32-bit" if "WOW6432Node" in ps_path else "64-bit"

                record = SoftwareRecord(
                    computer_id=computer_record.computer_id,
                    display_name=display_name,
                    display_version=str(item.get("DisplayVersion") or "").strip(),
                    publisher=str(item.get("Publisher") or "").strip(),
                    install_date=str(item.get("InstallDate") or "").strip(),
                    uninstall_string=str(item.get("UninstallString") or "").strip(),
                    install_location=str(item.get("InstallLocation") or "").strip(),
                    estimated_size=str(item.get("EstimatedSize") or "").strip(),
                    registry_key=str(item.get("PSChildName") or "").strip(),
                    registry_hive="HKLM",
                    architecture=architecture,
                    raw_json=json.dumps(item, ensure_ascii=False),
                    scan_id=computer_record.scan_id,
                )

                if record.is_valid():
                    records.append(record)

            return records, telemetry

        except json.JSONDecodeError as exc:
            AppLogger.log_message(
                "error",
                f"JSON parse failure for "
                f"{computer_record.hostname or computer_record.ip_address}: {exc}",
            )
            return [], {}

    def classify_error(self, error_text: str) -> str:
        """
        Convert raw PowerShell error text into a user-friendly category.
        """
        lowered = (error_text or "").lower()

        error_map = {
            "access is denied": "Access denied.",
            "access denied": "Access denied.",
            "winrm": "WinRM unavailable.",
            "cannot connect": "Host offline or unavailable.",
            "rpc server": "RPC server unavailable.",
            "timed out": "Inventory timeout.",
            "json": "JSON parse failure.",
        }

        for token, message in error_map.items():
            if token in lowered:
                return message

        return error_text.strip() or "Unknown inventory error."

# ------------------ Thread Task Manager ------------------ #
class ThreadTaskManager:
    """
    Manage application worker pools, submitted futures, cancellation, and
    shutdown for scan workflows.

    This class centralizes thread lifecycle behavior so scan classes do not
    create unmanaged ThreadPoolExecutor instances directly.
    """

    def __init__(self, stop_event: threading.Event):
        """
        Initialize thread task manager.

        Args:
            stop_event: Shared cancellation event.
        """
        self.stop_event = stop_event
        self._executors: Dict[str, ThreadPoolExecutor] = {}
        self._futures: Dict[str, List[Any]] = {}
        self._lock = threading.RLock()
        self._completion_event = threading.Event()

    def create_pool(
        self,
        name: str,
        max_workers: int,
        thread_name_prefix: str,
    ) -> ThreadPoolExecutor:
        """
        Create or replace a named worker pool.

        Args:
            name: Pool name.
            max_workers: Maximum worker count.
            thread_name_prefix: Thread name prefix.

        Returns:
            ThreadPoolExecutor for submitted work.
        """
        with self._lock:
            self.shutdown_pool(name, cancel_futures=True)

            executor = ThreadPoolExecutor(
                max_workers=max(1, int(max_workers)),
                thread_name_prefix=thread_name_prefix,
            )
            self._executors[name] = executor
            self._futures[name] = []
            AppLogger.log_message(
                "debug",
                f"Thread pool created: {name} ({max_workers} workers).",
            )
            return executor

    def submit(
        self,
        pool_name: str,
        function: Any,
        *args: Any,
        **kwargs: Any,
    ) -> Optional[Any]:
        """
        Submit work to a named pool.

        Args:
            pool_name: Existing pool name.
            function: Callable to execute.
            *args: Positional callable arguments.
            **kwargs: Keyword callable arguments.

        Returns:
            Future when submitted; otherwise None.
        """
        with self._lock:
            if self.stop_event.is_set():
                return None

            executor = self._executors.get(pool_name)
            if not executor:
                AppLogger.log_message(
                    "error",
                    f"Cannot submit task; pool does not exist: {pool_name}",
                )
                return None

            try:
                future = executor.submit(function, *args, **kwargs)
                future.add_done_callback(lambda _: self._completion_event.set())
                self._futures.setdefault(pool_name, []).append(future)
                return future
            except RuntimeError as exc:
                AppLogger.log_message(
                    "error",
                    f"Failed to submit task to pool {pool_name}: {exc}",
                )
                return None

    def collect_ready(self, pool_name: str) -> List[Any]:
        """
        Return completed futures for a pool and remove them from tracking.

        Args:
            pool_name: Pool name.

        Returns:
            List of completed futures.
        """
        with self._lock:
            futures = self._futures.get(pool_name, [])
            ready = [future for future in futures if future.done()]
            self._futures[pool_name] = [
                future for future in futures
                if not future.done()
            ]

            if not any(
                tracked_future.done()
                for tracked_futures in self._futures.values()
                for tracked_future in tracked_futures
            ):
                self._completion_event.clear()

            return ready


    def wait_for_completion(self, timeout: float) -> bool:
        """
        Wait for any tracked future to complete.

        Args:
            timeout: Maximum wait time in seconds.

        Returns:
            True when at least one completion was signaled, else False.
        """
        return self._completion_event.wait(timeout=max(0.0, float(timeout)))

    def has_pending(self, pool_name: Optional[str] = None) -> bool:
        """
        Check whether futures remain pending.

        Args:
            pool_name: Optional pool name. When omitted, all pools are checked.

        Returns:
            True when any pending futures remain.
        """
        with self._lock:
            pools = (
                [pool_name]
                if pool_name
                else list(self._futures.keys())
            )

            for name in pools:
                if any(not future.done() for future in self._futures.get(name, [])):
                    return True

            return False

    def cancel_all(self) -> None:
        """Signal cancellation and cancel all futures that have not started."""
        with self._lock:
            self.stop_event.set()

            for pool_name, futures in self._futures.items():
                for future in futures:
                    future.cancel()

                AppLogger.log_message(
                    "warning",
                    f"Cancellation requested for pool: {pool_name}",
                )

    def shutdown_pool(
        self,
        name: str,
        cancel_futures: bool = False,
    ) -> None:
        """
        Shutdown a named pool.

        Args:
            name: Pool name.
            cancel_futures: Cancel pending futures before shutdown.
        """
        executor = self._executors.pop(name, None)
        futures = self._futures.pop(name, [])

        if cancel_futures:
            for future in futures:
                future.cancel()

        if executor:
            try:
                executor.shutdown(wait=False, cancel_futures=cancel_futures)
                AppLogger.log_message("debug", f"Thread pool shut down: {name}")
            except TypeError:
                executor.shutdown(wait=False)
                AppLogger.log_message("debug", f"Thread pool shut down: {name}")

    def shutdown_all(self, cancel_futures: bool = False) -> None:
        """
        Shutdown all pools.

        Args:
            cancel_futures: Cancel pending futures before shutdown.
        """
        with self._lock:
            pool_names = list(self._executors.keys())

        for pool_name in pool_names:
            self.shutdown_pool(pool_name, cancel_futures=cancel_futures)

# ------------------ Scan Coordination ------------------ #
class ScanCoordinator:
    """
    Orchestrate the full scan lifecycle using a streaming pipeline.

    Ping, DNS, and inventory run concurrently. Ping submission is intentionally
    throttled with an in-flight window to prevent ICMP flooding from creating
    false offline results on busy networks.
    """

    PING_POOL = "ping"
    DNS_POOL = "dns"
    INVENTORY_POOL = "inventory"

    PHASE_PING_PERCENT_STEP = 10
    PHASE_RESOLVED_STEP = 10
    PHASE_INVENTORY_STEP = 1

    DEFAULT_PING_RETRY_COUNT = 1
    DEFAULT_MAX_PING_IN_FLIGHT = 64
    DEFAULT_DB_BATCH_SIZE = 50
    DEFAULT_DB_BATCH_SECONDS = 2.0
    INVENTORY_WINDOW_SIZE = 20
    INVENTORY_ADJUST_INTERVAL = 8
    INVENTORY_SCALE_UP_ERROR_RATE = 0.12
    INVENTORY_SCALE_DOWN_ERROR_RATE = 0.25
    INVENTORY_SCALE_DOWN_WINRM_RATE = 0.2

    def __init__(
        self,
        database_manager: DatabaseManager,
        result_queue: queue.Queue,
        credential_manager: Optional[CredentialManager] = None,
        preferences_manager: Optional[PreferencesManager] = None,
    ):
        """
        Initialize scan coordinator.

        Args:
            database_manager: Database service.
            result_queue: Queue used for UI-safe scan updates.
            credential_manager: Optional credential service.
        """
        self.database_manager = database_manager
        self.result_queue = result_queue
        self.credential_manager = credential_manager or CredentialManager()
        self.preferences_manager = preferences_manager or PreferencesManager()

        self.validator = ScanInputValidator()
        self.ip_scanner = IpRangeScanner()
        self.hostname_resolver = HostnameResolver()
        self.inventory_scanner = RegistryInventoryScanner(
            credential_manager=self.credential_manager
        )

        self.stop_event = threading.Event()
        self.task_manager = ThreadTaskManager(self.stop_event)
        self.scan_thread: Optional[threading.Thread] = None

        self.scan_id: Optional[int] = None
        self.ip_range: str = ""
        self.filter_text: str = ""
        self.filters: List[str] = []
        self.options: Dict[str, Any] = {}

        self.expanded_ips: List[str] = []
        self.pending_ping_ips: deque = deque()
        self.active_ping_targets: set = set()
        self.completed_ping_targets: set = set()

        self.pingable_ips: List[str] = []
        self.computer_records: List[ComputerRecord] = []
        self.inventory_results: List[Tuple[ComputerRecord, List[SoftwareRecord]]] = []
        self.pending_inventory_records: deque = deque()
        self.inventory_in_flight = 0
        self.inventory_target_workers = AppConfig.DEFAULT_MAX_INVENTORY_WORKERS
        self.inventory_worker_floor = AppConfig.MIN_INVENTORY_WORKERS
        self.inventory_worker_ceiling = AppConfig.DEFAULT_MAX_INVENTORY_WORKERS
        self.inventory_duration_window: deque = deque(maxlen=self.INVENTORY_WINDOW_SIZE)
        self.inventory_failure_window: deque = deque(maxlen=self.INVENTORY_WINDOW_SIZE)
        self.inventory_winrm_window: deque = deque(maxlen=self.INVENTORY_WINDOW_SIZE)

        self.ping_completed = 0
        self.resolved_count = 0
        self.inventory_completed = 0
        self.inventory_success_count = 0
        self.inventory_failed_count = 0
        self.total_software_records = 0
        self.unique_software_titles = 0
        self._software_keys: set = set()
        self._db_batch_write_count = 0
        self._db_last_commit_time = 0.0
        self._db_batch_size = self.DEFAULT_DB_BATCH_SIZE
        self._db_batch_seconds = self.DEFAULT_DB_BATCH_SECONDS

        self.ping_retry_count = self.DEFAULT_PING_RETRY_COUNT
        self.max_ping_in_flight = self.DEFAULT_MAX_PING_IN_FLIGHT

        self.started_at: float = 0.0
        self._last_phase_feed_state: Dict[str, Any] = {}
        self._last_dashboard_update_time = 0.0

    def start_scan(
        self,
        ip_range: str,
        filters: str,
        options: Dict[str, Any],
    ) -> bool:
        """
        Start a scan in a background thread.

        Args:
            ip_range: User-provided IP range.
            filters: User-provided device filter string.
            options: Scan options, including worker counts.

        Returns:
            True if scan started, otherwise False.
        """
        if self.scan_thread and self.scan_thread.is_alive():
            self.result_queue.put(
                ScanResult(
                    task_name="Coordinator",
                    status=AppConfig.SCAN_STATUS_WARNING,
                    message="A scan is already running.",
                )
            )
            return False

        self.ip_range = str(ip_range or "").strip()
        self.filter_text = str(filters or "").strip()
        self.filters = self.validator.parse_device_filters(self.filter_text)
        self.options = dict(options or {})

        self._reset_runtime_state()
        self.stop_event.clear()

        self.scan_thread = threading.Thread(
            target=self.run_scan_workflow,
            name="ScanCoordinator",
            daemon=True,
        )
        self.scan_thread.start()
        return True

    def cancel_scan(self) -> None:
        """Request cancellation of the active scan and stop managed pools."""
        self.stop_event.set()
        self.task_manager.cancel_all()

        self._emit_progress(
            phase="cancelled",
            message="Cancellation requested.",
            force_feed=True,
        )
        AppLogger.log_message("warning", "Scan cancellation requested.")

    def shutdown(self) -> None:
        """Shutdown all managed worker pools."""
        self.stop_event.set()
        self.task_manager.shutdown_all(cancel_futures=True)

    def run_scan_workflow(self) -> None:
        """Execute the full streaming scan workflow."""
        self.started_at = time.perf_counter()

        try:
            self._emit_progress(
                phase="validating",
                message="Validating scan input.",
                force_feed=True,
            )

            is_valid, message = self.validator.validate_ip_range(self.ip_range)
            if not is_valid:
                self.result_queue.put(
                    ScanResult(
                        task_name="Validation",
                        target=self.ip_range,
                        status=AppConfig.SCAN_STATUS_ERROR,
                        message=message,
                        error=message,
                    )
                )
                self.finalize_scan(failed=True)
                return

            self.expanded_ips = self.validator.expand_ip_range(self.ip_range)
            self.pending_ping_ips = deque(self.expanded_ips)

            self._emit_progress(
                phase="starting",
                total=self._progress_total_units(),
                completed=self._progress_completed_units(),
                message=(
                    f"Starting streaming scan for "
                    f"{len(self.expanded_ips)} IP addresses."
                ),
                force_feed=True,
            )

            self.scan_id = self.database_manager.create_scan(
                self.ip_range,
                self.filter_text,
            )

            if not self.scan_id:
                raise RuntimeError("Unable to create scan database record.")

            self._run_streaming_pipeline()

            if self.stop_event.is_set():
                self.finalize_scan(cancelled=True)
                return

            self.finalize_scan(cancelled=False)

        except Exception as exc:
            AppLogger.log_message("critical", f"Scan workflow failed: {exc}")
            self.result_queue.put(
                ScanResult(
                    scan_id=self.scan_id,
                    task_name="Coordinator",
                    status=AppConfig.SCAN_STATUS_ERROR,
                    message="Scan workflow failed.",
                    error=str(exc),
                )
            )
            self.finalize_scan(failed=True)
        finally:
            self.task_manager.shutdown_all(cancel_futures=False)

    def _run_streaming_pipeline(self) -> None:
        """
        Run ping, DNS, and inventory concurrently.

        Ping work is submitted incrementally. As soon as a ping slot opens,
        another IP is submitted. DNS and inventory work begin immediately after
        successful upstream results.
        """
        ping_workers = self._worker_count(
            self.options.get("max_ping_workers"),
            AppConfig.MIN_PING_WORKERS,
            AppConfig.MAX_PING_WORKERS,
            AppConfig.DEFAULT_MAX_PING_WORKERS,
        )
        inventory_workers = self._worker_count(
            self.options.get("max_inventory_workers"),
            AppConfig.MIN_INVENTORY_WORKERS,
            AppConfig.MAX_INVENTORY_WORKERS,
            AppConfig.DEFAULT_MAX_INVENTORY_WORKERS,
        )
        self._initialize_inventory_worker_bounds(inventory_workers)

        dns_workers = max(
            AppConfig.MIN_PING_WORKERS,
            min(AppConfig.MAX_PING_WORKERS, ping_workers),
        )

        self.ping_retry_count = self._safe_non_negative_int(
            self.options.get("ping_retry_count"),
            self.DEFAULT_PING_RETRY_COUNT,
        )
        self.max_ping_in_flight = self._safe_bounded_int(
            self.options.get("max_ping_in_flight"),
            minimum=1,
            maximum=max(1, ping_workers),
            default=min(self.DEFAULT_MAX_PING_IN_FLIGHT, ping_workers),
        )

        self.task_manager.create_pool(self.PING_POOL, ping_workers, "PingWorker")
        self.task_manager.create_pool(self.DNS_POOL, dns_workers, "DNSWorker")
        self.task_manager.create_pool(
            self.INVENTORY_POOL,
            inventory_workers,
            "InventoryWorker",
        )

        self._db_batch_size = self._safe_bounded_int(
            self.options.get("db_batch_size"),
            minimum=1,
            maximum=1000,
            default=self.DEFAULT_DB_BATCH_SIZE,
        )
        self._db_batch_seconds = float(
            self.options.get("db_batch_seconds", self.DEFAULT_DB_BATCH_SECONDS)
        )
        if self._db_batch_seconds <= 0:
            self._db_batch_seconds = self.DEFAULT_DB_BATCH_SECONDS

        self.database_manager.begin_transaction()
        self._db_batch_write_count = 0
        self._db_last_commit_time = time.monotonic()

        self._fill_ping_window()

        self._emit_progress(
            phase="pipeline",
            total=self._progress_total_units(),
            completed=self._progress_completed_units(),
            message=(
                f"Pipeline running. Ping workers: {ping_workers}; "
                f"Max active pings: {self.max_ping_in_flight}; "
                f"Ping retries: {self.ping_retry_count}; "
                f"DNS workers: {dns_workers}; "
                f"Inventory workers: {inventory_workers}."
            ),
            force_feed=True,
        )

        try:
            while self._pipeline_has_work():
                if self.stop_event.is_set():
                    break

                self._process_completed_ping_futures()
                self._process_completed_dns_futures()
                self._dispatch_inventory_work()
                self._process_completed_inventory_futures()
                self._dispatch_inventory_work()

                self._fill_ping_window()

                if self.stop_event.is_set():
                    break

                self.task_manager.wait_for_completion(timeout=0.05)

                if self.stop_event.is_set():
                    break

                self._emit_pipeline_progress_if_significant()
                self._emit_dashboard_progress_if_due()

            self._process_completed_ping_futures()
            self._process_completed_dns_futures()
            self._process_completed_inventory_futures()
            self._dispatch_inventory_work()
            self._flush_db_batch(force=True)
        except Exception:
            self.database_manager.rollback_transaction()
            raise

        self._emit_progress(
            phase="pipeline",
            total=self._progress_total_units(),
            completed=self._progress_completed_units(),
            success_count=self.inventory_success_count,
            warning_count=self.resolved_count,
            error_count=self.inventory_failed_count,
            message=(
                f"Pipeline finished. Pinged {self.ping_completed}/"
                f"{len(self.expanded_ips)} | Resolved {self.resolved_count} | "
                f"Inventoried {self.inventory_completed}/"
                f"{len(self.computer_records)}"
            ),
            force_feed=True,
        )

    def _fill_ping_window(self) -> None:
        """
        Submit ping jobs until the active ping window is full.

        This prevents a large thread pool from flooding the subnet all at once.
        """
        while (
            self.pending_ping_ips
            and len(self.active_ping_targets) < self.max_ping_in_flight
            and not self.stop_event.is_set()
        ):
            ip_address = self.pending_ping_ips.popleft()
            future = self.task_manager.submit(
                self.PING_POOL,
                self._ping_with_retry_context,
                ip_address,
                0,
            )

            if future:
                self.active_ping_targets.add(ip_address)

    def _ping_with_retry_context(
        self,
        ip_address: str,
        attempt: int,
    ) -> Tuple[str, int, ScanResult]:
        """
        Run one ping attempt and include retry metadata.

        Args:
            ip_address: Target IP address.
            attempt: Zero-based attempt count.

        Returns:
            Tuple of IP address, attempt number, and ScanResult.
        """
        result = self.ip_scanner.ping_host(ip_address)
        return ip_address, attempt, result

    def _process_completed_ping_futures(self) -> None:
        """Process completed ping futures and submit DNS work for successes."""
        for future in self.task_manager.collect_ready(self.PING_POOL):
            if self.stop_event.is_set():
                return

            try:
                ip_address, attempt, result = future.result()

                if result.status == AppConfig.SCAN_STATUS_SUCCESS:
                    self._finalize_ping_result(ip_address, result)
                    self.pingable_ips.append(result.target)
                    self.task_manager.submit(
                        self.DNS_POOL,
                        self.hostname_resolver._resolve_worker,
                        result.target,
                        self.filters,
                    )
                    continue

                if attempt < self.ping_retry_count:
                    self.task_manager.submit(
                        self.PING_POOL,
                        self._ping_with_retry_context,
                        ip_address,
                        attempt + 1,
                    )
                    continue

                self._finalize_ping_result(ip_address, result)

            except Exception as exc:
                self.ping_completed += 1
                AppLogger.log_message(
                    "error",
                    f"Unhandled ping future error: {exc}",
                )
                self.result_queue.put(
                    ScanResult(
                        task_name="Ping",
                        status=AppConfig.SCAN_STATUS_ERROR,
                        message="Unhandled ping future error.",
                        error=str(exc),
                    )
                )

    def _finalize_ping_result(
        self,
        ip_address: str,
        result: ScanResult,
    ) -> None:
        """
        Mark a ping target complete and emit its final result.

        Args:
            ip_address: Pinged IP address.
            result: Final ScanResult for the target.
        """
        if ip_address in self.completed_ping_targets:
            return

        self.completed_ping_targets.add(ip_address)
        self.active_ping_targets.discard(ip_address)
        self.ping_completed += 1
        self.result_queue.put(result)

    def _process_completed_dns_futures(self) -> None:
        """Process completed DNS futures and submit inventory work for matches."""
        for future in self.task_manager.collect_ready(self.DNS_POOL):
            if self.stop_event.is_set():
                return

            try:
                record, scan_result = future.result()
                self.result_queue.put(scan_result)
                self.resolved_count += 1

                if not record:
                    continue

                record.scan_id = self.scan_id
                record.computer_id = self.database_manager.insert_computer(
                    record,
                    commit=False,
                )
                self._mark_db_write()
                self.computer_records.append(record)
                self.pending_inventory_records.append(record)

            except Exception as exc:
                self.resolved_count += 1
                AppLogger.log_message(
                    "error",
                    f"Unhandled DNS future error: {exc}",
                )
                self.result_queue.put(
                    ScanResult(
                        task_name="Resolve",
                        status=AppConfig.SCAN_STATUS_ERROR,
                        message="Unhandled DNS future error.",
                        error=str(exc),
                    )
                )

    def _process_completed_inventory_futures(self) -> None:
        """Process completed inventory futures and save inventory records."""
        for future in self.task_manager.collect_ready(self.INVENTORY_POOL):
            if self.stop_event.is_set():
                return

            try:
                computer, software_records, scan_result = future.result()
                self.inventory_in_flight = max(0, self.inventory_in_flight - 1)
                self.result_queue.put(scan_result)
                self.inventory_results.append((computer, software_records))
                self.inventory_completed += 1

                if computer.inventory_status == AppConfig.INVENTORY_STATUS_SUCCESS:
                    self.inventory_success_count += 1
                elif computer.inventory_status == AppConfig.INVENTORY_STATUS_FAILED:
                    self.inventory_failed_count += 1
                    inventory_target = computer.hostname or computer.ip_address or "Unknown"
                    inventory_reason = computer.inventory_error or scan_result.error or "Unknown reason"
                    AppLogger.log_message(
                        "error",
                        f"Inventory failed for {inventory_target} ({computer.ip_address or 'no-ip'}): {inventory_reason}",
                    )

                self.total_software_records += len(software_records)
                self._record_inventory_metrics(scan_result, computer)
                self._adjust_inventory_workers_if_needed()

                for record in software_records:
                    self._software_keys.add(record.normalized_key())

                self.unique_software_titles = len(self._software_keys)

                if computer.computer_id:
                    self.database_manager.update_computer_inventory_status(
                        computer.computer_id,
                        computer.inventory_status,
                        computer.inventory_error,
                        commit=False,
                    )
                    self._mark_db_write()
                    self.database_manager.insert_software_records(
                        computer.computer_id,
                        software_records,
                        commit=False,
                    )
                    self._mark_db_write()

            except Exception as exc:
                self.inventory_in_flight = max(0, self.inventory_in_flight - 1)
                self.inventory_completed += 1
                self.inventory_failed_count += 1
                AppLogger.log_message(
                    "error",
                    f"Unhandled inventory future error: {exc}",
                )
                self.result_queue.put(
                    ScanResult(
                        task_name="Inventory",
                        status=AppConfig.SCAN_STATUS_ERROR,
                        message="Unhandled inventory future error.",
                        error=str(exc),
                    )
                )

    def _mark_db_write(self, count: int = 1) -> None:
        """Track buffered database writes and commit periodically."""
        self._db_batch_write_count += max(0, count)
        self._flush_db_batch()

    def _flush_db_batch(self, force: bool = False) -> None:
        """Commit current write transaction and start a new batch when due."""
        if self._db_batch_write_count <= 0 and not force:
            return

        elapsed = time.monotonic() - self._db_last_commit_time
        should_commit = force or (
            self._db_batch_write_count >= self._db_batch_size
            or elapsed >= self._db_batch_seconds
        )
        if not should_commit:
            return

        self.database_manager.commit_transaction()
        self.database_manager.begin_transaction()
        self._db_batch_write_count = 0
        self._db_last_commit_time = time.monotonic()

    def _emit_pipeline_progress_if_significant(self) -> None:
        """Emit a Live Feed phase message only for meaningful milestones."""
        total_ips = max(1, len(self.expanded_ips))
        ping_percent_bucket = (
            int((self.ping_completed / total_ips) * 100)
            // self.PHASE_PING_PERCENT_STEP
            * self.PHASE_PING_PERCENT_STEP
        )
        resolved_bucket = (
            self.resolved_count
            // self.PHASE_RESOLVED_STEP
            * self.PHASE_RESOLVED_STEP
        )
        inventory_bucket = (
            self.inventory_completed
            // self.PHASE_INVENTORY_STEP
            * self.PHASE_INVENTORY_STEP
        )

        current_state = {
            "ping_percent_bucket": ping_percent_bucket,
            "resolved_bucket": resolved_bucket,
            "inventory_bucket": inventory_bucket,
            "inventory_failed_count": self.inventory_failed_count,
            "ping_complete": self.ping_completed == len(self.expanded_ips),
            "inventory_complete": (
                bool(self.computer_records)
                and self.inventory_completed == len(self.computer_records)
            ),
        }

        if current_state == self._last_phase_feed_state:
            return

        significant = (
            current_state["ping_percent_bucket"]
            != self._last_phase_feed_state.get("ping_percent_bucket")
            or current_state["resolved_bucket"]
            != self._last_phase_feed_state.get("resolved_bucket")
            or current_state["inventory_bucket"]
            != self._last_phase_feed_state.get("inventory_bucket")
            or current_state["inventory_failed_count"]
            > self._last_phase_feed_state.get("inventory_failed_count", 0)
            or (
                current_state["ping_complete"]
                and not self._last_phase_feed_state.get("ping_complete", False)
            )
            or (
                current_state["inventory_complete"]
                and not self._last_phase_feed_state.get("inventory_complete", False)
            )
        )

        if not significant:
            return

        self._last_phase_feed_state = current_state
        self._emit_progress(
            phase="pipeline",
            total=self._progress_total_units(),
            completed=self._progress_completed_units(),
            success_count=self.inventory_success_count,
            warning_count=self.resolved_count,
            error_count=self.inventory_failed_count,
            message=(
                f"Pinged {self.ping_completed}/{len(self.expanded_ips)} | "
                f"Resolved {self.resolved_count} | "
                f"Inventoried {self.inventory_completed}/"
                f"{len(self.computer_records)}"
            ),
            force_feed=True,
        )

    def _emit_dashboard_progress_if_due(self) -> None:
        """Refresh dashboard/progress bar without adding Live Feed noise."""
        now = time.perf_counter()

        if now - self._last_dashboard_update_time < 0.5:
            return

        self._last_dashboard_update_time = now
        self._emit_progress(
            phase="pipeline",
            total=self._progress_total_units(),
            completed=self._progress_completed_units(),
            success_count=self.inventory_success_count,
            warning_count=self.resolved_count,
            error_count=self.inventory_failed_count,
            message=(
                f"Pinged {self.ping_completed}/{len(self.expanded_ips)} | "
                f"Resolved {self.resolved_count} | "
                f"Inventoried {self.inventory_completed}/"
                f"{len(self.computer_records)}"
            ),
            feed_visible=False,
        )

    def _pipeline_has_work(self) -> bool:
        """
        Check whether any pipeline stage still has pending work.

        Returns:
            True when pending IPs or futures remain.
        """
        return (
            bool(self.pending_ping_ips)
            or self.task_manager.has_pending(self.PING_POOL)
            or self.task_manager.has_pending(self.DNS_POOL)
            or bool(self.pending_inventory_records)
            or self.task_manager.has_pending(self.INVENTORY_POOL)
        )

    def _dispatch_inventory_work(self) -> None:
        while (
            self.pending_inventory_records
            and self.inventory_in_flight < self.inventory_target_workers
            and not self.stop_event.is_set()
        ):
            record = self.pending_inventory_records.popleft()
            future = self.task_manager.submit(
                self.INVENTORY_POOL,
                self.inventory_scanner.inventory_computer,
                record,
            )
            if future:
                self.inventory_in_flight += 1

    def _initialize_inventory_worker_bounds(self, configured_workers: int) -> None:
        self.inventory_worker_ceiling = max(
            AppConfig.MIN_INVENTORY_WORKERS,
            configured_workers,
        )
        self.inventory_worker_floor = max(
            AppConfig.MIN_INVENTORY_WORKERS,
            min(4, self.inventory_worker_ceiling),
        )
        subnet_hint = self._subnet_hint()
        adaptive = self.preferences_manager.get("adaptive_inventory_workers", {})
        if isinstance(adaptive, dict) and subnet_hint in adaptive:
            profile = adaptive.get(subnet_hint, {})
            minimum = int(profile.get("min", self.inventory_worker_floor))
            maximum = int(profile.get("max", self.inventory_worker_ceiling))
            self.inventory_worker_floor = max(AppConfig.MIN_INVENTORY_WORKERS, min(minimum, self.inventory_worker_ceiling))
            self.inventory_worker_ceiling = max(self.inventory_worker_floor, min(maximum, configured_workers))
        self.inventory_target_workers = self.inventory_worker_floor

    def _record_inventory_metrics(self, scan_result: ScanResult, computer: ComputerRecord) -> None:
        self.inventory_duration_window.append(max(0.0, float(scan_result.duration_seconds or 0.0)))
        failed = computer.inventory_status == AppConfig.INVENTORY_STATUS_FAILED
        self.inventory_failure_window.append(1 if failed else 0)
        reason = (computer.inventory_error or scan_result.error or "").lower()
        winrm_like = any(token in reason for token in ("winrm", "cannot connect", "host offline", "rpc"))
        self.inventory_winrm_window.append(1 if winrm_like else 0)

    def _adjust_inventory_workers_if_needed(self) -> None:
        if len(self.inventory_failure_window) < self.INVENTORY_ADJUST_INTERVAL:
            return
        if self.inventory_completed % self.INVENTORY_ADJUST_INTERVAL != 0:
            return
        median_latency = statistics.median(self.inventory_duration_window) if self.inventory_duration_window else 0.0
        error_rate = sum(self.inventory_failure_window) / max(1, len(self.inventory_failure_window))
        winrm_rate = sum(self.inventory_winrm_window) / max(1, len(self.inventory_winrm_window))
        queue_depth = len(self.pending_inventory_records)
        if (error_rate >= self.INVENTORY_SCALE_DOWN_ERROR_RATE or winrm_rate >= self.INVENTORY_SCALE_DOWN_WINRM_RATE) and self.inventory_target_workers > self.inventory_worker_floor:
            self.inventory_target_workers -= 1
            AppLogger.log_message("warning", f"Adaptive inventory throttle: workers={self.inventory_target_workers}, error_rate={error_rate:.2f}, winrm_rate={winrm_rate:.2f}, median_latency={median_latency:.2f}s, queue_depth={queue_depth}")
            return
        if error_rate <= self.INVENTORY_SCALE_UP_ERROR_RATE and winrm_rate < self.INVENTORY_SCALE_DOWN_WINRM_RATE and queue_depth > 0 and self.inventory_target_workers < self.inventory_worker_ceiling:
            self.inventory_target_workers += 1
            AppLogger.log_message("info", f"Adaptive inventory scale-up: workers={self.inventory_target_workers}, error_rate={error_rate:.2f}, winrm_rate={winrm_rate:.2f}, median_latency={median_latency:.2f}s, queue_depth={queue_depth}")

    def _subnet_hint(self) -> str:
        raw = str(self.ip_range or "").strip()
        if "-" in raw:
            raw = raw.split("-", 1)[0].strip()
        if "/" in raw:
            return raw
        try:
            ip_obj = ipaddress.ip_address(raw)
            if isinstance(ip_obj, ipaddress.IPv4Address):
                return ".".join(raw.split(".")[:3]) + ".0/24"
        except ValueError:
            pass
        return raw or "default"

    def finalize_scan(
        self,
        cancelled: bool = False,
        failed: bool = False,
    ) -> None:
        """
        Finalize scan and update database summary.

        Args:
            cancelled: True when scan was cancelled.
            failed: True when scan failed unexpectedly.
        """
        summary = self.calculate_summary(include_duration=True)

        if cancelled:
            summary["status"] = AppConfig.SCAN_STATUS_CANCELLED
            phase = "cancelled"
            message = "Scan cancelled."
        elif failed:
            summary["status"] = AppConfig.SCAN_STATUS_ERROR
            phase = "complete"
            message = "Scan completed with errors."
        else:
            summary["status"] = AppConfig.SCAN_STATUS_SUCCESS
            phase = "complete"
            message = "Scan complete."

        summary["phase"] = phase
        summary["message"] = message

        if self.scan_id:
            self.database_manager.complete_scan(self.scan_id, summary)
        self._persist_inventory_worker_profile()

        self.result_queue.put(
            ScanProgress(
                phase=phase,
                total=self._progress_total_units(),
                completed=self._progress_total_units(),
                success_count=summary.get("successful_inventory_count", 0),
                error_count=summary.get("failed_inventory_count", 0),
                message=message,
                data=summary,
            )
        )
        self.result_queue.put(
            ScanResult(
                scan_id=self.scan_id,
                task_name="Coordinator",
                status=summary["status"],
                message=message,
                data=summary,
            )
        )

        AppLogger.log_message("info", message)
        AppLogger.log_message(
            "info",
            (
                "Scan summary: "
                f"status={summary.get('status')}, "
                f"pingable={summary.get('pingable_count', 0)}/{summary.get('total_scanned', 0)}, "
                f"matched={summary.get('matched_count', 0)}, "
                f"inventory_success={summary.get('successful_inventory_count', 0)}, "
                f"inventory_failed={summary.get('failed_inventory_count', 0)}, "
                f"duration_seconds={summary.get('duration_seconds', 0)}"
            ),
        )

    def _persist_inventory_worker_profile(self) -> None:
        if self.inventory_completed <= 0:
            return
        subnet = self._subnet_hint()
        adaptive = self.preferences_manager.get("adaptive_inventory_workers", {})
        if not isinstance(adaptive, dict):
            adaptive = {}
        adaptive[subnet] = {
            "min": int(self.inventory_worker_floor),
            "max": int(max(self.inventory_worker_floor, self.inventory_target_workers)),
            "updated_at": datetime.now().strftime(AppConfig.STORAGE_TIMESTAMP_FORMAT),
        }
        self.preferences_manager.set("adaptive_inventory_workers", adaptive)

    def calculate_summary(self, include_duration: bool = True) -> Dict[str, Any]:
        """
        Calculate scan summary metrics.

        Args:
            include_duration: Include duration_seconds only for final summaries.

        Returns:
            Summary dictionary with separate ping and inventory progress fields.
        """
        ping_total = len(self.expanded_ips)
        ping_completed = self.ping_completed

        inventory_total = max(len(self.computer_records), self.inventory_completed)
        inventory_completed = self.inventory_completed

        summary = {
            "status": AppConfig.SCAN_STATUS_SUCCESS,
            "scan_id": self.scan_id,
            "total_ips": ping_total,
            "pingable_count": len(self.pingable_ips),
            "matched_count": len(self.computer_records),
            "successful_inventory_count": self.inventory_success_count,
            "failed_inventory_count": self.inventory_failed_count,
            "unique_software_titles": self.unique_software_titles,
            "total_software_records": self.total_software_records,
            "inventory_completed_count": inventory_completed,

            # Separate progress channels for the GUI.
            "ping_total": ping_total,
            "ping_completed": ping_completed,
            "ping_active": ping_total > 0 and ping_completed < ping_total,
            "ping_complete": ping_total > 0 and ping_completed >= ping_total,

            "inventory_total": inventory_total,
            "inventory_completed": inventory_completed,
            "inventory_active": (
                inventory_total > 0
                and inventory_completed < inventory_total
            ),
            "inventory_complete": (
                inventory_total > 0
                and inventory_completed >= inventory_total
            ),
        }

        if include_duration:
            summary["duration_seconds"] = round(
                time.perf_counter() - self.started_at,
                2,
            )

        return summary


    def _emit_progress(
        self,
        phase: str,
        total: int = 0,
        completed: int = 0,
        success_count: int = 0,
        warning_count: int = 0,
        error_count: int = 0,
        message: str = "",
        force_feed: bool = False,
        feed_visible: bool = True,
    ) -> None:
        """
        Emit progress with live dashboard data.

        The legacy total/completed fields are preserved for compatibility, but the
        GUI should now prefer ping_* and inventory_* fields from data.
        """
        dashboard_data = self.calculate_summary(include_duration=False)
        dashboard_data["phase"] = phase
        dashboard_data["message"] = message
        dashboard_data["_feed_visible"] = feed_visible or force_feed

        legacy_total = total or max(
            1,
            dashboard_data.get("ping_total", 0)
            + dashboard_data.get("inventory_total", 0),
        )
        legacy_completed = completed or min(
            legacy_total,
            dashboard_data.get("ping_completed", 0)
            + dashboard_data.get("inventory_completed", 0),
        )

        self.result_queue.put(
            ScanProgress(
                phase=phase,
                total=legacy_total,
                completed=legacy_completed,
                success_count=success_count,
                warning_count=warning_count,
                error_count=error_count,
                message=message,
                data=dashboard_data,
            )
        )


    def _progress_total_units(self) -> int:
        """
        Return legacy combined progress units.

        This is retained for older calls, but the GUI now displays separate ping and
        inventory bars.
        """
        return max(
            1,
            len(self.expanded_ips)
            + max(len(self.computer_records), self.inventory_completed),
        )


    def _progress_completed_units(self) -> int:
        """
        Return legacy combined completed units.

        This is retained for older calls, but the GUI now displays separate ping and
        inventory bars.
        """
        return min(
            self._progress_total_units(),
            self.ping_completed + self.inventory_completed,
        )

    def _worker_count(
        self,
        value: Any,
        minimum: int,
        maximum: int,
        default: int,
    ) -> int:
        """
        Validate worker-count input.

        Args:
            value: Raw worker count.
            minimum: Minimum allowed value.
            maximum: Maximum allowed value.
            default: Default fallback.

        Returns:
            Safe worker count.
        """
        try:
            return max(minimum, min(maximum, int(value)))
        except (TypeError, ValueError):
            AppLogger.log_message(
                "warning",
                f"Invalid worker count received; using default {default}.",
            )
            return default

    def _safe_non_negative_int(self, value: Any, default: int) -> int:
        """
        Convert a value to a non-negative integer.

        Args:
            value: Raw value.
            default: Fallback value.

        Returns:
            Non-negative integer.
        """
        try:
            return max(0, int(value))
        except (TypeError, ValueError):
            return default

    def _safe_bounded_int(
        self,
        value: Any,
        minimum: int,
        maximum: int,
        default: int,
    ) -> int:
        """
        Convert a value to a bounded integer.

        Args:
            value: Raw value.
            minimum: Minimum allowed value.
            maximum: Maximum allowed value.
            default: Fallback value.

        Returns:
            Bounded integer.
        """
        try:
            return max(minimum, min(maximum, int(value)))
        except (TypeError, ValueError):
            return default

    def _reset_runtime_state(self) -> None:
        """Reset per-scan runtime state."""
        self.scan_id = None
        self.expanded_ips = []
        self.pending_ping_ips = deque()
        self.active_ping_targets = set()
        self.completed_ping_targets = set()

        self.pingable_ips = []
        self.computer_records = []
        self.inventory_results = []

        self.ping_completed = 0
        self.resolved_count = 0
        self.inventory_completed = 0
        self.inventory_success_count = 0
        self.inventory_failed_count = 0
        self.total_software_records = 0
        self.unique_software_titles = 0
        self._software_keys = set()

        self.ping_retry_count = self.DEFAULT_PING_RETRY_COUNT
        self.max_ping_in_flight = self.DEFAULT_MAX_PING_IN_FLIGHT

        self._last_phase_feed_state = {}
        self._last_dashboard_update_time = 0.0
        self.started_at = 0.0

# ------------------ Sortable Treeview ------------------ #
class SortableTreeview(ttk.Frame):
    """
    Reusable themed table widget with sortable columns, copy support,
    context menu actions, CSV export, and IP-aware sorting.
    """

    def __init__(
        self,
        parent: tk.Widget,
        columns: Optional[Sequence[str]] = None,
        double_click_callback: Optional[Any] = None,
    ):
        """
        Initialize sortable table.

        Args:
            parent: Parent Tkinter widget.
            columns: Optional initial column names.
            double_click_callback: Optional callback for row double-click.
        """
        super().__init__(parent)

        self.columns: List[str] = []
        self.sort_states: Dict[str, bool] = {}
        self.double_click_callback = double_click_callback
        self._last_context_item: Optional[str] = None
        self._last_context_column: Optional[str] = None

        self.tree = ttk.Treeview(self, show="headings", selectmode="extended")
        self.vertical_scrollbar = ttk.Scrollbar(
            self,
            orient="vertical",
            command=self.tree.yview,
        )
        self.horizontal_scrollbar = ttk.Scrollbar(
            self,
            orient="horizontal",
            command=self.tree.xview,
        )

        self.tree.configure(
            yscrollcommand=self.vertical_scrollbar.set,
            xscrollcommand=self.horizontal_scrollbar.set,
        )

        self.tree.grid(row=0, column=0, sticky="nsew")
        self.vertical_scrollbar.grid(row=0, column=1, sticky="ns")
        self.horizontal_scrollbar.grid(row=1, column=0, sticky="ew")

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.context_menu = tk.Menu(self, tearoff=False)
        self.context_menu.add_command(
            label="Copy Cell",
            command=self.copy_selected_cell,
        )
        self.context_menu.add_command(
            label="Copy Row",
            command=self.copy_selected_row,
        )
        self.context_menu.add_command(
            label="Copy Table",
            command=self.copy_table,
        )
        self.context_menu.add_separator()
        self.context_menu.add_command(
            label="Export Table to CSV",
            command=self.export_table_to_csv,
        )

        self.tree.bind("<Button-3>", self._show_context_menu)
        self.tree.bind("<Control-c>", lambda _event: self.copy_selected_row())
        self.tree.bind("<Double-1>", self._handle_double_click)

        if columns:
            self.set_columns(columns)

    def set_columns(self, columns: Sequence[str]) -> None:
        """
        Configure table columns.

        Args:
            columns: Column names to display.
        """
        self.columns = list(columns)
        self.sort_states = {column: False for column in self.columns}
        self.tree.configure(columns=self.columns)

        for column in self.columns:
            self.tree.heading(
                column,
                text=column,
                command=lambda col=column: self.sort_by_column(col),
            )
            self.tree.column(column, width=140, minwidth=50, stretch=True)

    def populate(self, rows: Sequence[Any]) -> None:
        """
        Populate table rows.

        Args:
            rows: Sequence of dictionaries, sqlite rows, objects, or sequences.
        """
        self.clear()

        for row in rows:
            values = self._row_to_values(row)
            self.tree.insert("", "end", values=values)

        self.auto_size_columns()

    def sort_by_column(self, column: str) -> None:
        """
        Sort rows by selected column.

        Args:
            column: Column name.
        """
        if column not in self.columns:
            return

        reverse = self.sort_states.get(column, False)

        items = [
            (
                self.normalize_sort_value(self.tree.set(item_id, column)),
                item_id,
            )
            for item_id in self.tree.get_children("")
        ]

        items.sort(key=lambda item: item[0], reverse=reverse)

        for index, (_value, item_id) in enumerate(items):
            self.tree.move(item_id, "", index)

        self.sort_states[column] = not reverse

    def normalize_sort_value(self, value: Any) -> Tuple[int, Any]:
        """
        Normalize mixed values for safe sorting.

        IP addresses are sorted by their integer address value instead of as
        decimals or simple strings.

        Args:
            value: Raw cell value.

        Returns:
            Sort-safe tuple.
        """
        text = str(value or "").strip()

        if not text:
            return 5, ""

        try:
            return 0, int(ipaddress.ip_address(text))
        except ValueError:
            pass

        percent_text = text[:-1].strip() if text.endswith("%") else text
        try:
            return 1, float(percent_text.replace(",", ""))
        except ValueError:
            pass

        for date_format in (
            AppConfig.STORAGE_TIMESTAMP_FORMAT,
            "%Y-%m-%d",
            "%m/%d/%Y",
        ):
            try:
                return 2, datetime.strptime(text, date_format)
            except ValueError:
                continue

        return 3, text.lower()

    def copy_selected_row(self) -> None:
        """Copy selected table rows to clipboard."""
        selected = self.tree.selection()
        if not selected:
            return

        lines = [
            "\t".join(str(value) for value in self.tree.item(item_id, "values"))
            for item_id in selected
        ]
        self._copy_to_clipboard("\n".join(lines))

    def copy_selected_cell(self) -> None:
        """Copy the right-clicked or focused cell to clipboard."""
        item_id = self._last_context_item or self.tree.focus()
        column = self._last_context_column

        if not item_id:
            return

        values = self.tree.item(item_id, "values")
        column_index = self._column_identifier_to_index(column)

        if column_index is None or column_index >= len(values):
            return

        self._copy_to_clipboard(str(values[column_index]))

    def copy_table(self) -> None:
        """Copy entire table, including headers, to clipboard."""
        lines = ["\t".join(self.columns)]

        for item_id in self.tree.get_children(""):
            values = self.tree.item(item_id, "values")
            lines.append("\t".join(str(value) for value in values))

        self._copy_to_clipboard("\n".join(lines))

    def clear(self) -> None:
        """Clear all table rows."""
        for item_id in self.tree.get_children(""):
            self.tree.delete(item_id)

    def auto_size_columns(self) -> None:
        """Auto-size columns within reasonable limits."""
        for column in self.columns:
            max_width = max(80, len(column) * 9)

            for item_id in self.tree.get_children(""):
                value = self.tree.set(item_id, column)
                max_width = max(max_width, len(str(value)) * 8)

            self.tree.column(column, width=min(max_width, 300))

    def export_table_to_csv(self) -> None:
        """Export visible table content to CSV using a file dialog."""
        output_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=(("CSV Files", "*.csv"), ("All Files", "*.*")),
        )

        if not output_path:
            return

        try:
            with open(output_path, "w", newline="", encoding="utf-8") as file_obj:
                writer = csv.writer(file_obj)
                writer.writerow(self.columns)

                for item_id in self.tree.get_children(""):
                    writer.writerow(self.tree.item(item_id, "values"))

            AppLogger.log_message("info", f"Table exported to CSV: {output_path}")
        except OSError as exc:
            AppLogger.log_message("error", f"CSV export failed: {exc}")

    def _row_to_values(self, row: Any) -> List[Any]:
        """
        Convert supported row object into table values.

        Args:
            row: Row object.

        Returns:
            List of values aligned to configured columns.
        """
        if isinstance(row, sqlite3.Row):
            row = dict(row)

        if isinstance(row, dict):
            return [row.get(column, "") for column in self.columns]

        if hasattr(row, "to_dict"):
            row_dict = row.to_dict()
            return [row_dict.get(column, "") for column in self.columns]

        if isinstance(row, (list, tuple)):
            return list(row)

        return [getattr(row, column, "") for column in self.columns]

    def _show_context_menu(self, event: tk.Event) -> None:
        """
        Show context menu at pointer location.

        Args:
            event: Tkinter mouse event.
        """
        self._last_context_item = self.tree.identify_row(event.y)
        self._last_context_column = self.tree.identify_column(event.x)

        if self._last_context_item:
            self.tree.selection_set(self._last_context_item)

        self.context_menu.tk_popup(event.x_root, event.y_root)

    def _handle_double_click(self, _event: tk.Event) -> None:
        """Handle double-click row action."""
        if not self.double_click_callback:
            return

        selected = self.tree.selection()
        if not selected:
            return

        values = self.tree.item(selected[0], "values")
        self.double_click_callback(values)

    def _column_identifier_to_index(
        self,
        column_identifier: Optional[str],
    ) -> Optional[int]:
        """
        Convert Treeview column identifier to zero-based index.

        Args:
            column_identifier: Treeview identifier like '#1'.

        Returns:
            Zero-based index or None.
        """
        if not column_identifier or not column_identifier.startswith("#"):
            return None

        try:
            return int(column_identifier[1:]) - 1
        except ValueError:
            return None

    def _copy_to_clipboard(self, text: str) -> None:
        """
        Copy text to system clipboard.

        Args:
            text: Text to copy.
        """
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
            AppLogger.log_message("debug", "Copied table data to clipboard.")
        except tk.TclError as exc:
            AppLogger.log_message("error", f"Clipboard copy failed: {exc}")

# ------------------ Scan Control Panel ------------------ #
class ScanControlPanel(ttk.Frame):
    """
    User input panel for configuring scans.

    Ping workers define available thread capacity. Max Active Pings limits
    simultaneous ICMP requests to avoid network throttling and false negatives.
    """

    def __init__(
        self,
        parent: tk.Widget,
        preferences_manager: PreferencesManager,
        start_callback: Any,
        cancel_callback: Any,
        clear_callback: Any,
        export_callback: Any,
        credential_callback: Any,
        show_credentials_button: bool = False,
    ):
        """Initialize scan controls."""
        super().__init__(parent, padding=10)

        self.preferences_manager = preferences_manager
        self.start_callback = start_callback
        self.cancel_callback = cancel_callback
        self.clear_callback = clear_callback
        self.export_callback = export_callback
        self.credential_callback = credential_callback
        self.show_credentials_button = show_credentials_button

        self.ip_range_var = tk.StringVar()
        self.device_filter_var = tk.StringVar()
        self.ping_workers_var = tk.IntVar()
        self.max_ping_in_flight_var = tk.IntVar()
        self.ping_retry_count_var = tk.IntVar()
        self.inventory_workers_var = tk.IntVar()

        self._create_widgets()
        self.restore_preferences(preferences_manager.preferences)
        self.set_scan_running(False)

    def get_scan_options(self) -> Dict[str, Any]:
        """
        Return validated scan options.

        Returns:
            Dictionary with IP range, filters, worker counts, and ping controls.
        """
        ping_workers = self._safe_int(
            self.ping_workers_var.get(),
            AppConfig.MIN_PING_WORKERS,
            AppConfig.MAX_PING_WORKERS,
            AppConfig.DEFAULT_MAX_PING_WORKERS,
        )
        max_ping_in_flight = self._safe_int(
            self.max_ping_in_flight_var.get(),
            1,
            ping_workers,
            min(64, ping_workers),
        )

        return {
            "ip_range": self.ip_range_var.get().strip(),
            "filters": self.device_filter_var.get().strip(),
            "max_ping_workers": ping_workers,
            "max_ping_in_flight": max_ping_in_flight,
            "ping_retry_count": self._safe_int(
                self.ping_retry_count_var.get(),
                0,
                3,
                1,
            ),
            "max_inventory_workers": self._safe_int(
                self.inventory_workers_var.get(),
                AppConfig.MIN_INVENTORY_WORKERS,
                AppConfig.MAX_INVENTORY_WORKERS,
                AppConfig.DEFAULT_MAX_INVENTORY_WORKERS,
            ),
        }

    def set_export_running(self, is_running: bool) -> None:
        """
        Disable export button during export.

        Args:
            is_running: True while export is active.
        """
        state = "disabled" if is_running else "normal"
        label = "Exporting..." if is_running else "Export XLSX"

        self.export_button.configure(state=state, text=label)

    def set_scan_running(self, is_running: bool) -> None:
        """Enable or disable controls based on scan state."""
        normal_state = "disabled" if is_running else "normal"
        scan_state = "disabled" if is_running else "normal"
        cancel_state = "normal" if is_running else "disabled"

        widgets = (
            self.ip_entry,
            self.filter_entry,
            self.ping_worker_spinbox,
            self.max_active_ping_spinbox,
            self.ping_retry_spinbox,
            self.inventory_worker_spinbox,
            self.clear_button,
            self.export_button,
        )

        for widget in widgets:
            widget.configure(state=normal_state)

        if hasattr(self, "credential_button"):
            self.credential_button.configure(state=normal_state)

        self.scan_button.configure(state=scan_state)
        self.cancel_button.configure(state=cancel_state)

    def set_credentials_cached(self, is_cached: bool) -> None:
        """Update scan button appearance based on credential cache state."""
        if is_cached:
            self.scan_button.configure(bg="#D4FAD4", activebackground="#BFEFC0")
            if hasattr(self, "credential_button"):
                self.credential_button.configure(bg="#D4FAD4")
        else:
            self.scan_button.configure(bg="#E6E6E6", activebackground="#DADADA")
            if hasattr(self, "credential_button"):
                self.credential_button.configure(bg="#FFF2CC")

    def restore_preferences(self, preferences: Dict[str, Any]) -> None:
        """Restore saved preferences into the control panel."""
        ping_workers = int(preferences.get(
            "max_ping_workers",
            AppConfig.DEFAULT_MAX_PING_WORKERS,
        ))

        self.ip_range_var.set(preferences.get("last_ip_range", ""))
        self.device_filter_var.set(preferences.get("last_device_filter", ""))
        self.ping_workers_var.set(ping_workers)
        self.max_ping_in_flight_var.set(
            int(preferences.get(
                "max_ping_in_flight",
                min(64, ping_workers),
            ))
        )
        self.ping_retry_count_var.set(
            int(preferences.get("ping_retry_count", 1))
        )
        self.inventory_workers_var.set(
            int(preferences.get(
                "max_inventory_workers",
                AppConfig.DEFAULT_MAX_INVENTORY_WORKERS,
            ))
        )

    def save_current_preferences(self) -> None:
        """Persist current scan control settings."""
        options = self.get_scan_options()
        preferences = dict(self.preferences_manager.preferences)
        preferences["last_ip_range"] = options["ip_range"]
        preferences["last_device_filter"] = options["filters"]
        preferences["max_ping_workers"] = options["max_ping_workers"]
        preferences["max_ping_in_flight"] = options["max_ping_in_flight"]
        preferences["ping_retry_count"] = options["ping_retry_count"]
        preferences["max_inventory_workers"] = options["max_inventory_workers"]
        self.preferences_manager.save_preferences(preferences)

    def _create_widgets(self) -> None:
        """Create and arrange scan control widgets."""
        self.configure(style="App.TFrame")

        input_frame = ttk.Frame(self)
        input_frame.grid(row=0, column=0, sticky="ew")
        input_frame.grid_columnconfigure(1, weight=1)
        input_frame.grid_columnconfigure(3, weight=1)

        ttk.Label(input_frame, text="IP Range:").grid(
            row=0, column=0, sticky="w", padx=(0, 5), pady=3,
        )
        self.ip_entry = ttk.Entry(input_frame, textvariable=self.ip_range_var)
        self.ip_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10), pady=3)

        ttk.Label(input_frame, text="Device Filters:").grid(
            row=0, column=2, sticky="w", padx=(0, 5), pady=3,
        )
        self.filter_entry = ttk.Entry(
            input_frame,
            textvariable=self.device_filter_var,
        )
        self.filter_entry.grid(row=0, column=3, sticky="ew", pady=3)

        self.filter_help_label = ttk.Label(
            input_frame,
            text="Separate filters with commas, semicolons, pipes, or spaces.",
        )
        self.filter_help_label.grid(row=1, column=3, sticky="w", pady=(0, 5))

        options_frame = ttk.Frame(self)
        options_frame.grid(row=1, column=0, sticky="ew", pady=(5, 0))

        if self.show_credentials_button:
            self.credential_button = tk.Button(
                options_frame,
                text="Credentials",
                command=self.credential_callback,
                bg="#FFF2CC",
                activebackground="#FFE699",
                relief="raised",
            )
            self.credential_button.pack(side="left", padx=(0, 12))

        ttk.Label(options_frame, text="Ping Workers:").pack(side="left")
        self.ping_worker_spinbox = ttk.Spinbox(
            options_frame,
            from_=AppConfig.MIN_PING_WORKERS,
            to=AppConfig.MAX_PING_WORKERS,
            textvariable=self.ping_workers_var,
            width=6,
        )
        self.ping_worker_spinbox.pack(side="left", padx=(5, 12))

        ttk.Label(options_frame, text="Max Active Pings:").pack(side="left")
        self.max_active_ping_spinbox = ttk.Spinbox(
            options_frame,
            from_=1,
            to=AppConfig.MAX_PING_WORKERS,
            textvariable=self.max_ping_in_flight_var,
            width=6,
        )
        self.max_active_ping_spinbox.pack(side="left", padx=(5, 12))

        ttk.Label(options_frame, text="Ping Retries:").pack(side="left")
        self.ping_retry_spinbox = ttk.Spinbox(
            options_frame,
            from_=0,
            to=3,
            textvariable=self.ping_retry_count_var,
            width=4,
        )
        self.ping_retry_spinbox.pack(side="left", padx=(5, 12))

        ttk.Label(options_frame, text="Inventory Workers:").pack(side="left")
        self.inventory_worker_spinbox = ttk.Spinbox(
            options_frame,
            from_=AppConfig.MIN_INVENTORY_WORKERS,
            to=AppConfig.MAX_INVENTORY_WORKERS,
            textvariable=self.inventory_workers_var,
            width=6,
        )
        self.inventory_worker_spinbox.pack(side="left", padx=(5, 12))

        button_frame = ttk.Frame(self)
        button_frame.grid(row=2, column=0, sticky="w", pady=(8, 0))

        self.scan_button = tk.Button(
            button_frame,
            text="Scan",
            command=self._handle_scan_click,
            bg="#E6E6E6",
            activebackground="#D4FAD4",
            relief="raised",
            width=10,
        )
        self.scan_button.pack(side="left", padx=(0, 5))

        self.cancel_button = tk.Button(
            button_frame,
            text="Cancel",
            command=self.cancel_callback,
            bg="#FAD4D4",
            activebackground="#F4B6B6",
            relief="raised",
            width=10,
        )
        self.cancel_button.pack(side="left", padx=5)

        self.clear_button = tk.Button(
            button_frame,
            text="Clear",
            command=self.clear_callback,
            bg="#E6E6E6",
            activebackground="#DADADA",
            relief="raised",
            width=10,
        )
        self.clear_button.pack(side="left", padx=5)

        self.export_button = tk.Button(
            button_frame,
            text="Export XLSX",
            command=self.export_callback,
            bg="#D9EAF7",
            activebackground="#C7DFF0",
            relief="raised",
            width=12,
        )
        self.export_button.pack(side="left", padx=5)

        self.grid_columnconfigure(0, weight=1)

    def _handle_scan_click(self) -> None:
        """Validate controls, save preferences, and invoke scan callback."""
        options = self.get_scan_options()
        is_valid, message = ScanInputValidator.validate_ip_range(
            options["ip_range"],
        )

        if not is_valid:
            AppLogger.log_message("error", message)
            return

        self.save_current_preferences()
        AppLogger.log_message(
            "info",
            (
                f"Scan options collected: ping_workers="
                f"{options['max_ping_workers']}, max_active_pings="
                f"{options['max_ping_in_flight']}, ping_retries="
                f"{options['ping_retry_count']}, inventory_workers="
                f"{options['max_inventory_workers']}."
            ),
        )
        self.start_callback(options)

    def _safe_int(
        self,
        value: Any,
        minimum: int,
        maximum: int,
        default: int,
    ) -> int:
        """
        Convert worker count to a safe integer.

        Args:
            value: Raw value.
            minimum: Minimum allowed value.
            maximum: Maximum allowed value.
            default: Fallback value.

        Returns:
            Safe integer.
        """
        try:
            return max(minimum, min(maximum, int(value)))
        except (TypeError, ValueError):
            AppLogger.log_message(
                "warning",
                f"Invalid scan option entered; using default {default}.",
            )
            return default

# ------------------ Scan Progress Panel ------------------ #
class ScanProgressPanel(ttk.Frame):
    """
    Display separate ping and inventory progress bars.

    Each progress row is shown during an active scan and hidden as soon as its
    own phase completes. Rows are reset and shown again at the start of the next
    scan.
    """

    PING_LABEL = "Ping Progress"
    INVENTORY_LABEL = "Inventory Progress"

    def __init__(self, parent: tk.Widget):
        """
        Initialize scan progress panel.

        Args:
            parent: Parent widget.
        """
        super().__init__(parent, style="App.TFrame")

        self.ping_var = tk.IntVar(value=0)
        self.inventory_var = tk.IntVar(value=0)

        self.ping_text_var = tk.StringVar(value="0/0 (0%)")
        self.inventory_text_var = tk.StringVar(value="0/0 (0%)")

        self._ping_visible = False
        self._inventory_visible = False
        self._scan_active = False

        self._create_widgets()
        self.suppress_all()

    def reset_for_scan(self) -> None:
        """Reset progress rows and show them for a new scan."""
        self._scan_active = True

        self.ping_var.set(0)
        self.inventory_var.set(0)
        self.ping_text_var.set("0/0 (0%)")
        self.inventory_text_var.set("0/0 (0%)")

        self._show_ping()
        self._show_inventory()

    def suppress_all(self) -> None:
        """Hide all progress rows."""
        self._scan_active = False
        self._hide_ping()
        self._hide_inventory()

    def update_progress(self, progress: ScanProgress) -> None:
        """
        Update ping and inventory progress from ScanProgress data.

        Args:
            progress: Progress update from scanner pipeline.
        """
        if not self._scan_active and progress.phase not in {"validating", "starting"}:
            self.reset_for_scan()

        progress_data = dict(progress.data or {})

        ping_completed, ping_total = self._extract_ping_progress(
            progress,
            progress_data,
        )
        inventory_completed, inventory_total = self._extract_inventory_progress(
            progress,
            progress_data,
        )

        self._update_ping_progress(ping_completed, ping_total)
        self._update_inventory_progress(inventory_completed, inventory_total)

        if progress.phase in {"complete", "cancelled"}:
            self.suppress_all()

    def _create_widgets(self) -> None:
        """Create progress row widgets."""
        self.grid_columnconfigure(1, weight=1)

        self.ping_label = ttk.Label(
            self,
            text=self.PING_LABEL,
            width=18,
            anchor="w",
        )
        self.ping_bar = ttk.Progressbar(
            self,
            variable=self.ping_var,
            maximum=100,
            mode="determinate",
        )
        self.ping_value_label = ttk.Label(
            self,
            textvariable=self.ping_text_var,
            width=18,
            anchor="e",
        )

        self.inventory_label = ttk.Label(
            self,
            text=self.INVENTORY_LABEL,
            width=18,
            anchor="w",
        )
        self.inventory_bar = ttk.Progressbar(
            self,
            variable=self.inventory_var,
            maximum=100,
            mode="determinate",
        )
        self.inventory_value_label = ttk.Label(
            self,
            textvariable=self.inventory_text_var,
            width=18,
            anchor="e",
        )

    def _extract_ping_progress(
        self,
        progress: ScanProgress,
        progress_data: Dict[str, Any],
    ) -> Tuple[int, int]:
        """
        Extract ping progress values.

        Args:
            progress: ScanProgress update.
            progress_data: Progress data dictionary.

        Returns:
            Tuple of completed and total ping work.
        """
        completed = self._safe_int(
            progress_data.get(
                "ping_completed",
                progress_data.get("ping_completed_count", 0),
            )
        )
        total = self._safe_int(
            progress_data.get(
                "ping_total",
                progress_data.get("expanded_ip_count", 0),
            )
        )

        if total <= 0 and progress.phase in {"ping", "resolving"}:
            completed = self._safe_int(progress.completed)
            total = self._safe_int(progress.total)

        return completed, total

    def _extract_inventory_progress(
        self,
        progress: ScanProgress,
        progress_data: Dict[str, Any],
    ) -> Tuple[int, int]:
        """
        Extract inventory progress values.

        Args:
            progress: ScanProgress update.
            progress_data: Progress data dictionary.

        Returns:
            Tuple of completed and total inventory work.
        """
        completed = self._safe_int(
            progress_data.get(
                "inventory_completed",
                progress_data.get("inventory_completed_count", 0),
            )
        )
        total = self._safe_int(
            progress_data.get(
                "inventory_total",
                progress_data.get(
                    "computer_record_count",
                    progress_data.get("matched_count", 0),
                ),
            )
        )

        if total <= 0 and progress.phase == "inventory":
            completed = self._safe_int(progress.completed)
            total = self._safe_int(progress.total)

        return completed, total

    def _update_ping_progress(self, completed: int, total: int) -> None:
        """
        Update ping progress row.

        Args:
            completed: Completed ping count.
            total: Total ping count.
        """
        if total <= 0:
            return

        percent = self._percent(completed, total)
        self.ping_var.set(percent)
        self.ping_text_var.set(f"{min(completed, total)}/{total} ({percent}%)")

        if completed >= total:
            self._hide_ping()
            return

        self._show_ping()

    def _update_inventory_progress(self, completed: int, total: int) -> None:
        """
        Update inventory progress row.

        Args:
            completed: Completed inventory count.
            total: Total inventory count.
        """
        if total <= 0:
            return

        percent = self._percent(completed, total)
        self.inventory_var.set(percent)
        self.inventory_text_var.set(
            f"{min(completed, total)}/{total} ({percent}%)"
        )

        if completed >= total:
            self._hide_inventory()
            return

        self._show_inventory()

    def _show_ping(self) -> None:
        """Show ping progress row."""
        if self._ping_visible:
            return

        self.ping_label.grid(row=0, column=0, sticky="w", padx=(0, 8))
        self.ping_bar.grid(row=0, column=1, sticky="ew", padx=(0, 8))
        self.ping_value_label.grid(row=0, column=2, sticky="e")
        self._ping_visible = True

    def _hide_ping(self) -> None:
        """Hide ping progress row."""
        if not self._ping_visible:
            return

        self.ping_label.grid_remove()
        self.ping_bar.grid_remove()
        self.ping_value_label.grid_remove()
        self._ping_visible = False

    def _show_inventory(self) -> None:
        """Show inventory progress row."""
        if self._inventory_visible:
            return

        self.inventory_label.grid(row=1, column=0, sticky="w", padx=(0, 8))
        self.inventory_bar.grid(row=1, column=1, sticky="ew", padx=(0, 8))
        self.inventory_value_label.grid(row=1, column=2, sticky="e")
        self._inventory_visible = True

    def _hide_inventory(self) -> None:
        """Hide inventory progress row."""
        if not self._inventory_visible:
            return

        self.inventory_label.grid_remove()
        self.inventory_bar.grid_remove()
        self.inventory_value_label.grid_remove()
        self._inventory_visible = False

    def _percent(self, completed: int, total: int) -> int:
        """
        Calculate bounded percentage.

        Args:
            completed: Completed count.
            total: Total count.

        Returns:
            Integer percentage.
        """
        if total <= 0:
            return 0

        safe_completed = max(0, min(completed, total))
        return int((safe_completed / total) * 100)

    def _safe_int(self, value: Any) -> int:
        """
        Convert value to integer safely.

        Args:
            value: Raw value.

        Returns:
            Integer or zero.
        """
        try:
            return int(value or 0)
        except (TypeError, ValueError):
            return 0

# ------------------ Dashboard Panel ------------------ #
class DashboardPanel(ttk.Frame):
    """
    Prominent summary panel for scan state and inventory counts.

    Duration is updated only when a summary explicitly includes duration_seconds.
    """

    def __init__(self, parent: tk.Widget):
        """Initialize dashboard panel."""
        super().__init__(parent, padding=10)

        self.metric_vars: Dict[str, tk.StringVar] = {
            "total_ips": tk.StringVar(value="0"),
            "pingable_count": tk.StringVar(value="0"),
            "matched_count": tk.StringVar(value="0"),
            "successful_inventory_count": tk.StringVar(value="0"),
            "failed_inventory_count": tk.StringVar(value="0"),
            "unique_software_titles": tk.StringVar(value="0"),
            "total_software_records": tk.StringVar(value="0"),
            "duration_seconds": tk.StringVar(value="0.00s"),
        }
        self.phase_var = tk.StringVar(value="Ready")
        self.message_var = tk.StringVar(value="No scan running.")

        self._create_widgets()

    def update_dashboard(self, summary: Dict[str, Any]) -> None:
        """Update displayed dashboard metrics."""
        metric_map = {
            "total_ips": "total_ips",
            "pingable_count": "pingable_count",
            "matched_count": "matched_count",
            "successful_inventory_count": "successful_inventory_count",
            "failed_inventory_count": "failed_inventory_count",
            "unique_software_titles": "unique_software_titles",
            "total_software_records": "total_software_records",
        }

        for variable_name, summary_key in metric_map.items():
            if summary_key in summary:
                self.metric_vars[variable_name].set(str(summary.get(summary_key, 0)))

        if "duration_seconds" in summary:
            duration = summary.get("duration_seconds", 0)
            self.metric_vars["duration_seconds"].set(f"{float(duration):.2f}s")

        phase = summary.get("phase") or summary.get("status")
        message = summary.get("message")

        if phase:
            self.phase_var.set(str(phase).title())

        if message:
            self.message_var.set(str(message))

    def clear(self) -> None:
        """Reset dashboard metrics to defaults."""
        for key, variable in self.metric_vars.items():
            variable.set("0.00s" if key == "duration_seconds" else "0")

        self.phase_var.set("Ready")
        self.message_var.set("No scan running.")

    def _create_widgets(self) -> None:
        """Create dashboard layout."""
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(
            header_frame,
            text="Dashboard",
            font=("Arial", 14, "bold"),
        ).pack(side="left")

        ttk.Label(
            header_frame,
            textvariable=self.phase_var,
            font=("Arial", 11, "bold"),
        ).pack(side="right")

        self.cards_frame = ttk.Frame(self)
        self.cards_frame.pack(fill="x")

        cards = (
            ("IPs Scanned", "total_ips", "#D9EAF7"),
            ("Pingable", "pingable_count", "#D4FAD4"),
            ("Matching", "matched_count", "#D9EAF7"),
            ("Inventory Success", "successful_inventory_count", "#D4FAD4"),
            ("Inventory Failed", "failed_inventory_count", "#FAD4D4"),
            ("Unique Software", "unique_software_titles", "#FFF2CC"),
            ("Software Records", "total_software_records", "#D9EAF7"),
            ("Duration", "duration_seconds", "#E6E6E6"),
        )

        for index, (label, key, color) in enumerate(cards):
            card = tk.Frame(
                self.cards_frame,
                bg=color,
                padx=12,
                pady=8,
                relief="groove",
                bd=1,
            )
            card.grid(
                row=index // 4,
                column=index % 4,
                sticky="nsew",
                padx=5,
                pady=5,
            )

            tk.Label(
                card,
                text=label,
                bg=color,
                font=("Arial", 9, "bold"),
            ).pack(anchor="w")

            tk.Label(
                card,
                textvariable=self.metric_vars[key],
                bg=color,
                font=("Arial", 16, "bold"),
            ).pack(anchor="w")

        for column in range(4):
            self.cards_frame.grid_columnconfigure(column, weight=1)

        self.message_label = ttk.Label(
            self,
            textvariable=self.message_var,
            anchor="w",
        )
        self.message_label.pack(fill="x", pady=(10, 0))

# ------------------ Software Summary Panel ------------------ #
class SoftwareSummaryPanel(ttk.Frame):
    """
    Display software occurrence summary with publisher grouping.

    The Outdated Software checkbox controls display mode:
        - Checked: version-specific rows are shown and outdated rows are marked.
        - Unchecked: rows are grouped by software name/publisher, and multiple
          versions are displayed as a version range.

    Double-clicking a software row sends a software key to ResultsNotebook.
    """

    FILTER_DEBOUNCE_MS = 175

    def __init__(
        self,
        parent: tk.Widget,
        selection_callback: Optional[Any] = None,
    ):
        """
        Initialize software summary panel.

        Args:
            parent: Parent widget.
            selection_callback: Callback for selected software.
        """
        super().__init__(parent, padding=10)

        self.selection_callback = selection_callback
        self.grouping_analyzer = SoftwareGroupingAnalyzer()

        self.raw_rows: List[Dict[str, Any]] = []
        self.display_rows: List[Dict[str, Any]] = []
        self.row_lookup: Dict[str, Dict[str, Any]] = {}

        self.search_var = tk.StringVar()
        self.outdated_software_var = tk.BooleanVar(value=True)
        self.summary_var = tk.StringVar(value="Software Summary: 0")

        self._filter_after_id: Optional[str] = None
        self._suspend_traces = False

        self._create_widgets()

        self.search_var.trace_add(
            "write",
            lambda *_args: self._schedule_filter(),
        )
        self.outdated_software_var.trace_add(
            "write",
            lambda *_args: self._handle_outdated_toggle(),
        )

    def populate(
        self,
        rows: Sequence[Any],
        successful_inventory_count: int,
    ) -> None:
        """
        Populate software summary rows.

        Args:
            rows: Source software occurrence rows.
            successful_inventory_count: Number of inventoried computers.
        """
        self.raw_rows = [
            self._normalize_row(row, successful_inventory_count)
            for row in rows
        ]
        self._rebuild_display_rows()

    def clear(self) -> None:
        """Clear software summary table."""
        self._cancel_filter_callback()
        self.raw_rows = []
        self.display_rows = []
        self.row_lookup.clear()
        self.summary_var.set("Software Summary: 0")
        self.tree.delete(*self.tree.get_children())

    def _create_widgets(self) -> None:
        """Create software summary widgets."""
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(
            header_frame,
            text="Software Summary",
            font=("Arial", 14, "bold"),
        ).pack(side="left")

        ttk.Label(
            header_frame,
            textvariable=self.summary_var,
            font=("Arial", 10, "bold"),
        ).pack(side="right")

        filter_frame = ttk.Frame(self)
        filter_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(filter_frame, text="Search:").pack(side="left")

        self.search_entry = ttk.Entry(
            filter_frame,
            textvariable=self.search_var,
        )
        self.search_entry.pack(side="left", fill="x", expand=True, padx=(5, 12))

        self.outdated_checkbutton = ttk.Checkbutton(
            filter_frame,
            text="Outdated Software",
            variable=self.outdated_software_var,
        )
        self.outdated_checkbutton.pack(side="right")

        columns = (
            "Software",
            "Version",
            "Publisher",
            "Occurrences",
            "GroupedCount",
            "Percent",
            "Outdated",
        )

        self.tree = ttk.Treeview(
            self,
            columns=columns,
            show="tree headings",
            selectmode="browse",
        )
        self.tree.pack(fill="both", expand=True)

        self.tree.heading("#0", text="Publisher Group")
        self.tree.column("#0", width=220, anchor="w")

        column_settings = {
            "Software": (330, "w"),
            "Version": (150, "w"),
            "Publisher": (260, "w"),
            "Occurrences": (110, "e"),
            "GroupedCount": (140, "e"),
            "Percent": (110, "e"),
            "Outdated": (90, "center"),
        }

        for column_name, settings in column_settings.items():
            width, anchor = settings
            self.tree.heading(column_name, text=column_name)
            self.tree.column(column_name, width=width, anchor=anchor)

        vertical_scrollbar = ttk.Scrollbar(
            self,
            orient="vertical",
            command=self.tree.yview,
        )
        self.tree.configure(yscrollcommand=vertical_scrollbar.set)
        vertical_scrollbar.place(relx=1.0, rely=0.14, relheight=0.86, anchor="ne")

        self.tree.tag_configure("outdated", background="#FAD4D4")
        self.tree.tag_configure("current", background="#D4FAD4")
        self.tree.tag_configure("grouped", background="#FFF2CC")

        self.tree.bind("<Double-1>", self._handle_double_click)

    def _normalize_row(
        self,
        row: Any,
        successful_inventory_count: int,
    ) -> Dict[str, Any]:
        """
        Normalize source row into panel row.

        Args:
            row: Source row.
            successful_inventory_count: Successful inventory denominator.

        Returns:
            Normalized row dictionary.
        """
        if isinstance(row, sqlite3.Row):
            row_dict = dict(row)
        elif isinstance(row, dict):
            row_dict = row
        else:
            row_dict = dict(row)

        occurrence_count = int(row_dict.get("occurrence_count", 0) or 0)
        grouped_count = int(
            row_dict.get(
                "grouped_version_occurrence_count",
                occurrence_count,
            ) or 0
        )
        percent = 0.0

        if successful_inventory_count > 0:
            percent = min(
                100.0,
                (occurrence_count / successful_inventory_count) * 100,
            )

        return {
            "DisplayName": row_dict.get("display_name", ""),
            "DisplayVersion": row_dict.get("display_version", ""),
            "Publisher": row_dict.get("publisher", ""),
            "OccurrenceCount": occurrence_count,
            "GroupedVersionOccurrenceCount": grouped_count,
            "PercentOfInventoriedComputers": f"{percent:.2f}%",
            "Outdated": False,
            "GroupedVersions": False,
            "GroupedSourceRows": [],
        }

    def _rebuild_display_rows(self) -> None:
        """Rebuild display rows based on Outdated Software checkbox."""
        group_versions = not bool(self.outdated_software_var.get())
        self.display_rows = self.grouping_analyzer.enrich_rows(
            self.raw_rows,
            group_versions=group_versions,
        )
        self._schedule_filter(delay_ms=1)

    def _schedule_filter(self, delay_ms: int = FILTER_DEBOUNCE_MS) -> None:
        """
        Debounce filter application.

        Args:
            delay_ms: Delay in milliseconds.
        """
        if self._suspend_traces:
            return

        self._cancel_filter_callback()
        self._filter_after_id = self.after(delay_ms, self._apply_filter)

    def _apply_filter(self) -> None:
        """Apply search filter and rebuild Treeview."""
        self._filter_after_id = None
        search_text = self.search_var.get().strip().lower()

        if search_text:
            rows = [
                row for row in self.display_rows
                if search_text in self._row_search_text(row)
            ]
        else:
            rows = list(self.display_rows)

        self._populate_tree(rows)

    def _populate_tree(self, rows: Sequence[Dict[str, Any]]) -> None:
        """
        Populate Treeview from filtered rows.

        Args:
            rows: Rows to display.
        """
        self.tree.delete(*self.tree.get_children())
        self.row_lookup.clear()

        grouped_by_publisher: Dict[str, List[Dict[str, Any]]] = {}

        for row in rows:
            publisher_category = row.get("PublisherCategory", "Unknown Publisher")
            grouped_by_publisher.setdefault(publisher_category, []).append(row)

        total_rows = 0

        for publisher_category in sorted(grouped_by_publisher):
            publisher_rows = grouped_by_publisher[publisher_category]
            occurrence_summary = self.grouping_analyzer.summarize_occurrence_range(
                publisher_rows
            )

            publisher_item = self.tree.insert(
                "",
                "end",
                text=publisher_category,
                values=(
                    "",
                    "",
                    "",
                    occurrence_summary,
                    occurrence_summary,
                    "",
                    "",
                ),
                open=False,
            )

            self.row_lookup[publisher_item] = {
                "PublisherCategory": publisher_category,
                "DisplayName": "",
                "DisplayVersion": "",
                "Publisher": publisher_category,
                "OccurrenceCount": occurrence_summary,
                "IsPublisherRow": True,
            }

            for row in publisher_rows:
                total_rows += 1
                tags = self._row_tags(row)
                item_id = self.tree.insert(
                    publisher_item,
                    "end",
                    text="",
                    values=(
                        row.get("DisplayName", ""),
                        row.get("DisplayVersion", ""),
                        row.get("Publisher", ""),
                        row.get("OccurrenceCount", 0),
                        row.get("GroupedVersionOccurrenceCount", 0),
                        row.get("PercentOfInventoriedComputers", ""),
                        "Yes" if row.get("Outdated") else "No",
                    ),
                    tags=tags,
                )
                self.row_lookup[item_id] = row

        mode_label = (
            "Version-specific"
            if self.outdated_software_var.get()
            else "Grouped by software"
        )
        self.summary_var.set(
            f"Software Summary: {total_rows} | Mode: {mode_label}"
        )

    def _handle_outdated_toggle(self) -> None:
        """Handle Outdated Software checkbox changes."""
        if self._suspend_traces:
            return

        self._rebuild_display_rows()

    def _handle_double_click(self, _event: tk.Event) -> None:
        """Follow selected software to Computer Breakout tab on double click."""
        selected = self.tree.selection()

        if not selected:
            return

        row = self.row_lookup.get(selected[0])

        if not row or row.get("IsPublisherRow"):
            return

        if not self.selection_callback:
            return

        software_key = {
            "publisher_category": row.get("PublisherCategory", ""),
            "display_name": row.get("DisplayName", ""),
            "display_version": row.get("DisplayVersion", ""),
            "publisher": row.get("Publisher", ""),
            "grouped_versions": bool(row.get("GroupedVersions")),
            "outdated_mode": bool(self.outdated_software_var.get()),
            "lowest_version": row.get("LowestVersion", ""),
            "highest_version": row.get("HighestVersion", ""),
        }
        self.selection_callback(software_key)

    def _row_tags(self, row: Dict[str, Any]) -> Tuple[str, ...]:
        """
        Build Treeview tags for a row.

        Args:
            row: Display row.

        Returns:
            Treeview tags.
        """
        if row.get("GroupedVersions"):
            return ("grouped",)

        if row.get("Outdated"):
            return ("outdated",)

        return ("current",)

    def _row_search_text(self, row: Dict[str, Any]) -> str:
        """
        Build searchable text for a row.

        Args:
            row: Software row.

        Returns:
            Lowercase searchable text.
        """
        values = (
            row.get("PublisherCategory", ""),
            row.get("DisplayName", ""),
            row.get("DisplayVersion", ""),
            row.get("Publisher", ""),
            row.get("OccurrenceCount", ""),
            row.get("GroupedVersionOccurrenceCount", ""),
        )
        return " ".join(str(value).lower() for value in values)

    def _cancel_filter_callback(self) -> None:
        """Cancel pending filter callback."""
        if not self._filter_after_id:
            return

        try:
            self.after_cancel(self._filter_after_id)
        except tk.TclError:
            pass

        self._filter_after_id = None

# ------------------ Installed Software Panel ------------------ #
class InstalledSoftwarePanel(ttk.Frame):
    """
    Display software installed on a selected computer.

    The table includes Software, Version, Publisher, and Install Date columns.
    """

    def __init__(
        self,
        parent: tk.Widget,
        detail_callback: Optional[Any] = None,
    ):
        """
        Initialize installed software panel.

        Args:
            parent: Parent Tkinter widget.
            detail_callback: Called when a software row is double-clicked.
        """
        super().__init__(parent, padding=10)

        self.detail_callback = detail_callback
        self.all_rows: List[Dict[str, Any]] = []
        self.search_var = tk.StringVar()
        self.summary_var = tk.StringVar(value="Installed Software: 0")

        self._create_widgets()
        self.search_var.trace_add("write", lambda *_args: self._apply_filter())

    def populate_installed_software(
        self,
        rows: Sequence[Any],
        computer_label: str = "",
    ) -> None:
        """
        Populate installed software rows.

        Args:
            rows: Software rows from the database.
            computer_label: Optional selected computer label.
        """
        self.all_rows = [self._normalize_row(row) for row in rows]

        label = f"Installed Software: {len(self.all_rows)}"
        if computer_label:
            label = f"{label} on {computer_label}"

        self.summary_var.set(label)
        self._apply_filter()

    def clear(self) -> None:
        """Clear installed software results."""
        self.all_rows = []
        self.search_var.set("")
        self.summary_var.set("Installed Software: 0")
        self.table.clear()

    def _create_widgets(self) -> None:
        """Create installed software layout."""
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(
            header_frame,
            text="Installed Software",
            font=("Arial", 14, "bold"),
        ).pack(side="left")

        ttk.Label(
            header_frame,
            textvariable=self.summary_var,
            font=("Arial", 10, "bold"),
        ).pack(side="right")

        filter_frame = ttk.Frame(self)
        filter_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(filter_frame, text="Search:").pack(side="left")
        self.search_entry = ttk.Entry(
            filter_frame,
            textvariable=self.search_var,
        )
        self.search_entry.pack(side="left", fill="x", expand=True, padx=(5, 0))

        self.table = SortableTreeview(
            self,
            columns=(
                "Software",
                "Version",
                "Publisher",
                "Install Date",
            ),
            double_click_callback=self._handle_double_click,
        )
        self.table.pack(fill="both", expand=True)

    def _apply_filter(self) -> None:
        """Filter visible software rows by search text."""
        search_text = self.search_var.get().strip().lower()

        visible_rows = self.all_rows
        if search_text:
            visible_rows = [
                row for row in self.all_rows
                if search_text in " ".join(
                    str(value).lower() for value in row.values()
                )
            ]

        self.table.populate(visible_rows)

    def _normalize_row(self, row: Any) -> Dict[str, Any]:
        """
        Normalize source row into table-friendly dictionary.

        Args:
            row: Source software row.

        Returns:
            Normalized row dictionary.
        """
        if isinstance(row, sqlite3.Row):
            row_dict = dict(row)
        elif isinstance(row, dict):
            row_dict = row
        else:
            row_dict = {
                "display_name": getattr(row, "display_name", ""),
                "display_version": getattr(row, "display_version", ""),
                "publisher": getattr(row, "publisher", ""),
                "install_date": getattr(row, "install_date", ""),
            }

        return {
            "Software": row_dict.get("display_name", ""),
            "Version": row_dict.get("display_version", ""),
            "Publisher": row_dict.get("publisher", ""),
            "Install Date": row_dict.get("install_date", ""),
            "_SoftwareId": row_dict.get("software_id", ""),
        }

    def _handle_double_click(self, values: Sequence[Any]) -> None:
        """
        Show installed software details when a row is double-clicked.

        Args:
            values: Selected row values.
        """
        if not self.detail_callback:
            return

        selected = self.table.tree.selection()
        if not selected:
            return

        selected_row = self.table.rows_by_item.get(selected[0], {})
        software_id = selected_row.get("_SoftwareId", "")
        self.detail_callback(software_id)

# ------------------ Computer Inventory Panel ------------------ #
class ComputerInventoryPanel(ttk.Frame):
    """
    Display inventoried computers with search, sorting, and double-click
    navigation behavior.

    A single click selects only. A double click opens the installed software
    table for the selected computer.
    """

    def __init__(
        self,
        parent: tk.Widget,
        selection_callback: Optional[Any] = None,
        detail_callback: Optional[Any] = None,
    ):
        """
        Initialize computer inventory panel.

        Args:
            parent: Parent Tkinter widget.
            selection_callback: Called when a computer row is double-clicked.
            detail_callback: Called when details are requested.
        """
        super().__init__(parent, padding=10)

        self.selection_callback = selection_callback
        self.detail_callback = detail_callback
        self.all_rows: List[Dict[str, Any]] = []

        self.search_var = tk.StringVar()
        self.summary_var = tk.StringVar(value="Computers: 0")

        self._create_widgets()
        self.search_var.trace_add("write", lambda *_args: self._apply_filter())

    def populate_computers(self, rows: Sequence[Any]) -> None:
        """
        Populate computer inventory rows.

        Args:
            rows: Computer rows from database or scanner.
        """
        self.all_rows = [self._normalize_row(row) for row in rows]
        self.summary_var.set(f"Computers: {len(self.all_rows)}")
        self._apply_filter()

    def clear(self) -> None:
        """Clear computer inventory results."""
        self.all_rows = []
        self.search_var.set("")
        self.summary_var.set("Computers: 0")
        self.table.clear()

    def _create_widgets(self) -> None:
        """Create computer inventory layout."""
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(
            header_frame,
            text="Computers",
            font=("Arial", 14, "bold"),
        ).pack(side="left")

        ttk.Label(
            header_frame,
            textvariable=self.summary_var,
            font=("Arial", 10, "bold"),
        ).pack(side="right")

        filter_frame = ttk.Frame(self)
        filter_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(filter_frame, text="Search:").pack(side="left")
        self.search_entry = ttk.Entry(
            filter_frame,
            textvariable=self.search_var,
        )
        self.search_entry.pack(side="left", fill="x", expand=True, padx=(5, 0))

        self.table = SortableTreeview(
            self,
            columns=(
                "ComputerId",
                "Hostname",
                "IPAddress",
                "InventoryStatus",
                "SoftwareCount",
                "LastSeen",
            ),
            double_click_callback=self._handle_double_click,
        )
        self.table.pack(fill="both", expand=True)

        # Selection no longer changes tabs.
        self.table.tree.bind("<<TreeviewSelect>>", self._handle_selection)

        self._configure_status_tags()

    def _apply_filter(self) -> None:
        """Filter visible rows by search text."""
        search_text = self.search_var.get().strip().lower()

        visible_rows = self.all_rows
        if search_text:
            visible_rows = [
                row for row in self.all_rows
                if search_text in " ".join(
                    str(value).lower() for value in row.values()
                )
            ]

        self.table.clear()

        for row in visible_rows:
            values = [row.get(column, "") for column in self.table.columns]
            item_id = self.table.tree.insert("", "end", values=values)
            self._apply_status_tag(item_id, row.get("InventoryStatus", ""))

        self.table.auto_size_columns()

    def _normalize_row(self, row: Any) -> Dict[str, Any]:
        """
        Normalize source row into table-friendly dictionary.

        Args:
            row: Source row object.

        Returns:
            Normalized row dictionary.
        """
        if isinstance(row, sqlite3.Row):
            row_dict = dict(row)
        elif isinstance(row, dict):
            row_dict = row
        elif hasattr(row, "to_dict"):
            row_dict = row.to_dict()
        else:
            row_dict = {
                "computer_id": getattr(row, "computer_id", ""),
                "hostname": getattr(row, "hostname", ""),
                "ip_address": getattr(row, "ip_address", ""),
                "inventory_status": getattr(row, "inventory_status", ""),
                "software_count": getattr(row, "software_count", 0),
                "last_seen": getattr(row, "last_seen", ""),
            }

        return {
            "ComputerId": row_dict.get("computer_id", ""),
            "Hostname": row_dict.get("hostname", ""),
            "IPAddress": row_dict.get("ip_address", ""),
            "InventoryStatus": row_dict.get("inventory_status", ""),
            "SoftwareCount": row_dict.get("software_count", 0),
            "LastSeen": row_dict.get("last_seen", ""),
        }

    def _configure_status_tags(self) -> None:
        """Configure status color tags."""
        self.table.tree.tag_configure(
            AppConfig.INVENTORY_STATUS_FAILED,
            background="#FAD4D4",
        )
        self.table.tree.tag_configure(
            AppConfig.INVENTORY_STATUS_PARTIAL,
            background="#FFF2CC",
        )
        self.table.tree.tag_configure(
            AppConfig.INVENTORY_STATUS_SUCCESS,
            background="#D4FAD4",
        )

    def _apply_status_tag(self, item_id: str, status: str) -> None:
        """
        Apply color tag based on inventory status.

        Args:
            item_id: Treeview item ID.
            status: Inventory status.
        """
        normalized_status = str(status).strip().lower()

        if normalized_status in {
            AppConfig.INVENTORY_STATUS_SUCCESS,
            AppConfig.INVENTORY_STATUS_PARTIAL,
            AppConfig.INVENTORY_STATUS_FAILED,
        }:
            self.table.tree.item(item_id, tags=(normalized_status,))

    def _handle_selection(self, _event: tk.Event) -> None:
        """Select row only; do not navigate tabs on single click."""
        selected = self.table.tree.selection()

        if not selected:
            return

        values = self.table.tree.item(selected[0], "values")
        hostname = values[1] if len(values) > 1 else ""
        AppLogger.log_message("debug", f"Selected computer: {hostname}")

    def _handle_double_click(self, values: Sequence[Any]) -> None:
        """
        Follow selected computer to Installed Software tab on double click.

        Args:
            values: Selected row values.
        """
        computer_id = values[0] if values else ""

        if self.selection_callback:
            self.selection_callback(computer_id)

# ------------------ Computer Breakout Panel ------------------ #
class ComputerBreakoutPanel(ttk.Frame):
    """
    Display all successfully inventoried computers for a selected software item.

    Installed computers are highlighted green. Computers where the selected
    software was not found are highlighted red. The Group Versions control is
    hidden when Software Summary is already in grouped mode because Computer
    Breakout must ignore versions for that selection.
    """

    RENDER_CHUNK_SIZE = 150
    SEARCH_DEBOUNCE_MS = 200
    GROUP_RELOAD_DEBOUNCE_MS = 300

    def __init__(
        self,
        parent: tk.Widget,
        reload_callback: Optional[Any] = None,
        selection_callback: Optional[Any] = None,
    ):
        """
        Initialize computer breakout panel.

        Args:
            parent: Parent Tkinter widget.
            reload_callback: Called when Group Versions changes.
            selection_callback: Called when a computer row is double-clicked.
        """
        super().__init__(parent, padding=10)

        self.reload_callback = reload_callback
        self.selection_callback = selection_callback

        self.all_rows: List[Dict[str, Any]] = []
        self.current_software_key: Dict[str, Any] = {}

        self.search_var = tk.StringVar()
        self.group_versions_var = tk.BooleanVar(value=False)
        self.summary_var = tk.StringVar(value="Computer Breakout: 0")
        self.software_title_var = tk.StringVar(value="No software selected")
        self.group_versions_note_var = tk.StringVar(value="")

        self._search_after_id: Optional[str] = None
        self._group_after_id: Optional[str] = None
        self._render_after_id: Optional[str] = None
        self._render_token = 0
        self._suspend_traces = False
        self._is_busy = False
        self._selected_is_already_grouped = False

        self._create_widgets()

        self.search_var.trace_add(
            "write",
            lambda *_args: self._schedule_filter(),
        )
        self.group_versions_var.trace_add(
            "write",
            lambda *_args: self._schedule_group_versions_reload(),
        )

    def populate_breakout(
        self,
        rows: Sequence[Any],
        software_key: Dict[str, Any],
    ) -> None:
        """
        Populate computer breakout rows.

        Args:
            rows: Computer breakout rows.
            software_key: Selected software identity.
        """
        self.current_software_key = dict(software_key or {})
        self._selected_is_already_grouped = bool(
            self.current_software_key.get("grouped_versions")
        )

        # Always reset Group Versions for non-grouped software selections so
        # each Computer Breakout view starts unchecked unless the source row is
        # already grouped by the Software Summary mode.
        if not self._selected_is_already_grouped and self.group_versions_var.get():
            self._suspend_traces = True
            try:
                self.group_versions_var.set(False)
            finally:
                self._suspend_traces = False

        self._configure_group_versions_visibility()
        self.all_rows = [self._normalize_row(row) for row in rows]

        installed_count = sum(
            1 for row in self.all_rows
            if row.get("Installed") == "Yes"
        )
        missing_count = len(self.all_rows) - installed_count

        name = self.current_software_key.get("display_name", "")
        version = self.current_software_key.get("display_version", "")
        publisher = self.current_software_key.get("publisher", "")

        if self._selected_is_already_grouped:
            version_text = f" | {version}" if version else ""
        elif self.group_versions_var.get():
            version_text = " | All Versions"
        else:
            version_text = f" | {version}" if version else ""

        self.software_title_var.set(
            f"{name}{version_text} | {publisher}".strip(" |")
        )
        self.summary_var.set(
            f"Inventoried Computers: {len(self.all_rows)} | "
            f"Installed: {installed_count} | Not Installed: {missing_count}"
        )

        self.set_busy(False)
        self._schedule_filter(delay_ms=1)

    def clear(self) -> None:
        """Clear computer breakout results."""
        self._cancel_pending_work()

        self._suspend_traces = True
        try:
            self.all_rows = []
            self.current_software_key = {}
            self._selected_is_already_grouped = False
            self.search_var.set("")
            self.group_versions_var.set(False)
            self.summary_var.set("Computer Breakout: 0")
            self.software_title_var.set("No software selected")
            self.group_versions_note_var.set("")
            self.group_versions_checkbutton.pack(side="right")
            self.group_versions_note_label.pack_forget()
            self.table.clear()
        finally:
            self._suspend_traces = False

    def set_busy(self, is_busy: bool, message: str = "Loading breakout...") -> None:
        """
        Set busy state for expensive reloads.

        Args:
            is_busy: True while a reload is running.
            message: Busy message.
        """
        self._is_busy = is_busy

        if is_busy:
            self.summary_var.set(message)
            self.group_versions_checkbutton.configure(state="disabled")
            self.search_entry.configure(state="disabled")
            return

        self.search_entry.configure(state="normal")

        if not self._selected_is_already_grouped:
            self.group_versions_checkbutton.configure(state="normal")

    def reset_for_new_selection(self, software_key: Dict[str, Any]) -> bool:
        """
        Prepare the panel for a newly selected software row.

        Args:
            software_key: Selected software identity.

        Returns:
            True when breakout should query ignoring versions.
        """
        self.current_software_key = dict(software_key or {})
        self._selected_is_already_grouped = bool(
            self.current_software_key.get("grouped_versions")
        )

        self._suspend_traces = True
        try:
            if self._selected_is_already_grouped:
                self.group_versions_var.set(True)
            else:
                self.group_versions_var.set(False)
        finally:
            self._suspend_traces = False

        self._configure_group_versions_visibility()
        return self.should_group_versions()

    def should_group_versions(self) -> bool:
        """
        Return whether breakout matching should ignore versions.

        Returns:
            True when version should be ignored.
        """
        return (
            self._selected_is_already_grouped
            or bool(self.group_versions_var.get())
        )

    def _create_widgets(self) -> None:
        """Create computer breakout layout."""
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(
            header_frame,
            text="Computer Breakout",
            font=("Arial", 14, "bold"),
        ).pack(side="left")

        ttk.Label(
            header_frame,
            textvariable=self.summary_var,
            font=("Arial", 10, "bold"),
        ).pack(side="right")

        selected_frame = ttk.Frame(self)
        selected_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(
            selected_frame,
            text="Selected Software:",
            font=("Arial", 10, "bold"),
        ).pack(side="left")

        ttk.Label(
            selected_frame,
            textvariable=self.software_title_var,
            font=("Arial", 10),
        ).pack(side="left", padx=(6, 0))

        filter_frame = ttk.Frame(self)
        filter_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(filter_frame, text="Search:").pack(side="left")
        self.search_entry = ttk.Entry(
            filter_frame,
            textvariable=self.search_var,
        )
        self.search_entry.pack(side="left", fill="x", expand=True, padx=(5, 12))

        self.group_versions_note_label = ttk.Label(
            filter_frame,
            textvariable=self.group_versions_note_var,
            font=("Arial", 9, "italic"),
        )

        self.group_versions_checkbutton = ttk.Checkbutton(
            filter_frame,
            text="Group Versions",
            variable=self.group_versions_var,
        )
        self.group_versions_checkbutton.pack(side="right")

        self.table = SortableTreeview(
            self,
            columns=(
                "Installed",
                "ComputerId",
                "Hostname",
                "IPAddress",
                "FQDN",
                "Domain",
                "Pingable",
                "MatchedFilter",
                "InventoryStatus",
                "SoftwareCount",
                "LastSeen",
            ),
            double_click_callback=self._handle_double_click,
        )
        self.table.pack(fill="both", expand=True)

        self.table.tree.tag_configure("installed_yes", background="#D4FAD4")
        self.table.tree.tag_configure("installed_no", background="#FAD4D4")

    def _configure_group_versions_visibility(self) -> None:
        """Show or hide Group Versions controls based on selected row mode."""
        if self._selected_is_already_grouped:
            self.group_versions_checkbutton.pack_forget()
            self.group_versions_note_var.set(
                "Versions are already grouped by the Software Summary view."
            )
            self.group_versions_note_label.pack(side="right")
            return

        self.group_versions_note_label.pack_forget()
        self.group_versions_note_var.set("")

        if not self.group_versions_checkbutton.winfo_manager():
            self.group_versions_checkbutton.pack(side="right")

    def _schedule_filter(self, delay_ms: int = SEARCH_DEBOUNCE_MS) -> None:
        """
        Debounce search/filter rendering.

        Args:
            delay_ms: Delay in milliseconds.
        """
        if self._suspend_traces or self._is_busy:
            return

        if self._search_after_id:
            self.after_cancel(self._search_after_id)

        self._search_after_id = self.after(delay_ms, self._apply_filter)

    def _apply_filter(self) -> None:
        """Filter visible breakout rows by search text."""
        self._search_after_id = None
        search_text = self.search_var.get().strip().lower()

        if search_text:
            visible_rows = [
                row for row in self.all_rows
                if search_text in row.get("_SearchText", "")
            ]
        else:
            visible_rows = list(self.all_rows)

        self._render_rows_chunked(visible_rows)

    def _render_rows_chunked(self, rows: Sequence[Dict[str, Any]]) -> None:
        """
        Render rows in small chunks to avoid blocking the GUI.

        Args:
            rows: Visible rows.
        """
        if self._render_after_id:
            self.after_cancel(self._render_after_id)
            self._render_after_id = None

        self._render_token += 1
        render_token = self._render_token
        rows_to_render = list(rows)

        self.table.clear()

        self._render_next_chunk(
            rows_to_render,
            start_index=0,
            render_token=render_token,
        )

    def _render_next_chunk(
        self,
        rows: List[Dict[str, Any]],
        start_index: int,
        render_token: int,
    ) -> None:
        """
        Render the next chunk of rows.

        Args:
            rows: Rows to render.
            start_index: Starting row index.
            render_token: Token used to cancel stale renders.
        """
        if render_token != self._render_token:
            return

        end_index = min(start_index + self.RENDER_CHUNK_SIZE, len(rows))

        for row in rows[start_index:end_index]:
            values = [row.get(column, "") for column in self.table.columns]
            tag = "installed_yes" if row.get("Installed") == "Yes" else "installed_no"
            self.table.tree.insert("", "end", values=values, tags=(tag,))

        if end_index < len(rows):
            self._render_after_id = self.after(
                1,
                lambda: self._render_next_chunk(
                    rows,
                    end_index,
                    render_token,
                ),
            )
            return

        self._render_after_id = None
        self.table.auto_size_columns()

    def _normalize_row(self, row: Any) -> Dict[str, Any]:
        """
        Normalize source row into breakout table row.

        Args:
            row: Source row.

        Returns:
            Normalized row.
        """
        if isinstance(row, sqlite3.Row):
            row_dict = dict(row)
        elif isinstance(row, dict):
            row_dict = row
        else:
            row_dict = dict(row)

        normalized = {
            "Installed": row_dict.get("installed", "No"),
            "ComputerId": row_dict.get("computer_id", ""),
            "Hostname": row_dict.get("hostname", ""),
            "IPAddress": row_dict.get("ip_address", ""),
            "FQDN": row_dict.get("fqdn", ""),
            "Domain": row_dict.get("domain", ""),
            "Pingable": row_dict.get("pingable", ""),
            "MatchedFilter": row_dict.get("matched_filter", ""),
            "InventoryStatus": row_dict.get("inventory_status", ""),
            "SoftwareCount": row_dict.get("software_count", 0),
            "LastSeen": row_dict.get("last_seen", ""),
        }
        normalized["_SearchText"] = " ".join(
            str(value).lower()
            for key, value in normalized.items()
            if not key.startswith("_")
        )
        return normalized

    def _schedule_group_versions_reload(self) -> None:
        """Debounce Group Versions reloads."""
        if (
            self._suspend_traces
            or self._is_busy
            or self._selected_is_already_grouped
        ):
            return

        if self._group_after_id:
            self.after_cancel(self._group_after_id)

        self._group_after_id = self.after(
            self.GROUP_RELOAD_DEBOUNCE_MS,
            self._handle_group_versions_changed,
        )

    def _handle_group_versions_changed(self) -> None:
        """Reload breakout rows when Group Versions changes."""
        self._group_after_id = None

        if not self.current_software_key or not self.reload_callback:
            return

        self.set_busy(True)
        self.reload_callback(
            self.current_software_key,
            self.should_group_versions(),
        )

    def _handle_double_click(self, values: Sequence[Any]) -> None:
        """
        Follow selected computer to Installed Software tab.

        Args:
            values: Selected row values.
        """
        if not self.selection_callback:
            return

        computer_id = values[1] if len(values) > 1 else ""
        self.selection_callback(computer_id)

    def _cancel_pending_work(self) -> None:
        """Cancel pending debounce/render callbacks."""
        for after_id in (
            self._search_after_id,
            self._group_after_id,
            self._render_after_id,
        ):
            if after_id:
                try:
                    self.after_cancel(after_id)
                except tk.TclError:
                    pass

        self._search_after_id = None
        self._group_after_id = None
        self._render_after_id = None
        self._render_token += 1

# ------------------ Details Panel ------------------ #
class DetailsPanel(ttk.Frame):
    """
    Display detailed selected object properties in organized sections.

    The panel uses grouped cards and read-only text fields so computer,
    software, scan, and error details are easier to read than a flat table.
    """

    FIELD_GROUPS = {
        "identity": {
            "title": "Identity",
            "keys": {
                "ComputerId",
                "Hostname",
                "IPAddress",
                "Display Name",
                "DisplayName",
                "Display Version",
                "DisplayVersion",
                "Publisher",
                "Publisher Group",
                "SoftwareGroup",
            },
        },
        "inventory": {
            "title": "Inventory",
            "keys": {
                "InventoryStatus",
                "SoftwareCount",
                "Occurrence Count",
                "OccurrenceCount",
                "Percent of Inventoried Computers",
                "PercentOfInventoriedComputers",
                "Outdated",
                "LastSeen",
            },
        },
        "scan": {
            "title": "Scan",
            "keys": {
                "scan_id",
                "status",
                "started_at",
                "completed_at",
                "duration_seconds",
                "ip_range",
                "filters",
            },
        },
    }

    def __init__(self, parent: tk.Widget):
        """
        Initialize details panel.

        Args:
            parent: Parent Tkinter widget.
        """
        super().__init__(parent, padding=10)

        self.detail_type_var = tk.StringVar(value="Details")
        self.current_properties: Dict[str, Any] = {}

        self._create_widgets()

    def show_detail_properties(
        self,
        properties: Dict[str, Any],
        detail_type: str = "Details",
    ) -> None:
        """
        Display key/value properties in organized sections.

        Args:
            properties: Dictionary of properties to display.
            detail_type: Label describing the selected detail type.
        """
        self.clear()
        self.current_properties = dict(properties or {})
        self.detail_type_var.set(detail_type)

        grouped_properties = self._group_properties(self.current_properties)

        for group_title, group_values in grouped_properties:
            self._add_section(group_title, group_values)

    def clear(self) -> None:
        """Clear displayed details."""
        self.current_properties = {}

        for child in self.content_frame.winfo_children():
            child.destroy()

    def copy_selected_property(self) -> None:
        """Copy all displayed properties to clipboard."""
        self.copy_all_properties()

    def copy_all_properties(self) -> None:
        """Copy all displayed properties."""
        if not self.current_properties:
            return

        lines = [
            f"{key}\t{'' if value is None else value}"
            for key, value in self.current_properties.items()
        ]

        try:
            self.clipboard_clear()
            self.clipboard_append("\n".join(lines))
            AppLogger.log_message("debug", "Details copied to clipboard.")
        except tk.TclError as exc:
            AppLogger.log_message("error", f"Failed to copy details: {exc}")

    def _create_widgets(self) -> None:
        """Create details panel layout."""
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(
            header_frame,
            textvariable=self.detail_type_var,
            font=("Arial", 14, "bold"),
        ).pack(side="left")

        button_frame = ttk.Frame(header_frame)
        button_frame.pack(side="right")

        self.copy_all_button = ttk.Button(
            button_frame,
            text="Copy Details",
            command=self.copy_all_properties,
        )
        self.copy_all_button.pack(side="left")

        self.canvas = tk.Canvas(
            self,
            bg=AppConfig.DEFAULT_BACKGROUND,
            highlightthickness=0,
        )
        self.scrollbar = ttk.Scrollbar(
            self,
            orient="vertical",
            command=self.canvas.yview,
        )
        self.content_frame = ttk.Frame(self.canvas)

        self.content_frame.bind(
            "<Configure>",
            lambda _event: self.canvas.configure(
                scrollregion=self.canvas.bbox("all"),
            ),
        )

        self.canvas_window = self.canvas.create_window(
            (0, 0),
            window=self.content_frame,
            anchor="nw",
        )
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.canvas.bind("<Configure>", self._resize_canvas_window)

    def _group_properties(
        self,
        properties: Dict[str, Any],
    ) -> List[Tuple[str, Dict[str, Any]]]:
        """
        Group properties into display sections.

        Args:
            properties: Raw property dictionary.

        Returns:
            Ordered list of section title and property dictionaries.
        """
        remaining = dict(properties)
        grouped: List[Tuple[str, Dict[str, Any]]] = []

        for group_config in self.FIELD_GROUPS.values():
            group_values = {}

            for key in list(remaining):
                if key in group_config["keys"]:
                    group_values[key] = remaining.pop(key)

            if group_values:
                grouped.append((group_config["title"], group_values))

        if remaining:
            grouped.append(("Additional Properties", remaining))

        if not grouped:
            grouped.append(("Details", {"Message": "No details available."}))

        return grouped

    def _add_section(self, title: str, properties: Dict[str, Any]) -> None:
        """
        Add a visual section card.

        Args:
            title: Section title.
            properties: Section properties.
        """
        section = tk.Frame(
            self.content_frame,
            bg="#FFFFFF",
            bd=1,
            relief="groove",
            padx=10,
            pady=8,
        )
        section.pack(fill="x", expand=True, padx=4, pady=6)

        tk.Label(
            section,
            text=title,
            bg="#FFFFFF",
            font=("Arial", 11, "bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 6))

        for row_index, (key, value) in enumerate(properties.items(), start=1):
            display_value = "" if value is None else str(value)

            tk.Label(
                section,
                text=self._prettify_key(key),
                bg="#FFFFFF",
                font=("Arial", 9, "bold"),
                anchor="w",
            ).grid(row=row_index, column=0, sticky="nw", padx=(0, 12), pady=3)

            value_widget = tk.Text(
                section,
                height=self._text_height(display_value),
                wrap="word",
                bg="#F8F8F8",
                relief="flat",
                padx=5,
                pady=3,
            )
            value_widget.insert("1.0", display_value)
            value_widget.configure(state="disabled")
            value_widget.grid(row=row_index, column=1, sticky="ew", pady=3)

        section.grid_columnconfigure(1, weight=1)

    def _prettify_key(self, key: Any) -> str:
        """
        Convert internal key names into cleaner labels.

        Args:
            key: Raw property key.

        Returns:
            Display label.
        """
        text = str(key)
        text = text.replace("_", " ")

        words = []
        current = ""

        for char in text:
            if current and char.isupper() and current[-1].islower():
                words.append(current)
                current = char
            else:
                current += char

        if current:
            words.append(current)

        return " ".join(words).title()

    def _text_height(self, value: str) -> int:
        """
        Estimate a compact text widget height.

        Args:
            value: Text value.

        Returns:
            Height in text lines.
        """
        line_count = max(1, value.count("\n") + 1)
        wrapped_count = max(1, len(value) // 90 + 1)
        return min(6, max(line_count, wrapped_count))

    def _resize_canvas_window(self, event: tk.Event) -> None:
        """
        Keep the scrollable content width aligned to the canvas.

        Args:
            event: Tkinter configure event.
        """
        self.canvas.itemconfigure(self.canvas_window, width=event.width)

# ------------------ Live Feed Panel ------------------ #
class LiveFeedPanel(ttk.Frame):
    """
    Human-readable real-time scan feed.

    The panel supports auto-scroll, pause auto-scroll, clear, and copy behavior.
    """

    def __init__(self, parent: tk.Widget):
        """
        Initialize live feed panel.

        Args:
            parent: Parent Tkinter widget.
        """
        super().__init__(parent, padding=10)

        self.auto_scroll_var = tk.BooleanVar(value=True)
        self._last_normalized_entry = ""
        self._create_widgets()

    def append_live_feed(self, text: str) -> None:
        """
        Append a line of scan feed text.

        Args:
            text: Message to append.
        """
        if not text:
            return

        normalized = text.strip()
        if not normalized:
            return

        if normalized == self._last_normalized_entry:
            return

        self._last_normalized_entry = normalized

        self.text.configure(state="normal")
        self.text.insert("end", normalized + "\n")
        self.text.configure(state="disabled")

        if self.auto_scroll_var.get():
            self.text.see("end")

    def clear_feed(self) -> None:
        """Clear the live feed."""
        self.text.configure(state="normal")
        self.text.delete("1.0", "end")
        self.text.configure(state="disabled")
        self._last_normalized_entry = ""
        AppLogger.log_message("debug", "Live feed cleared.")

    def copy_feed(self) -> None:
        """Copy the full live feed to clipboard."""
        try:
            content = self.text.get("1.0", "end").strip()
            self.clipboard_clear()
            self.clipboard_append(content)
            AppLogger.log_message("debug", "Live feed copied to clipboard.")
        except tk.TclError as exc:
            AppLogger.log_message("error", f"Failed to copy live feed: {exc}")

    def _create_widgets(self) -> None:
        """Create live feed widgets."""
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(
            header_frame,
            text="Live Feed",
            font=("Arial", 14, "bold"),
        ).pack(side="left")

        control_frame = ttk.Frame(header_frame)
        control_frame.pack(side="right")

        self.auto_scroll_check = ttk.Checkbutton(
            control_frame,
            text="Auto-scroll",
            variable=self.auto_scroll_var,
        )
        self.auto_scroll_check.pack(side="left", padx=(0, 5))

        self.clear_button = ttk.Button(
            control_frame,
            text="Clear",
            command=self.clear_feed,
        )
        self.clear_button.pack(side="left", padx=5)

        self.copy_button = ttk.Button(
            control_frame,
            text="Copy",
            command=self.copy_feed,
        )
        self.copy_button.pack(side="left")

        text_frame = ttk.Frame(self)
        text_frame.pack(fill="both", expand=True)

        self.text = tk.Text(
            text_frame,
            height=12,
            state="disabled",
            bg="#FFFFFF",
            wrap="word",
        )
        scrollbar = ttk.Scrollbar(
            text_frame,
            orient="vertical",
            command=self.text.yview,
        )
        self.text.configure(yscrollcommand=scrollbar.set)

        self.text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

# ------------------ Logs Panel ------------------ #
class LogsPanel(ttk.Frame):
    """
    Display application logs inside the main UI.

    This panel consumes log messages from AppLogger's queue and displays
    them in real time.
    """

    def __init__(self, parent: tk.Widget, log_queue: queue.Queue):
        """
        Initialize logs panel.

        Args:
            parent: Parent Tkinter widget.
            log_queue: Queue provided by AppLogger.
        """
        super().__init__(parent, padding=10)

        self.log_queue = log_queue
        self.level_var = tk.StringVar(value="INFO+")
        self.auto_scroll_var = tk.BooleanVar(value=True)
        self._records: List[str] = []
        self._create_widgets()

    def append_log(self, text: str) -> None:
        """
        Append a log message to the panel.

        Args:
            text: Log message.
        """
        if not text:
            return

        self._records.append(text.rstrip())
        if not self._passes_level_filter(text):
            return

        self.text.configure(state="normal")
        self.text.insert("end", text.rstrip() + "\n")
        self._apply_color(text)
        self.text.configure(state="disabled")
        if self.auto_scroll_var.get():
            self.text.see("end")

    def process_log_queue(self) -> None:
        """
        Drain log queue and append logs to UI.

        This should be called periodically from the main thread.
        """
        try:
            while True:
                record = self.log_queue.get_nowait()
                self.append_log(record)
        except queue.Empty:
            pass

    def clear_logs(self) -> None:
        """Clear log display."""
        self._records.clear()
        self.text.configure(state="normal")
        self.text.delete("1.0", "end")
        self.text.configure(state="disabled")
        AppLogger.log_message("debug", "Logs cleared.")

    def copy_logs(self) -> None:
        """Copy all logs to clipboard."""
        try:
            content = self.text.get("1.0", "end").strip()
            self.clipboard_clear()
            self.clipboard_append(content)
            AppLogger.log_message("debug", "Logs copied to clipboard.")
        except tk.TclError as exc:
            AppLogger.log_message("error", f"Failed to copy logs: {exc}")

    def _apply_color(self, text: str) -> None:
        """
        Apply basic color-coding based on log level.

        Args:
            text: Log message.
        """
        lowered = text.lower()

        tag = None
        if "error" in lowered or "critical" in lowered:
            tag = "error"
        elif "warning" in lowered:
            tag = "warning"
        elif "info" in lowered:
            tag = "info"
        elif "debug" in lowered:
            tag = "debug"

        if tag:
            start_index = self.text.index("end-2l")
            end_index = self.text.index("end-1l")
            self.text.tag_add(tag, start_index, end_index)

    def _passes_level_filter(self, text: str) -> bool:
        """
        Check whether a log message should be displayed for current filter.

        Args:
            text: Full log message line.
        """
        selected = self.level_var.get()
        lowered = text.lower()
        level_order = {
            "DEBUG": 10,
            "INFO": 20,
            "WARNING": 30,
            "ERROR": 40,
            "CRITICAL": 50,
        }
        selected_threshold = {
            "DEBUG+": "DEBUG",
            "INFO+": "INFO",
            "WARNING+": "WARNING",
            "ERROR+": "ERROR",
            "CRITICAL": "CRITICAL",
        }.get(selected, "INFO")

        record_level_name = "INFO"
        for candidate in ("CRITICAL", "ERROR", "WARNING", "INFO", "DEBUG"):
            if candidate.lower() in lowered:
                record_level_name = candidate
                break

        return (
            level_order[record_level_name]
            >= level_order[selected_threshold]
        )

    def _refresh_visible_logs(self) -> None:
        """Rebuild text area from cached records using current filter."""
        self.text.configure(state="normal")
        self.text.delete("1.0", "end")

        for record in self._records:
            if not self._passes_level_filter(record):
                continue
            self.text.insert("end", record + "\n")
            self._apply_color(record)

        self.text.configure(state="disabled")
        if self.auto_scroll_var.get():
            self.text.see("end")

    def _on_level_changed(self, _event: Optional[tk.Event] = None) -> None:
        """Handle user log-level dropdown changes."""
        self._refresh_visible_logs()

    def _create_widgets(self) -> None:
        """Create logs panel layout."""
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(
            header_frame,
            text="Logs",
            font=("Arial", 14, "bold"),
        ).pack(side="left")

        control_frame = ttk.Frame(header_frame)
        control_frame.pack(side="right")

        self.clear_button = ttk.Button(
            control_frame,
            text="Clear",
            command=self.clear_logs,
        )
        self.clear_button.pack(side="left", padx=5)

        self.copy_button = ttk.Button(
            control_frame,
            text="Copy",
            command=self.copy_logs,
        )
        self.copy_button.pack(side="left")

        ttk.Label(
            control_frame,
            text="Level:",
        ).pack(side="left", padx=(12, 4))

        self.level_combobox = ttk.Combobox(
            control_frame,
            textvariable=self.level_var,
            values=("DEBUG+", "INFO+", "WARNING+", "ERROR+", "CRITICAL"),
            state="readonly",
            width=10,
        )
        self.level_combobox.pack(side="left")
        self.level_combobox.bind("<<ComboboxSelected>>", self._on_level_changed)

        self.auto_scroll_check = ttk.Checkbutton(
            control_frame,
            text="Auto-scroll",
            variable=self.auto_scroll_var,
        )
        self.auto_scroll_check.pack(side="left", padx=(12, 0))

        text_frame = ttk.Frame(self)
        text_frame.pack(fill="both", expand=True)

        self.text = tk.Text(
            text_frame,
            height=12,
            state="disabled",
            bg="#FFFFFF",
            wrap="word",
        )

        scrollbar = ttk.Scrollbar(
            text_frame,
            orient="vertical",
            command=self.text.yview,
        )
        self.text.configure(yscrollcommand=scrollbar.set)

        self.text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Configure color tags
        self.text.tag_configure("error", foreground="red")
        self.text.tag_configure("warning", foreground="orange")
        self.text.tag_configure("info", foreground="black")
        self.text.tag_configure("debug", foreground="#5C6370")
    
# ------------------ Status Bar ------------------ #
class StatusBar(ttk.Frame):
    """
    Compact application status display.

    Displays current scan phase, last action, and overall state.
    """

    def __init__(self, parent: tk.Widget):
        """
        Initialize status bar.

        Args:
            parent: Parent Tkinter widget.
        """
        super().__init__(parent, padding=(5, 2))

        self.phase_var = tk.StringVar(value="Ready")
        self.message_var = tk.StringVar(value="Application ready.")
        self.selection_var = tk.StringVar(value="")

        self._create_widgets()

    def update_status(
        self,
        phase: Optional[str] = None,
        message: Optional[str] = None,
        selection_count: Optional[int] = None,
    ) -> None:
        """
        Update status bar values.

        Args:
            phase: Current phase (e.g., scanning, inventorying).
            message: Status message.
            selection_count: Number of selected items.
        """
        if phase is not None:
            self.phase_var.set(str(phase).title())

        if message is not None:
            self.message_var.set(str(message))

        if selection_count is not None:
            self.selection_var.set(f"Selected: {selection_count}")

    def set_ready(self) -> None:
        """Set status to ready."""
        self.update_status(
            phase="ready",
            message="Application ready.",
            selection_count=0,
        )

    def set_scanning(self, ip_range: str) -> None:
        """
        Set status to scanning.

        Args:
            ip_range: IP range being scanned.
        """
        self.update_status(
            phase="scanning",
            message=f"Scanning {ip_range}...",
        )

    def set_inventorying(self, hostname: str) -> None:
        """
        Set status to inventorying.

        Args:
            hostname: Target hostname.
        """
        self.update_status(
            phase="inventorying",
            message=f"Inventorying {hostname}...",
        )

    def set_complete(self, summary: Dict[str, Any]) -> None:
        """
        Set status to scan complete.

        Args:
            summary: Scan summary dictionary.
        """
        count = summary.get("successful_inventory_count", 0)
        self.update_status(
            phase="complete",
            message=f"Scan complete: {count} computers inventoried",
        )

    def set_cancelled(self) -> None:
        """Set status to cancelled."""
        self.update_status(
            phase="cancelled",
            message="Scan cancelled.",
        )

    def set_error(self, message: str) -> None:
        """
        Set status to error state.

        Args:
            message: Error message.
        """
        self.update_status(
            phase="error",
            message=message,
        )

    def _create_widgets(self) -> None:
        """Create status bar layout."""
        self.configure(relief="sunken")

        self.phase_label = ttk.Label(
            self,
            textvariable=self.phase_var,
            font=("Arial", 9, "bold"),
        )
        self.phase_label.pack(side="left", padx=(5, 10))

        self.message_label = ttk.Label(
            self,
            textvariable=self.message_var,
            anchor="w",
        )
        self.message_label.pack(side="left", fill="x", expand=True)

        self.selection_label = ttk.Label(
            self,
            textvariable=self.selection_var,
        )
        self.selection_label.pack(side="right", padx=5)

# ------------------ Results Notebook ------------------ #
class ResultsNotebook(ttk.Frame):
    """
    Own all result tabs and tab-to-tab navigation.

    Tabs remain open after navigation and are ordered by category.
    The selected tab is color-tinted by category so the full tab header is
    highlighted instead of a small color swatch.
    """

    TAB_CATEGORIES = {
        "Dashboard": ("System", "#D9EAF7", 10),
        "Software Summary": ("Software", "#FFF2CC", 20),
        "Computer Breakout": ("Software", "#FFF2CC", 21),
        "Computers": ("Computer", "#D4FAD4", 30),
        "Installed Software": ("Computer", "#D4FAD4", 31),
        "Details": ("Computer", "#D4FAD4", 32),
        "Live Feed": ("System", "#D9EAF7", 40),
        "Logs": ("System", "#D9EAF7", 41),
    }

    def __init__(
        self,
        parent: tk.Widget,
        log_queue: queue.Queue,
        database_manager: DatabaseManager,
    ):
        """
        Initialize results notebook.

        Args:
            parent: Parent widget.
            log_queue: GUI log queue.
            database_manager: Database service.
        """
        super().__init__(parent)

        self.log_queue = log_queue
        self.database_manager = database_manager
        self.current_scan_id: Optional[int] = None

        self._tab_widgets: Dict[str, tk.Widget] = {}
        self._breakout_request_id = 0

        self._create_widgets()
        self._create_tabs()
        self._sort_tabs_by_category()
        self._apply_selected_tab_color()

    def set_current_scan_id(self, scan_id: Optional[int]) -> None:
        """
        Store current scan ID.

        Args:
            scan_id: Current scan ID.
        """
        self.current_scan_id = scan_id

    def populate_all_results(
        self,
        scan_id: int,
        summary: Optional[Dict[str, Any]] = None,
    ) -> None:
        """
        Populate all primary result panels after scan completion.

        Args:
            scan_id: Scan ID.
            summary: Optional scan summary.
        """
        self.current_scan_id = scan_id

        if summary is None:
            summary = self.database_manager.get_scan_summary(scan_id)

        self.update_dashboard(summary)

        software_rows = self.database_manager.get_software_occurrences(scan_id)
        successful_count = int(
            summary.get("successful_inventory_count", 0) or 0
        )
        self.software_summary_panel.populate(
            software_rows,
            successful_inventory_count=successful_count,
        )

        computers = self.database_manager.get_all_computers_with_counts(scan_id)
        self.computer_inventory_panel.populate_computers(computers)

    def populate_software_summary(
        self,
        rows: Sequence[Any],
        successful_inventory_count: int = 0,
    ) -> None:
        """
        Populate Software Summary tab.

        Args:
            rows: Software summary rows.
            successful_inventory_count: Number of successful inventory targets.
        """
        self.software_summary_panel.populate(
            rows,
            successful_inventory_count=successful_inventory_count,
        )

    def populate_computers(self, rows: Sequence[Any]) -> None:
        """
        Populate main Computers tab.

        Args:
            rows: Computer rows.
        """
        self.computer_inventory_panel.populate_computers(rows)

    def update_dashboard(self, summary: Dict[str, Any]) -> None:
        """
        Update Dashboard panel.

        Args:
            summary: Dashboard summary.
        """
        self.dashboard_panel.update_dashboard(summary)

    def show_computer_breakout_for_software(
        self,
        software_key: Dict[str, Any],
    ) -> None:
        """
        Show installed/not-installed computer breakout for selected software.

        Args:
            software_key: Selected software identity.
        """
        if not self.current_scan_id:
            return

        group_versions = self.computer_breakout_panel.reset_for_new_selection(
            software_key
        )
        self.reload_computer_breakout(
            software_key=software_key,
            group_versions=group_versions,
        )

    def reload_computer_breakout(
        self,
        software_key: Dict[str, Any],
        group_versions: bool = False,
    ) -> None:
        """
        Reload Computer Breakout for selected software without blocking the GUI.

        Args:
            software_key: Selected software identity.
            group_versions: True to ignore versions while matching.
        """
        if not self.current_scan_id:
            return

        self._breakout_request_id += 1
        request_id = self._breakout_request_id

        self._select_tab("Computer Breakout")
        self.computer_breakout_panel.set_busy(True)

        scan_id = self.current_scan_id
        display_name = software_key.get("display_name", "")
        display_version = software_key.get("display_version", "")
        publisher = software_key.get("publisher", "")

        def worker() -> None:
            """Load breakout rows away from Tkinter's main thread."""
            try:
                rows = self.database_manager.get_computer_breakout_for_software(
                    scan_id=scan_id,
                    display_name=display_name,
                    display_version=display_version,
                    publisher=publisher,
                    group_versions=group_versions,
                )
                self.after(
                    0,
                    lambda: self._finish_breakout_reload(
                        request_id,
                        rows,
                        software_key,
                    ),
                )
            except Exception as exc:
                AppLogger.log_message(
                    "error",
                    f"Failed to load computer breakout: {exc}",
                )
                self.after(
                    0,
                    lambda: self.computer_breakout_panel.set_busy(False),
                )

        thread = threading.Thread(
            target=worker,
            name="ComputerBreakoutLoader",
            daemon=True,
        )
        thread.start()

    def show_computers_for_software(
        self,
        software_key: Dict[str, Any],
    ) -> None:
        """
        Backward-compatible wrapper for older callbacks.

        Args:
            software_key: Selected software identity.
        """
        self.show_computer_breakout_for_software(software_key)

    def show_software_for_computer(self, computer_id: Any) -> None:
        """
        Show Installed Software tab for selected computer.

        Args:
            computer_id: Computer ID.
        """
        if not self.current_scan_id:
            return

        try:
            safe_computer_id = int(computer_id)
        except (TypeError, ValueError):
            AppLogger.log_message(
                "warning",
                f"Invalid computer ID selected: {computer_id}",
            )
            return

        rows = self.database_manager.get_software_for_computer(
            self.current_scan_id,
            safe_computer_id,
        )
        self.installed_software_panel.populate_installed_software(rows)
        self._select_tab("Installed Software")

    def show_detail_properties(
        self,
        detail_type: str,
        record_id: Any,
    ) -> None:
        """
        Show full property details.

        Args:
            detail_type: Detail type, computer or software.
            record_id: Record ID.
        """
        try:
            safe_record_id = int(record_id)
        except (TypeError, ValueError):
            AppLogger.log_message(
                "warning",
                f"Invalid detail record ID selected: {record_id}",
            )
            return

        if detail_type == "computer":
            properties = self.database_manager.get_computer_properties(
                safe_record_id
            )
        else:
            properties = self.database_manager.get_software_properties(
                safe_record_id
            )

        self.details_panel.show_detail_properties(properties, detail_type=detail_type.title())
        self._select_tab("Details")

    def append_live_feed(self, text: str) -> None:
        """
        Append text to Live Feed tab.

        Args:
            text: Feed text.
        """
        self.live_feed_panel.append_live_feed(text)

    def append_log(self, text: str) -> None:
        """
        Append text to Logs tab.

        Args:
            text: Log text.
        """
        self.logs_panel.append_log(text)

    def process_log_queue(self) -> None:
        """Drain queued log messages into Logs tab."""
        try:
            while True:
                record = self.log_queue.get_nowait()
                self.append_log(record)
        except queue.Empty:
            return

    def clear_results(self) -> None:
        """Clear all result panels."""
        self.current_scan_id = None
        self.dashboard_panel.clear()
        self.software_summary_panel.clear()
        self.computer_inventory_panel.clear()
        self.computer_breakout_panel.clear()
        self.installed_software_panel.clear()
        self.details_panel.clear()
        self.live_feed_panel.clear()
        self.logs_panel.clear()
        self._select_tab("Dashboard")

    def restore_all_computers(self) -> None:
        """Compatibility method; Computers tab always shows all computers."""
        self._select_tab("Computers")

    def _create_widgets(self) -> None:
        """Create Notebook widget."""
        self._configure_notebook_style()
        self.notebook = ttk.Notebook(self, style="Results.TNotebook")
        self.notebook.pack(fill="both", expand=True)
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

    def _create_tabs(self) -> None:
        """Create all result tabs."""
        self.dashboard_panel = DashboardPanel(self.notebook)
        self.software_summary_panel = SoftwareSummaryPanel(
            self.notebook,
            selection_callback=self.show_computer_breakout_for_software,
        )
        self.computer_breakout_panel = ComputerBreakoutPanel(
            self.notebook,
            reload_callback=self.reload_computer_breakout,
            selection_callback=self.show_software_for_computer,
        )
        self.computer_inventory_panel = ComputerInventoryPanel(
            self.notebook,
            selection_callback=self.show_software_for_computer,
        )
        self.installed_software_panel = InstalledSoftwarePanel(
            self.notebook,
            detail_callback=lambda software_id: self.show_detail_properties(
                "software",
                software_id,
            ),
        )
        self.details_panel = DetailsPanel(self.notebook)
        self.live_feed_panel = LiveFeedPanel(self.notebook)
        self.logs_panel = LogsPanel(self.notebook, self.log_queue)

        self._tab_widgets = {
            "Dashboard": self.dashboard_panel,
            "Software Summary": self.software_summary_panel,
            "Computer Breakout": self.computer_breakout_panel,
            "Computers": self.computer_inventory_panel,
            "Installed Software": self.installed_software_panel,
            "Details": self.details_panel,
            "Live Feed": self.live_feed_panel,
            "Logs": self.logs_panel,
        }

        for tab_name, widget in self._tab_widgets.items():
            self.notebook.add(
                widget,
                text=self._tab_label(tab_name),
                padding=(8, 3),
            )

    def _finish_breakout_reload(
        self,
        request_id: int,
        rows: Sequence[Any],
        software_key: Dict[str, Any],
    ) -> None:
        """
        Apply async breakout query results on the GUI thread.

        Args:
            request_id: Reload request ID.
            rows: Breakout rows.
            software_key: Selected software identity.
        """
        if request_id != self._breakout_request_id:
            AppLogger.log_message(
                "debug",
                "Ignored stale computer breakout reload result.",
            )
            return

        self.computer_breakout_panel.populate_breakout(rows, software_key)
        self._select_tab("Computer Breakout")

    def _sort_tabs_by_category(self) -> None:
        """Sort tabs by configured category/order."""
        ordered_names = sorted(
            self._tab_widgets,
            key=lambda name: self.TAB_CATEGORIES.get(name, ("System", "", 99))[2],
        )

        for index, tab_name in enumerate(ordered_names):
            self.notebook.insert(index, self._tab_widgets[tab_name])

    def _select_tab(self, tab_name: str) -> None:
        """
        Select a tab by name.

        Args:
            tab_name: Tab name.
        """
        widget = self._tab_widgets.get(tab_name)

        if not widget:
            AppLogger.log_message(
                "warning",
                f"Requested tab does not exist: {tab_name}",
            )
            return

        self.notebook.select(widget)
        self._apply_selected_tab_color()

    def _tab_label(self, tab_name: str) -> str:
        """
        Build category-prefixed tab label.

        Args:
            tab_name: Raw tab name.

        Returns:
            Display tab label.
        """
        return tab_name

    def _configure_notebook_style(self) -> None:
        """Configure base notebook styles used for full-tab color highlighting."""
        style = ttk.Style(self)
        style.configure(
            "Results.TNotebook.Tab",
            padding=(12, 6),
            background="#E6E6E6",
        )
        style.map(
            "Results.TNotebook.Tab",
            background=[("selected", "#D9EAF7")],
        )

    def _on_tab_changed(self, _event: tk.Event) -> None:
        """Refresh selected tab coloring when the active tab changes."""
        self._apply_selected_tab_color()

    def _apply_selected_tab_color(self) -> None:
        """Tint the selected tab using the tab category color."""
        selected = self.notebook.select()
        if not selected:
            return

        selected_widget = self.nametowidget(selected)
        selected_name = next(
            (name for name, widget in self._tab_widgets.items() if widget == selected_widget),
            None,
        )
        if not selected_name:
            return

        _category, color, _order = self.TAB_CATEGORIES.get(
            selected_name,
            ("System", "#D9EAF7", 99),
        )
        style = ttk.Style(self)
        style.map(
            "Results.TNotebook.Tab",
            background=[("selected", color)],
        )

# ------------------ Main Window Layout ------------------ #
class MainWindowLayout(ttk.Frame):
    """
    Compose the main application layout.

    This class owns the top-level UI sections and exposes a small set of helper
    methods used by MainApplication.
    """

    def __init__(
        self,
        parent: tk.Widget,
        preferences_manager: PreferencesManager,
        database_manager: DatabaseManager,
        log_queue: queue.Queue,
        start_callback: Any,
        cancel_callback: Any,
        clear_callback: Any,
        export_callback: Any,
        credential_callback: Any,
        show_credentials_button: bool = False,
    ):
        """
        Initialize main window layout.

        Args:
            parent: Parent Tkinter widget.
            preferences_manager: Preferences service.
            database_manager: Database service.
            log_queue: GUI log queue.
            start_callback: Scan start callback.
            cancel_callback: Scan cancel callback.
            clear_callback: Clear results callback.
            export_callback: Export callback.
            credential_callback: Credential prompt callback.
            show_credentials_button: Whether to show Credentials button.
        """
        super().__init__(parent, padding=8, style="App.TFrame")

        self.parent = parent
        self.preferences_manager = preferences_manager
        self.database_manager = database_manager
        self.log_queue = log_queue

        self.start_callback = start_callback
        self.cancel_callback = cancel_callback
        self.clear_callback = clear_callback
        self.export_callback = export_callback
        self.credential_callback = credential_callback
        self.show_credentials_button = show_credentials_button

        self.admin_status_var = tk.StringVar(value="Admin: Unknown")
        self.credential_status_var = tk.StringVar(value="Credentials: Not Cached")

        self._create_widgets()
        self.pack(fill="both", expand=True)

    def set_admin_status(self, is_admin: bool) -> None:
        """
        Update admin status text.

        Args:
            is_admin: True when running elevated.
        """
        status = "Admin: Yes" if is_admin else "Admin: No"
        self.admin_status_var.set(status)

    def set_credentials_cached(self, is_cached: bool) -> None:
        """
        Update credential cache status.

        Args:
            is_cached: True when delegated credentials are cached.
        """
        if is_cached:
            self.credential_status_var.set("Credentials: Cached")
            self.credential_status_label.configure(
                foreground="#107C10",
                font=("Arial", 9, "bold"),
            )
        else:
            self.credential_status_var.set("Credentials: Not Cached")
            self.credential_status_label.configure(
                foreground="#7A4E00",
                font=("Arial", 9, "bold"),
            )

        self.scan_control_panel.set_credentials_cached(is_cached)

    def set_export_running(self, is_running: bool) -> None:
        """
        Enable or disable export controls.

        Args:
            is_running: True while export is active.
        """
        self.scan_control_panel.set_export_running(is_running)

        if is_running:
            self.status_bar.update_status(
                phase="export",
                message="Export is running in the background.",
            )
        else:
            self.status_bar.update_status(
                phase="export",
                message="Export finished.",
            )

    def set_scan_running(self, is_running: bool) -> None:
        """
        Update scan-control state.

        Args:
            is_running: True while a scan is active.
        """
        self.scan_control_panel.set_scan_running(is_running)

    def save_current_preferences(self) -> None:
        """Persist current scan-control preferences."""
        self.scan_control_panel.save_current_preferences()

    def update_status(
        self,
        phase: Optional[str] = None,
        message: Optional[str] = None,
        selection_count: Optional[int] = None,
    ) -> None:
        """
        Update status bar.

        Args:
            phase: Current phase.
            message: Status message.
            selection_count: Optional selected row count.
        """
        self.status_bar.update_status(
            phase=phase,
            message=message,
            selection_count=selection_count,
        )

    def update_dashboard(self, summary: Dict[str, Any]) -> None:
        """
        Update dashboard tab.

        Args:
            summary: Scan summary data.
        """
        self.results_notebook.update_dashboard(summary)

    def append_live_feed(self, text: str) -> None:
        """
        Append a message to the Live Feed tab.

        Args:
            text: Display text.
        """
        self.results_notebook.append_live_feed(text)

    def process_log_queue(self) -> None:
        """Process queued logs."""
        self.results_notebook.process_log_queue()

    def reset_scan_progress(self) -> None:
        """Reset and show scan progress bars for a new scan."""
        self.ping_progress_var.set(0)
        self.inventory_progress_var.set(0)
        self.ping_progress_label_var.set("Ping: 0/0 (0%)")
        self.inventory_progress_label_var.set("Inventory: 0/0 (0%)")
        self._show_progress_bar("ping")
        self._hide_progress_bar("inventory")


    def update_progress(self, progress: ScanProgress) -> None:
        """
        Update independent ping and inventory progress bars.

        Args:
            progress: Scan progress update.
        """
        data = dict(progress.data or {})

        ping_total = int(data.get("ping_total", 0) or 0)
        ping_completed = int(data.get("ping_completed", 0) or 0)

        inventory_total = int(data.get("inventory_total", 0) or 0)
        inventory_completed = int(data.get("inventory_completed", 0) or 0)

        # Backward compatibility: support legacy progress payloads that only
        # populate top-level total/completed fields.
        if not ping_total and not inventory_total:
            inventory_total = max(0, int(progress.total or 0))
            inventory_completed = max(0, int(progress.completed or 0))

        self._update_named_progress_bar(
            name="ping",
            label="Ping",
            completed=ping_completed,
            total=ping_total,
        )
        self._update_named_progress_bar(
            name="inventory",
            label="Inventory",
            completed=inventory_completed,
            total=inventory_total,
        )

        if progress.message:
            self.update_status(
                phase=progress.phase,
                message=progress.message,
            )

        if progress.phase in {"complete", "cancelled"}:
            self._hide_progress_bar("ping")
            self._hide_progress_bar("inventory")


    def clear_results(self) -> None:
        """Clear result panels and reset progress."""
        self.results_notebook.clear_results()
        self.ping_progress_var.set(0)
        self.inventory_progress_var.set(0)
        self.ping_progress_label_var.set("Ping: 0/0 (0%)")
        self.inventory_progress_label_var.set("Inventory: 0/0 (0%)")
        self._hide_progress_bar("ping")
        self._hide_progress_bar("inventory")
        self.status_bar.set_ready()


    def _update_named_progress_bar(
        self,
        name: str,
        label: str,
        completed: int,
        total: int,
    ) -> None:
        """
        Update one named progress bar.

        Args:
            name: Progress bar name, either ping or inventory.
            label: Display label.
            completed: Completed unit count.
            total: Total unit count.
        """
        if total <= 0:
            self._hide_progress_bar(name)
            return

        safe_completed = max(0, min(total, completed))
        percent = int((safe_completed / total) * 100)

        if name == "ping":
            self.ping_progress_var.set(percent)
            self.ping_progress_label_var.set(
                f"{label}: {safe_completed}/{total} ({percent}%)"
            )
        elif name == "inventory":
            self.inventory_progress_var.set(percent)
            self.inventory_progress_label_var.set(
                f"{label}: {safe_completed}/{total} ({percent}%)"
            )

        if safe_completed >= total:
            self._hide_progress_bar(name)
            return

        self._show_progress_bar(name)


    def _show_progress_bar(self, name: str) -> None:
        """
        Show one progress bar if hidden.

        Args:
            name: Progress bar name.
        """
        if name == "ping":
            if self._ping_progress_visible:
                return

            self.ping_progress_name_label.grid(
                row=0,
                column=0,
                sticky="w",
                padx=(0, 6),
                pady=2,
            )
            self.ping_progress_bar.grid(
                row=0,
                column=1,
                sticky="ew",
                padx=(0, 8),
                pady=2,
            )
            self.ping_progress_value_label.grid(
                row=0,
                column=2,
                sticky="e",
                pady=2,
            )
            self._ping_progress_visible = True
            return

        if name == "inventory":
            if self._inventory_progress_visible:
                return

            self.inventory_progress_name_label.grid(
                row=1,
                column=0,
                sticky="w",
                padx=(0, 6),
                pady=2,
            )
            self.inventory_progress_bar.grid(
                row=1,
                column=1,
                sticky="ew",
                padx=(0, 8),
                pady=2,
            )
            self.inventory_progress_value_label.grid(
                row=1,
                column=2,
                sticky="e",
                pady=2,
            )
            self._inventory_progress_visible = True


    def _hide_progress_bar(self, name: str) -> None:
        """
        Hide one progress bar.

        Args:
            name: Progress bar name.
        """
        if name == "ping":
            self.ping_progress_name_label.grid_remove()
            self.ping_progress_bar.grid_remove()
            self.ping_progress_value_label.grid_remove()
            self._ping_progress_visible = False
            return

        if name == "inventory":
            self.inventory_progress_name_label.grid_remove()
            self.inventory_progress_bar.grid_remove()
            self.inventory_progress_value_label.grid_remove()
            self._inventory_progress_visible = False

    def _create_widgets(self) -> None:
        """Create main layout widgets."""
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        self.header_frame = ttk.Frame(self, style="App.TFrame")
        self.header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        self.header_frame.grid_columnconfigure(1, weight=1)

        self.title_label = ttk.Label(
            self.header_frame,
            text=AppConfig.APP_NAME,
            font=("Arial", 16, "bold"),
        )
        self.title_label.grid(row=0, column=0, sticky="w")

        self.status_frame = ttk.Frame(self.header_frame, style="App.TFrame")
        self.status_frame.grid(row=0, column=1, sticky="e")

        self.admin_status_label = ttk.Label(
            self.status_frame,
            textvariable=self.admin_status_var,
            font=("Arial", 9, "bold"),
        )
        self.admin_status_label.pack(side="left", padx=(0, 12))

        self.credential_status_label = ttk.Label(
            self.status_frame,
            textvariable=self.credential_status_var,
            font=("Arial", 9, "bold"),
            foreground="#7A4E00",
        )
        self.credential_status_label.pack(side="left")

        self.scan_control_panel = ScanControlPanel(
            self,
            preferences_manager=self.preferences_manager,
            start_callback=self.start_callback,
            cancel_callback=self.cancel_callback,
            clear_callback=self.clear_callback,
            export_callback=self.export_callback,
            credential_callback=self.credential_callback,
            show_credentials_button=self.show_credentials_button,
        )
        self.scan_control_panel.grid(row=1, column=0, sticky="ew", pady=(0, 8))

        self.progress_frame = ttk.Frame(self, style="App.TFrame")
        self.progress_frame.grid(row=2, column=0, sticky="ew", pady=(0, 8))
        self.progress_frame.grid_columnconfigure(1, weight=1)

        self.ping_progress_var = tk.IntVar(value=0)
        self.ping_progress_label_var = tk.StringVar(value="Ping: 0/0 (0%)")

        self.inventory_progress_var = tk.IntVar(value=0)
        self.inventory_progress_label_var = tk.StringVar(value="Inventory: 0/0 (0%)")

        self.ping_progress_name_label = ttk.Label(
            self.progress_frame,
            text="Ping",
            width=12,
            anchor="w",
        )
        self.ping_progress_bar = ttk.Progressbar(
            self.progress_frame,
            variable=self.ping_progress_var,
            maximum=100,
            mode="determinate",
        )
        self.ping_progress_value_label = ttk.Label(
            self.progress_frame,
            textvariable=self.ping_progress_label_var,
            width=24,
            anchor="e",
        )

        self.inventory_progress_name_label = ttk.Label(
            self.progress_frame,
            text="Inventory",
            width=12,
            anchor="w",
        )
        self.inventory_progress_bar = ttk.Progressbar(
            self.progress_frame,
            variable=self.inventory_progress_var,
            maximum=100,
            mode="determinate",
        )
        self.inventory_progress_value_label = ttk.Label(
            self.progress_frame,
            textvariable=self.inventory_progress_label_var,
            width=24,
            anchor="e",
        )

        self._ping_progress_visible = False
        self._inventory_progress_visible = False
        self._hide_progress_bar("ping")
        self._hide_progress_bar("inventory")

        self.results_notebook = ResultsNotebook(
            self,
            log_queue=self.log_queue,
            database_manager=self.database_manager,
        )
        self.results_notebook.grid(row=3, column=0, sticky="nsew")

        self.status_bar = StatusBar(self)
        self.status_bar.grid(row=4, column=0, sticky="ew", pady=(6, 0))

# ------------------ Export Management ------------------ #
class ExcelExportManager:
    """
    Export scan results to a polished XLSX workbook.

    Workbook sheets:
        Dashboard
        Software Summary
        Computers
        Inventory Details
        Export Notes

    Features:
        - Styled dashboard
        - Frozen panes
        - Excel tables with filters
        - Conditional formatting
        - Data bars
        - Status highlighting
        - Charts
        - Workbook navigation links
        - Safe cell-by-cell writing, no invalid range assignment
    """

    THEME = {
        "background": "ECECEC",
        "title": "1F4E79",
        "header": "D9EAF7",
        "header_dark": "5B9BD5",
        "success": "D4FAD4",
        "warning": "FFF2CC",
        "error": "FAD4D4",
        "neutral": "F8F8F8",
        "white": "FFFFFF",
        "border": "A6A6A6",
        "text": "1F1F1F",
    }

    def __init__(self, database_manager: DatabaseManager):
        """
        Initialize Excel export manager.

        Args:
            database_manager: Database service used to retrieve scan results.
        """
        self.database_manager = database_manager
        self.grouping_analyzer = SoftwareGroupingAnalyzer()

    def export_scan_to_xlsx(
        self,
        scan_id: int,
        output_path: str,
    ) -> ExportResult:
        """
        Export a scan to an XLSX workbook.

        Args:
            scan_id: Scan ID to export.
            output_path: Destination workbook path.

        Returns:
            ExportResult describing success or failure.
        """
        started = time.perf_counter()
        safe_path = Path(output_path).with_suffix(".xlsx")

        try:
            if not scan_id:
                return ExportResult(
                    success=False,
                    output_path=str(safe_path),
                    error="XLSX export failed: missing scan ID.",
                )

            AppLogger.log_message(
                "info",
                f"Starting XLSX export for scan {scan_id}.",
            )

            export_data = self._load_export_data(scan_id)

            workbook = Workbook()
            dashboard_sheet = workbook.active
            dashboard_sheet.title = "Dashboard"

            software_sheet = workbook.create_sheet("Software Summary")
            computers_sheet = workbook.create_sheet("Computers")
            details_sheet = workbook.create_sheet("Inventory Details")
            notes_sheet = workbook.create_sheet("Export Notes")

            self._build_dashboard_sheet(
                dashboard_sheet,
                export_data["summary"],
                export_data["software_rows"],
                export_data["computer_rows"],
            )
            self._build_software_summary_sheet(
                software_sheet,
                export_data["software_rows"],
            )
            self._build_computers_sheet(
                computers_sheet,
                export_data["computer_rows"],
            )
            self._build_inventory_details_sheet(
                details_sheet,
                export_data["inventory_rows"],
            )
            self._build_export_notes_sheet(
                notes_sheet,
                scan_id,
                safe_path,
            )

            workbook.active = 0
            workbook.save(safe_path)

            duration = round(time.perf_counter() - started, 2)
            AppLogger.log_message(
                "info",
                f"XLSX export completed in {duration}s: {safe_path}",
            )

            return ExportResult(
                success=True,
                output_path=str(safe_path),
                message="XLSX export completed.",
                duration_seconds=duration,
            )

        except Exception as exc:
            duration = round(time.perf_counter() - started, 2)
            AppLogger.log_message("error", f"XLSX export failed: {exc}")
            return ExportResult(
                success=False,
                output_path=str(safe_path),
                error=f"XLSX export failed: {exc}",
                duration_seconds=duration,
            )

    def _load_export_data(self, scan_id: int) -> Dict[str, Any]:
        """
        Load all export data before workbook creation.

        Args:
            scan_id: Scan ID.

        Returns:
            Export data dictionary.
        """
        summary = self.database_manager.get_scan_summary(scan_id)
        software_rows = self._load_software_summary(scan_id, summary)
        computer_rows = self.database_manager.get_all_computers_with_counts(scan_id)
        inventory_rows = self._load_inventory_details(scan_id)

        return {
            "summary": summary,
            "software_rows": software_rows,
            "computer_rows": computer_rows,
            "inventory_rows": inventory_rows,
        }

    def _load_software_summary(
        self,
        scan_id: int,
        summary: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        """
        Load and normalize software summary rows.

        Args:
            scan_id: Scan ID.
            summary: Scan summary dictionary.

        Returns:
            Enriched software rows.
        """
        rows = self.database_manager.get_software_occurrences(scan_id)
        successful_count = int(
            summary.get("successful_inventory_count", 0) or 0,
        )

        normalized_rows = []

        for row in rows:
            row_dict = dict(row)
            occurrence_count = int(row_dict.get("occurrence_count", 0) or 0)
            grouped_count = int(
                row_dict.get(
                    "grouped_version_occurrence_count",
                    occurrence_count,
                ) or 0
            )
            percent = 0.0

            if successful_count > 0:
                percent = min(1.0, occurrence_count / successful_count)

            normalized_rows.append({
                "DisplayName": row_dict.get("display_name", ""),
                "DisplayVersion": row_dict.get("display_version", ""),
                "Publisher": row_dict.get("publisher", ""),
                "OccurrenceCount": occurrence_count,
                "GroupedVersionOccurrenceCount": grouped_count,
                "PercentOfInventoriedComputers": percent,
                "Outdated": False,
            })

        return self.grouping_analyzer.enrich_rows(normalized_rows)

    def _load_inventory_details(self, scan_id: int) -> List[Dict[str, Any]]:
        """
        Load detailed computer/software inventory rows.

        Args:
            scan_id: Scan ID.

        Returns:
            Detailed joined inventory rows.
        """
        try:
            cursor = self.database_manager.connection.cursor()
            cursor.execute("""
                SELECT
                    c.hostname,
                    c.ip_address,
                    c.domain,
                    c.inventory_status,
                    s.display_name,
                    s.display_version,
                    s.publisher,
                    s.install_date,
                    s.architecture,
                    s.install_location
                FROM software s
                JOIN computers c
                    ON c.computer_id = s.computer_id
                WHERE s.scan_id = ?
                ORDER BY
                    LOWER(c.hostname),
                    c.ip_address,
                    LOWER(s.display_name),
                    LOWER(s.display_version)
            """, (scan_id,))
            return [dict(row) for row in cursor.fetchall()]
        except sqlite3.Error as exc:
            AppLogger.log_message(
                "error",
                f"Inventory detail export query failed: {exc}",
            )
            return [], {}

    def _build_dashboard_sheet(
        self,
        worksheet: Any,
        summary: Dict[str, Any],
        software_rows: List[Dict[str, Any]],
        computer_rows: List[Dict[str, Any]],
    ) -> None:
        """
        Build Dashboard worksheet.

        Args:
            worksheet: Dashboard worksheet.
            summary: Scan summary.
            software_rows: Software summary rows.
            computer_rows: Computer rows.
        """
        self._setup_sheet(worksheet)
        worksheet.sheet_view.showGridLines = False

        worksheet.merge_cells("A1:H1")
        title_cell = worksheet["A1"]
        title_cell.value = "Inventory Scan Dashboard"
        title_cell.font = Font(size=20, bold=True, color=self.THEME["white"])
        title_cell.fill = PatternFill("solid", fgColor=self.THEME["title"])
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 32

        self._write_navigation_links(worksheet)

        self._write_card_table(
            worksheet,
            start_row=4,
            start_col=1,
            title="Scan Metadata",
            rows=[
                ["Scan ID", summary.get("scan_id", "")],
                ["Status", summary.get("status", "")],
                ["Started", summary.get("started_at", "")],
                ["Completed", summary.get("completed_at", "")],
                ["IP Range", summary.get("ip_range", "")],
                ["Filters", summary.get("filters", "")],
            ],
        )

        self._write_card_table(
            worksheet,
            start_row=4,
            start_col=4,
            title="Key Metrics",
            rows=[
                ["Pingable", summary.get("pingable_count", "")],
                ["Matched", summary.get("matched_count", "")],
                ["Inventoried", summary.get("successful_inventory_count", "")],
                ["Failed Inventory", summary.get("failed_inventory_count", "")],
                ["Unique Software", len(software_rows)],
                ["Computer Rows", len(computer_rows)],
            ],
        )

        status_counts = self._computer_status_counts(computer_rows)
        status_rows = [
            ["Status", "Count"],
            ["Success", status_counts.get(AppConfig.INVENTORY_STATUS_SUCCESS, 0)],
            ["Partial", status_counts.get(AppConfig.INVENTORY_STATUS_PARTIAL, 0)],
            ["Failed", status_counts.get(AppConfig.INVENTORY_STATUS_FAILED, 0)],
            ["Other", status_counts.get("other", 0)],
        ]

        worksheet["A14"] = "Inventory Status Breakdown"
        self._style_section_title(worksheet["A14"])
        status_range = self._write_matrix(
            worksheet,
            start_row=15,
            start_col=1,
            rows=status_rows,
        )
        self._style_table_block(worksheet, status_range)

        top_software = sorted(
            software_rows,
            key=lambda row: int(row.get("OccurrenceCount", 0) or 0),
            reverse=True,
        )[:10]
        top_rows = [["Software", "Occurrences"]]
        top_rows.extend([
            [
                row.get("DisplayName", ""),
                int(row.get("OccurrenceCount", 0) or 0),
            ]
            for row in top_software
        ])

        worksheet["D14"] = "Top Software by Occurrence"
        self._style_section_title(worksheet["D14"])
        top_range = self._write_matrix(
            worksheet,
            start_row=15,
            start_col=4,
            rows=top_rows,
        )
        self._style_table_block(worksheet, top_range)

        self._add_dashboard_charts(
            worksheet,
            status_start_row=15,
            status_end_row=15 + len(status_rows) - 1,
            top_start_row=15,
            top_end_row=15 + len(top_rows) - 1,
        )
        self._add_dashboard_notes(worksheet)

        self._set_column_widths(
            worksheet,
            {
                "A": 24,
                "B": 22,
                "C": 4,
                "D": 42,
                "E": 18,
                "F": 4,
                "G": 24,
                "H": 28,
            },
        )

    def _build_software_summary_sheet(
        self,
        worksheet: Any,
        software_rows: List[Dict[str, Any]],
    ) -> None:
        """
        Build Software Summary worksheet.

        Args:
            worksheet: Software summary worksheet.
            software_rows: Software summary rows.
        """
        self._setup_sheet(worksheet)

        headers = [
            "Publisher Group",
            "Software",
            "Version",
            "Publisher",
            "Occurrence Count",
            "Grouped Version Count",
            "% Inventoried",
            "Outdated",
        ]

        data_rows = [
            [
                row.get("PublisherCategory", ""),
                row.get("DisplayName", ""),
                row.get("DisplayVersion", ""),
                row.get("Publisher", ""),
                int(row.get("OccurrenceCount", 0) or 0),
                int(row.get("GroupedVersionOccurrenceCount", 0) or 0),
                float(row.get("PercentOfInventoriedComputers", 0) or 0),
                "Yes" if row.get("Outdated") else "No",
            ]
            for row in software_rows
        ]

        if not data_rows:
            data_rows = [["", "", "", "", 0, 0, 0, "No"]]

        self._write_table_sheet(
            worksheet=worksheet,
            headers=headers,
            rows=data_rows,
            table_name="SoftwareSummaryTable",
            style_name="TableStyleMedium2",
        )

        max_row = worksheet.max_row
        self._format_column_number(worksheet, "G", 2, max_row, "0.00%")
        self._add_data_bar(worksheet, f"E2:E{max_row}")
        self._add_data_bar(worksheet, f"F2:F{max_row}")

        worksheet.conditional_formatting.add(
            f"H2:H{max_row}",
            CellIsRule(
                operator="equal",
                formula=['"Yes"'],
                fill=PatternFill("solid", fgColor=self.THEME["error"]),
            ),
        )

        self._set_column_widths(
            worksheet,
            {
                "A": 22,
                "B": 44,
                "C": 18,
                "D": 32,
                "E": 18,
                "F": 22,
                "G": 16,
                "H": 12,
            },
        )

    def _build_computers_sheet(
        self,
        worksheet: Any,
        computer_rows: List[Dict[str, Any]],
    ) -> None:
        """
        Build Computers worksheet.

        Args:
            worksheet: Computers worksheet.
            computer_rows: Computer rows.
        """
        self._setup_sheet(worksheet)

        headers = [
            "Computer ID",
            "Hostname",
            "IP Address",
            "FQDN",
            "Domain",
            "Pingable",
            "Matched Filter",
            "Inventory Status",
            "Software Count",
            "Last Seen",
            "Operating System",
            "Manufacturer",
            "Model",
            "Serial Number",
            "Inventory Error",
        ]

        data_rows = [
            [
                row.get("computer_id", ""),
                row.get("hostname", ""),
                row.get("ip_address", ""),
                row.get("fqdn", ""),
                row.get("domain", ""),
                bool(row.get("pingable", False)),
                bool(row.get("matched_filter", False)),
                row.get("inventory_status", ""),
                int(row.get("software_count", 0) or 0),
                row.get("last_seen", ""),
                row.get("operating_system", ""),
                row.get("manufacturer", ""),
                row.get("model", ""),
                row.get("serial_number", ""),
                row.get("inventory_error", ""),
            ]
            for row in computer_rows
        ]

        if not data_rows:
            data_rows = [["", "", "", "", "", False, False, "", 0, "", "", "", "", "", ""]]

        self._write_table_sheet(
            worksheet=worksheet,
            headers=headers,
            rows=data_rows,
            table_name="ComputersTable",
            style_name="TableStyleMedium2",
        )

        max_row = worksheet.max_row
        self._add_data_bar(worksheet, f"I2:I{max_row}")
        self._add_status_rules(worksheet, f"H2:H{max_row}")

        self._set_column_widths(
            worksheet,
            {
                "A": 12,
                "B": 24,
                "C": 18,
                "D": 34,
                "E": 20,
                "F": 12,
                "G": 16,
                "H": 18,
                "I": 16,
                "J": 22,
                "K": 32,
                "L": 22,
                "M": 24,
                "N": 22,
                "O": 48,
            },
        )

        self._wrap_column(worksheet, "O", 2, max_row)

    def _build_inventory_details_sheet(
        self,
        worksheet: Any,
        inventory_rows: List[Dict[str, Any]],
    ) -> None:
        """
        Build detailed computer/software inventory worksheet.

        Args:
            worksheet: Inventory details worksheet.
            inventory_rows: Detailed inventory rows.
        """
        self._setup_sheet(worksheet)

        headers = [
            "Hostname",
            "IP Address",
            "Domain",
            "Inventory Status",
            "Software",
            "Version",
            "Publisher",
            "Install Date",
            "Architecture",
            "Install Location",
        ]

        data_rows = [
            [
                row.get("hostname", ""),
                row.get("ip_address", ""),
                row.get("domain", ""),
                row.get("inventory_status", ""),
                row.get("display_name", ""),
                row.get("display_version", ""),
                row.get("publisher", ""),
                row.get("install_date", ""),
                row.get("architecture", ""),
                row.get("install_location", ""),
            ]
            for row in inventory_rows
        ]

        if not data_rows:
            data_rows = [["", "", "", "", "", "", "", "", "", ""]]

        self._write_table_sheet(
            worksheet=worksheet,
            headers=headers,
            rows=data_rows,
            table_name="InventoryDetailsTable",
            style_name="TableStyleMedium2",
        )

        max_row = worksheet.max_row
        self._add_status_rules(worksheet, f"D2:D{max_row}")

        self._set_column_widths(
            worksheet,
            {
                "A": 24,
                "B": 18,
                "C": 20,
                "D": 18,
                "E": 44,
                "F": 18,
                "G": 32,
                "H": 16,
                "I": 16,
                "J": 48,
            },
        )
        self._wrap_column(worksheet, "J", 2, max_row)

    def _build_export_notes_sheet(
        self,
        worksheet: Any,
        scan_id: int,
        output_path: Path,
    ) -> None:
        """
        Build export metadata and notes worksheet.

        Args:
            worksheet: Export notes worksheet.
            scan_id: Scan ID.
            output_path: Output path.
        """
        self._setup_sheet(worksheet)
        worksheet.sheet_view.showGridLines = False

        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "Export Notes"
        worksheet["A1"].font = Font(size=18, bold=True, color=self.THEME["white"])
        worksheet["A1"].fill = PatternFill("solid", fgColor=self.THEME["title"])
        worksheet["A1"].alignment = Alignment(horizontal="center")

        notes = [
            ["Scan ID", scan_id],
            ["Export Path", str(output_path)],
            [
                "Generated At",
                datetime.now().strftime(AppConfig.STORAGE_TIMESTAMP_FORMAT),
            ],
            ["Format", "XLSX"],
            [
                "Notes",
                (
                    "Dashboard, Software Summary, Computers, and Inventory "
                    "Details are exported as separate worksheets with filters."
                ),
            ],
        ]

        self._write_card_table(
            worksheet,
            start_row=3,
            start_col=1,
            title="Workbook Metadata",
            rows=notes,
        )

        self._set_column_widths(
            worksheet,
            {
                "A": 24,
                "B": 80,
                "C": 18,
                "D": 18,
            },
        )

    def _write_table_sheet(
        self,
        worksheet: Any,
        headers: List[str],
        rows: List[List[Any]],
        table_name: str,
        style_name: str,
    ) -> None:
        """
        Write a standard table worksheet.

        Args:
            worksheet: Target worksheet.
            headers: Header row values.
            rows: Data rows.
            table_name: Excel table name.
            style_name: Built-in Excel table style.
        """
        worksheet.append(headers)

        for row in rows:
            worksheet.append(row)

        max_row = max(1, worksheet.max_row)
        max_col = max(1, len(headers))
        table_range = f"A1:{get_column_letter(max_col)}{max_row}"

        self._style_header_row(worksheet, 1, max_col)
        self._add_table(
            worksheet,
            table_range,
            table_name=table_name,
            style_name=style_name,
        )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = table_range

    def _write_matrix(
        self,
        worksheet: Any,
        start_row: int,
        start_col: int,
        rows: List[List[Any]],
    ) -> str:
        """
        Write a matrix of values cell-by-cell.

        Args:
            worksheet: Target worksheet.
            start_row: Starting row.
            start_col: Starting column.
            rows: Matrix rows.

        Returns:
            A1 range written.
        """
        if not rows:
            rows = [[""]]

        for row_offset, row_values in enumerate(rows):
            for col_offset, value in enumerate(row_values):
                worksheet.cell(
                    row=start_row + row_offset,
                    column=start_col + col_offset,
                    value=value,
                )

        end_row = start_row + len(rows) - 1
        end_col = start_col + max(len(row) for row in rows) - 1

        return (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )

    def _setup_sheet(self, worksheet: Any) -> None:
        """
        Apply base worksheet settings.

        Args:
            worksheet: Target worksheet.
        """
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"

    def _write_navigation_links(self, worksheet: Any) -> None:
        """Write dashboard navigation links."""
        links = [
            ("A2", "Dashboard", "#'Dashboard'!A1"),
            ("B2", "Software Summary", "#'Software Summary'!A1"),
            ("C2", "Computers", "#'Computers'!A1"),
            ("D2", "Inventory Details", "#'Inventory Details'!A1"),
            ("E2", "Export Notes", "#'Export Notes'!A1"),
        ]

        for cell_ref, label, target in links:
            cell = worksheet[cell_ref]
            cell.value = label
            cell.hyperlink = target
            cell.style = "Hyperlink"
            cell.alignment = Alignment(horizontal="center")

    def _write_card_table(
        self,
        worksheet: Any,
        start_row: int,
        start_col: int,
        title: str,
        rows: List[List[Any]],
    ) -> None:
        """
        Write a formatted two-column card table.

        Args:
            worksheet: Target worksheet.
            start_row: Top row.
            start_col: Left column.
            title: Card title.
            rows: Label/value rows.
        """
        title_cell = worksheet.cell(start_row, start_col)
        title_cell.value = title
        self._style_section_title(title_cell)

        for offset, row_values in enumerate(rows, start=1):
            label_cell = worksheet.cell(start_row + offset, start_col)
            value_cell = worksheet.cell(start_row + offset, start_col + 1)

            label_cell.value = row_values[0]
            value_cell.value = row_values[1]

            label_cell.fill = PatternFill("solid", fgColor=self.THEME["neutral"])
            label_cell.font = Font(bold=True, color=self.THEME["text"])
            value_cell.fill = PatternFill("solid", fgColor=self.THEME["white"])
            value_cell.alignment = Alignment(wrap_text=True)

            for cell in (label_cell, value_cell):
                cell.border = self._thin_border()

    def _style_section_title(self, cell: Any) -> None:
        """
        Style a section title cell.

        Args:
            cell: Target cell.
        """
        cell.font = Font(bold=True, size=12, color=self.THEME["white"])
        cell.fill = PatternFill("solid", fgColor=self.THEME["header_dark"])
        cell.alignment = Alignment(horizontal="center")

    def _style_header_row(
        self,
        worksheet: Any,
        row_index: int,
        column_count: int,
    ) -> None:
        """
        Style a worksheet header row.

        Args:
            worksheet: Target worksheet.
            row_index: Header row index.
            column_count: Header column count.
        """
        for column_index in range(1, column_count + 1):
            cell = worksheet.cell(row_index, column_index)
            cell.font = Font(bold=True, color=self.THEME["white"])
            cell.fill = PatternFill("solid", fgColor=self.THEME["title"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._thin_border()

    def _style_table_block(self, worksheet: Any, cell_range: str) -> None:
        """
        Style a compact dashboard table block.

        Args:
            worksheet: Target worksheet.
            cell_range: A1-style range.
        """
        for row_index, row in enumerate(worksheet[cell_range], start=1):
            for cell in row:
                cell.border = self._thin_border()
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                if row_index == 1:
                    cell.font = Font(bold=True, color=self.THEME["white"])
                    cell.fill = PatternFill("solid", fgColor=self.THEME["title"])
                else:
                    cell.fill = PatternFill("solid", fgColor=self.THEME["white"])

    def _add_table(
        self,
        worksheet: Any,
        cell_range: str,
        table_name: str,
        style_name: str,
    ) -> None:
        """
        Add an Excel table to a worksheet.

        Args:
            worksheet: Target worksheet.
            cell_range: Table range.
            table_name: Unique table name.
            style_name: Excel built-in table style.
        """
        safe_table_name = "".join(
            char for char in table_name
            if char.isalnum() or char == "_"
        )

        table = Table(displayName=safe_table_name, ref=cell_range)
        style = TableStyleInfo(
            name=style_name,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)

    def _add_dashboard_charts(
        self,
        worksheet: Any,
        status_start_row: int,
        status_end_row: int,
        top_start_row: int,
        top_end_row: int,
    ) -> None:
        """
        Add dashboard charts.

        Args:
            worksheet: Dashboard worksheet.
            status_start_row: Status helper table start row.
            status_end_row: Status helper table end row.
            top_start_row: Top software helper table start row.
            top_end_row: Top software helper table end row.
        """
        if status_end_row > status_start_row:
            pie_chart = PieChart()
            pie_chart.title = "Inventory Status"
            pie_data = Reference(
                worksheet,
                min_col=2,
                min_row=status_start_row + 1,
                max_row=status_end_row,
            )
            pie_labels = Reference(
                worksheet,
                min_col=1,
                min_row=status_start_row + 1,
                max_row=status_end_row,
            )
            pie_chart.add_data(pie_data, titles_from_data=False)
            pie_chart.set_categories(pie_labels)
            pie_chart.height = 7
            pie_chart.width = 9
            worksheet.add_chart(pie_chart, "A22")

        if top_end_row > top_start_row:
            bar_chart = BarChart()
            bar_chart.title = "Top Software Occurrences"
            bar_chart.y_axis.title = "Software"
            bar_chart.x_axis.title = "Occurrences"
            bar_chart.type = "bar"
            bar_chart.style = 10

            data = Reference(
                worksheet,
                min_col=5,
                min_row=top_start_row,
                max_row=top_end_row,
            )
            labels = Reference(
                worksheet,
                min_col=4,
                min_row=top_start_row + 1,
                max_row=top_end_row,
            )
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(labels)
            bar_chart.height = 9
            bar_chart.width = 16
            worksheet.add_chart(bar_chart, "D22")

    def _add_dashboard_notes(self, worksheet: Any) -> None:
        """
        Add explanatory dashboard notes.

        Args:
            worksheet: Dashboard worksheet.
        """
        worksheet["G4"] = "Workbook Tips"
        self._style_section_title(worksheet["G4"])

        notes = [
            "Use table filters on each sheet for quick narrowing.",
            "Software occurrence bars show the most widespread software.",
            "Outdated rows indicate versions older than the latest detected version.",
            "Computer status colors match the application UI style.",
        ]

        for offset, note in enumerate(notes, start=5):
            cell = worksheet.cell(offset, 7)
            cell.value = note
            cell.fill = PatternFill("solid", fgColor=self.THEME["neutral"])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = self._thin_border()

    def _add_data_bar(self, worksheet: Any, cell_range: str) -> None:
        """
        Add data bar conditional formatting.

        Args:
            worksheet: Target worksheet.
            cell_range: Cell range.
        """
        worksheet.conditional_formatting.add(
            cell_range,
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="max",
                color="5B9BD5",
            ),
        )

    def _add_status_rules(self, worksheet: Any, cell_range: str) -> None:
        """
        Add status-based conditional formatting.

        Args:
            worksheet: Target worksheet.
            cell_range: Status cell range.
        """
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{AppConfig.INVENTORY_STATUS_SUCCESS}"'],
                fill=PatternFill("solid", fgColor=self.THEME["success"]),
            ),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{AppConfig.INVENTORY_STATUS_PARTIAL}"'],
                fill=PatternFill("solid", fgColor=self.THEME["warning"]),
            ),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{AppConfig.INVENTORY_STATUS_FAILED}"'],
                fill=PatternFill("solid", fgColor=self.THEME["error"]),
            ),
        )

    def _format_column_number(
        self,
        worksheet: Any,
        column_letter: str,
        start_row: int,
        end_row: int,
        number_format: str,
    ) -> None:
        """
        Apply number format to cells one-by-one.

        Args:
            worksheet: Target worksheet.
            column_letter: Column letter.
            start_row: First row.
            end_row: Last row.
            number_format: Excel number format.
        """
        column_index = openpyxl.utils.column_index_from_string(column_letter)

        for row_index in range(start_row, end_row + 1):
            worksheet.cell(row_index, column_index).number_format = number_format

    def _wrap_column(
        self,
        worksheet: Any,
        column_letter: str,
        start_row: int,
        end_row: int,
    ) -> None:
        """
        Enable text wrapping for a column range.

        Args:
            worksheet: Target worksheet.
            column_letter: Column letter.
            start_row: First row.
            end_row: Last row.
        """
        column_index = openpyxl.utils.column_index_from_string(column_letter)

        for row_index in range(start_row, end_row + 1):
            worksheet.cell(row_index, column_index).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    def _set_column_widths(
        self,
        worksheet: Any,
        widths: Dict[str, int],
    ) -> None:
        """
        Set worksheet column widths.

        Args:
            worksheet: Target worksheet.
            widths: Mapping of column letter to width.
        """
        for column_letter, width in widths.items():
            worksheet.column_dimensions[column_letter].width = width

    def _computer_status_counts(
        self,
        computer_rows: List[Dict[str, Any]],
    ) -> Dict[str, int]:
        """
        Count inventory statuses.

        Args:
            computer_rows: Computer rows.

        Returns:
            Status count dictionary.
        """
        counts = {
            AppConfig.INVENTORY_STATUS_SUCCESS: 0,
            AppConfig.INVENTORY_STATUS_PARTIAL: 0,
            AppConfig.INVENTORY_STATUS_FAILED: 0,
            "other": 0,
        }

        for row in computer_rows:
            status = str(row.get("inventory_status", "")).strip().lower()

            if status in counts:
                counts[status] += 1
            else:
                counts["other"] += 1

        return counts

    def _thin_border(self) -> Border:
        """
        Build a standard thin border.

        Returns:
            OpenPyXL Border.
        """
        side = Side(style="thin", color=self.THEME["border"])
        return Border(left=side, right=side, top=side, bottom=side)

# ------------------ Software Grouping Analyzer ------------------ #
class SoftwareGroupingAnalyzer:
    """
    Enrich and group software rows for Software Summary display.

    This class supports two display modes:
        - Outdated mode enabled: keep version-specific rows so stale versions
          are visible.
        - Outdated mode disabled: group matching software names/publishers and
          display a version range when multiple versions exist.

    All methods are CPU-local and avoid database access so they can be safely
    called before rendering the Treeview.
    """

    SYSTEM_CATEGORY = "System / Microsoft"

    SYSTEM_PUBLISHER_NAMES = {
        "microsoft",
        "microsoft corporation",
        "microsoft windows",
        "microsoft inc.",
    }

    COMPANY_SUFFIXES = (
        "inc",
        "inc.",
        "llc",
        "l.l.c.",
        "ltd",
        "ltd.",
        "limited",
        "corp",
        "corp.",
        "corporation",
        "company",
        "co",
        "co.",
        "gmbh",
        "ag",
        "sa",
        "s.a.",
        "plc",
    )

    @classmethod
    def enrich_rows(
        cls,
        rows: Sequence[Dict[str, Any]],
        group_versions: bool = False,
    ) -> List[Dict[str, Any]]:
        """
        Enrich software rows with publisher category and outdated markers.

        Args:
            rows: Normalized software rows.
            group_versions: True to collapse versions by software name.

        Returns:
            Enriched rows ready for UI rendering.
        """
        normalized_rows = [dict(row) for row in rows]

        for row in normalized_rows:
            row["PublisherCategory"] = cls.assign_publisher_category(
                row.get("Publisher", "")
            )
            row["VersionSortKey"] = cls._version_tuple(
                row.get("DisplayVersion", "")
            )

        cls._mark_outdated_versions(normalized_rows)

        if group_versions:
            return cls.group_rows_by_name(normalized_rows)

        normalized_rows.sort(
            key=lambda row: (
                cls._normalize_text(row.get("PublisherCategory", "")),
                cls._normalize_text(row.get("DisplayName", "")),
                cls._version_tuple(row.get("DisplayVersion", "")),
                cls._normalize_text(row.get("Publisher", "")),
            )
        )
        return normalized_rows

    @classmethod
    def group_rows_by_name(
        cls,
        rows: Sequence[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """
        Group version-specific rows into one display row per software title.

        Args:
            rows: Enriched software rows.

        Returns:
            Grouped software rows.
        """
        grouped_rows: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}

        for row in rows:
            key = (
                cls._normalize_text(row.get("DisplayName", "")),
                cls._normalize_text(row.get("Publisher", "")),
            )
            grouped_rows.setdefault(key, []).append(dict(row))

        output_rows: List[Dict[str, Any]] = []

        for group in grouped_rows.values():
            if not group:
                continue

            group.sort(
                key=lambda row: cls._version_tuple(row.get("DisplayVersion", ""))
            )

            base_row = dict(group[-1])
            version_values = [
                str(row.get("DisplayVersion", "") or "").strip()
                for row in group
                if str(row.get("DisplayVersion", "") or "").strip()
            ]
            unique_versions = list(dict.fromkeys(version_values))

            occurrence_count = max(
                int(row.get("GroupedVersionOccurrenceCount", 0) or 0)
                for row in group
            )
            fallback_count = max(
                int(row.get("OccurrenceCount", 0) or 0)
                for row in group
            )
            grouped_count = occurrence_count or fallback_count

            base_row["OccurrenceCount"] = grouped_count
            base_row["GroupedVersionOccurrenceCount"] = grouped_count
            base_row["GroupedVersions"] = True
            base_row["GroupedSourceRows"] = group
            base_row["Outdated"] = any(bool(row.get("Outdated")) for row in group)

            if len(unique_versions) > 1:
                base_row["DisplayVersion"] = (
                    f"{unique_versions[0]} - {unique_versions[-1]}"
                )
                base_row["LowestVersion"] = unique_versions[0]
                base_row["HighestVersion"] = unique_versions[-1]
            elif unique_versions:
                base_row["DisplayVersion"] = unique_versions[0]
                base_row["LowestVersion"] = unique_versions[0]
                base_row["HighestVersion"] = unique_versions[0]
            else:
                base_row["DisplayVersion"] = ""
                base_row["LowestVersion"] = ""
                base_row["HighestVersion"] = ""

            output_rows.append(base_row)

        output_rows.sort(
            key=lambda row: (
                cls._normalize_text(row.get("PublisherCategory", "")),
                cls._normalize_text(row.get("DisplayName", "")),
                cls._normalize_text(row.get("Publisher", "")),
            )
        )
        return output_rows

    @classmethod
    def assign_publisher_category(cls, publisher: str) -> str:
        """
        Assign a stable publisher grouping label.

        Args:
            publisher: Raw publisher value.

        Returns:
            Publisher category.
        """
        normalized = cls._normalize_text(publisher)

        if not normalized:
            return "Unknown Publisher"

        if normalized in cls.SYSTEM_PUBLISHER_NAMES:
            return cls.SYSTEM_CATEGORY

        tokens = cls._publisher_tokens(normalized)

        if not tokens:
            return publisher.strip() or "Unknown Publisher"

        return " ".join(token.capitalize() for token in tokens[:2])

    @classmethod
    def summarize_occurrence_range(
        cls,
        rows: Sequence[Dict[str, Any]],
    ) -> str:
        """
        Summarize grouped-version occurrence counts for a publisher row.

        Args:
            rows: Child software rows.

        Returns:
            Single count or range string.
        """
        grouped_counts: Dict[Tuple[str, str], int] = {}

        for row in rows:
            key = (
                cls._normalize_text(row.get("DisplayName", "")),
                cls._normalize_text(row.get("Publisher", "")),
            )
            grouped_count = int(
                row.get(
                    "GroupedVersionOccurrenceCount",
                    row.get("OccurrenceCount", 0),
                ) or 0
            )
            grouped_counts[key] = max(grouped_counts.get(key, 0), grouped_count)

        values = sorted(set(grouped_counts.values()))

        if not values:
            return "0"

        if len(values) == 1:
            return str(values[0])

        return f"{values[0]} - {values[-1]}"

    @classmethod
    def _mark_outdated_versions(cls, rows: List[Dict[str, Any]]) -> None:
        """
        Mark rows as outdated when a newer version exists for same title.

        Args:
            rows: Rows to mutate.
        """
        grouped_rows: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}

        for row in rows:
            key = (
                cls._normalize_text(row.get("DisplayName", "")),
                cls._normalize_text(row.get("Publisher", "")),
            )
            grouped_rows.setdefault(key, []).append(row)

        for group in grouped_rows.values():
            if len(group) < 2:
                for row in group:
                    row["Outdated"] = False
                continue

            latest_version = max(
                cls._version_tuple(row.get("DisplayVersion", ""))
                for row in group
            )

            for row in group:
                row["Outdated"] = (
                    cls._version_tuple(row.get("DisplayVersion", ""))
                    < latest_version
                )

    @classmethod
    def _publisher_tokens(cls, normalized_publisher: str) -> List[str]:
        """
        Tokenize publisher name and remove legal suffixes.

        Args:
            normalized_publisher: Normalized publisher text.

        Returns:
            Publisher tokens.
        """
        tokens = [
            token.strip("., ")
            for token in normalized_publisher.split()
            if token.strip("., ")
        ]

        return [
            token for token in tokens
            if token not in cls.COMPANY_SUFFIXES
        ]

    @classmethod
    def _normalize_text(cls, value: Any) -> str:
        """
        Normalize text for stable grouping.

        Args:
            value: Raw value.

        Returns:
            Lowercase whitespace-normalized text.
        """
        return " ".join(str(value or "").strip().lower().split())

    @classmethod
    def _version_tuple(cls, value: Any) -> Tuple[Any, ...]:
        """
        Convert a version string into a comparable tuple.

        Args:
            value: Raw version value.

        Returns:
            Sortable version tuple.
        """
        text = str(value or "").strip().lower()

        if not text:
            return tuple()

        parts: List[Any] = []
        current = ""

        for char in text:
            if char.isalnum():
                current += char
                continue

            if current:
                parts.append(cls._coerce_version_part(current))
                current = ""

        if current:
            parts.append(cls._coerce_version_part(current))

        return tuple(parts)

    @staticmethod
    def _coerce_version_part(value: str) -> Any:
        """
        Convert numeric version fragments to integers.

        Args:
            value: Version fragment.

        Returns:
            Integer or string fragment.
        """
        try:
            return int(value)
        except ValueError:
            return value

# ------------------ Main Application ------------------ #
class MainApplication(tk.Tk):
    """
    Root Tkinter application.
    """

    def __init__(self):
        """Initialize the main application."""
        super().__init__()

        self.title(AppConfig.APP_NAME)
        self.geometry(AppConfig.DEFAULT_WINDOW_GEOMETRY)
        self.configure(bg=AppConfig.DEFAULT_BACKGROUND)

        self.result_queue: queue.Queue = queue.Queue()
        self.current_scan_id: Optional[int] = None
        self.is_scan_running = False
        self.is_credential_prompt_running = False
        self.is_export_running = False

        self.create_services()
        self.create_widgets()

        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.after(100, self.process_result_queue)
        self.after(100, self.process_log_queue)

        AppLogger.log_message("info", "Application started.")

    def create_services(self) -> None:
        """Create application services."""
        try:
            AppLogger.configure_logging()

            self.preferences_manager = PreferencesManager()
            self.database_manager = DatabaseManager()
            self.credential_manager = CredentialManager()
            self.export_manager = ExcelExportManager(self.database_manager)

            self.scan_coordinator = ScanCoordinator(
                database_manager=self.database_manager,
                result_queue=self.result_queue,
                credential_manager=self.credential_manager,
                preferences_manager=self.preferences_manager,
            )

        except Exception as exc:
            AppLogger.log_message(
                "critical",
                f"Application service initialization failed: {exc}",
            )
            raise

    def create_widgets(self) -> None:
        """Create main application widgets."""
        self._configure_styles()

        self.layout = MainWindowLayout(
            self,
            preferences_manager=self.preferences_manager,
            database_manager=self.database_manager,
            log_queue=AppLogger.get_log_queue(),
            start_callback=self._start_scan,
            cancel_callback=self._cancel_scan,
            clear_callback=self._clear_results,
            export_callback=self._export_xlsx,
            credential_callback=self._prompt_for_credentials,
            show_credentials_button=not self.credential_manager.is_admin(),
        )

        self.layout.set_admin_status(self.credential_manager.is_admin())
        self.layout.set_credentials_cached(
            self.credential_manager.has_cached_credentials()
        )

        if not self.credential_manager.is_admin():
            self.layout.update_status(
                phase="limited",
                message=(
                    "Running without admin privileges. Use Credentials before "
                    "inventory scans that require elevated remote access."
                ),
            )

    def on_closing(self) -> None:
        """Gracefully close the application and unload runtime resources."""
        try:
            if self.is_scan_running:
                self.scan_coordinator.cancel_scan()
                AppLogger.log_message(
                    "warning",
                    "Application closing while scan is running; cancellation requested.",
                )

            self.scan_coordinator.shutdown()
            self.layout.save_current_preferences()
            self.preferences_manager.set("window_geometry", self.geometry())
            self.credential_manager.clear_cached_credentials()

            if self.database_manager.connection:
                self.database_manager.connection.close()

            AppLogger.log_message("info", "Application closed.")
        except Exception as exc:
            AppLogger.log_message(
                "error",
                f"Error during application shutdown: {exc}",
            )
        finally:
            self.destroy()

    def process_result_queue(self) -> None:
        """Process scan result queue on the main thread."""
        try:
            while True:
                item = self.result_queue.get_nowait()

                if isinstance(item, ScanProgress):
                    self._handle_scan_progress(item)
                elif isinstance(item, ScanResult):
                    self._handle_scan_result(item)

        except queue.Empty:
            pass
        except Exception as exc:
            AppLogger.log_message(
                "error",
                f"Failed to process result queue: {exc}",
            )

        if self.winfo_exists():
            self.after(100, self.process_result_queue)

    def process_log_queue(self) -> None:
        """Process log queue for the Logs tab."""
        try:
            self.layout.process_log_queue()
        except Exception as exc:
            AppLogger.log_message(
                "error",
                f"Failed to process log queue: {exc}",
            )

        if self.winfo_exists():
            self.after(150, self.process_log_queue)

    def _prompt_for_credentials(self) -> None:
        """Prompt for credentials without blocking the Tkinter GUI."""
        if self.is_credential_prompt_running:
            self.layout.update_status(
                phase="credentials",
                message="Credential prompt is already running.",
            )
            return

        self.is_credential_prompt_running = True
        self._set_credential_button_state("disabled")
        self.layout.update_status(
            phase="credentials",
            message="Credential prompt started. Complete the PowerShell credential window.",
        )

        thread = threading.Thread(
            target=self._credential_prompt_worker,
            name="CredentialPromptWorker",
            daemon=True,
        )
        thread.start()

    def _credential_prompt_worker(self) -> None:
        """Run credential prompt in a worker thread."""
        success = self.credential_manager.prompt_for_credentials()

        self.result_queue.put(
            ScanResult(
                task_name="Credentials",
                status=(
                    AppConfig.SCAN_STATUS_SUCCESS
                    if success
                    else AppConfig.SCAN_STATUS_WARNING
                ),
                message=(
                    "PowerShell credentials cached for this session."
                    if success
                    else "Credential prompt cancelled or failed."
                ),
                data={"credential_prompt_complete": True},
            )
        )

    def _start_scan(self, options: Dict[str, Any]) -> None:
        """
        Start scan from UI options.

        Args:
            options: Scan options from ScanControlPanel.get_scan_options.
        """
        try:
            safe_options = dict(options or {})

            if (
                self.credential_manager.credentials_required()
                and not self.credential_manager.has_cached_credentials()
            ):
                self.layout.update_status(
                    phase="credentials",
                    message=(
                        "Credentials are not cached. Click Credentials first "
                        "if remote inventory requires delegated credentials."
                    ),
                )

            self.layout.reset_scan_progress()

            started = self.scan_coordinator.start_scan(
                ip_range=safe_options.get("ip_range", ""),
                filters=safe_options.get("filters", ""),
                options=safe_options,
            )

            if not started:
                self.layout.clear_results()
                return

            self.is_scan_running = True
            self.layout.set_scan_running(True)
            self.layout.update_status(
                phase="scanning",
                message=(
                    f"Scanning {safe_options.get('ip_range', '')} with "
                    f"{safe_options.get('max_ping_workers')} ping workers and "
                    f"{safe_options.get('max_inventory_workers')} inventory workers."
                ),
            )
            AppLogger.log_message("info", "Scan started.")

        except Exception as exc:
            AppLogger.log_message("error", f"Failed to start scan: {exc}")
            self.layout.update_status(
                phase="error",
                message=f"Failed to start scan: {exc}",
            )

    def _cancel_scan(self) -> None:
        """Cancel the active scan."""
        self.scan_coordinator.cancel_scan()
        self.layout.update_status(
            phase="cancelled",
            message="Cancellation requested.",
        )

    def _clear_results(self) -> None:
        """Clear displayed scan results but preserve logs."""
        self.layout.clear_results()
        self.current_scan_id = None
        self.layout.update_status(
            phase="ready",
            message="Results cleared.",
            selection_count=0,
        )

    def _export_xlsx(self) -> None:
        """Prompt for XLSX path and start export on a background thread."""
        if getattr(self, "is_export_running", False):
            self.layout.update_status(
                phase="export",
                message="An export is already running.",
            )
            AppLogger.log_message(
                "warning",
                "XLSX export request ignored because another export is running.",
            )
            return

        if not self.current_scan_id:
            self.layout.update_status(
                phase="export",
                message="No completed scan is available to export.",
            )
            AppLogger.log_message(
                "warning",
                "XLSX export requested without a completed scan.",
            )
            return

        initial_dir = self.preferences_manager.get(
            "last_export_directory",
            str(Path.cwd()),
        )

        output_path = filedialog.asksaveasfilename(
            initialdir=initial_dir,
            defaultextension=".xlsx",
            filetypes=(("Excel Workbook", "*.xlsx"), ("All Files", "*.*")),
        )

        if not output_path:
            return

        safe_output_path = str(Path(output_path).with_suffix(".xlsx"))
        self.preferences_manager.set(
            "last_export_directory",
            str(Path(safe_output_path).parent),
        )

        self.is_export_running = True
        self._set_export_running(True)

        self.layout.update_status(
            phase="export",
            message="Exporting XLSX workbook in background...",
        )
        AppLogger.log_message(
            "info",
            f"Background XLSX export started: {safe_output_path}",
        )

        export_thread = threading.Thread(
            target=self._export_xlsx_worker,
            args=(self.current_scan_id, safe_output_path),
            name="XlsxExportWorker",
            daemon=True,
        )
        export_thread.start()


    def _export_xlsx_worker(self, scan_id: int, output_path: str) -> None:
        """
        Run XLSX export away from Tkinter's main thread.

        Args:
            scan_id: Scan ID to export.
            output_path: Destination path.
        """
        result = self.export_manager.export_scan_to_xlsx(scan_id, output_path)
        self.after(0, lambda: self._handle_export_finished(result))


    def _handle_export_finished(self, result: ExportResult) -> None:
        """
        Apply export completion result on the Tkinter main thread.

        Args:
            result: Export result from worker thread.
        """
        self.is_export_running = False
        self._set_export_running(False)

        if result.success:
            self.layout.update_status(
                phase="export",
                message=result.to_display_text(),
            )
            self.layout.append_live_feed(
                f"XLSX export completed in {result.duration_seconds}s: "
                f"{result.output_path}"
            )
            return

        self.layout.update_status(
            phase="export",
            message=result.to_display_text(),
        )
        self.layout.append_live_feed(result.to_display_text())


    def _set_export_running(self, is_running: bool) -> None:
        """
        Update export-related UI state.

        Args:
            is_running: True while export is running.
        """
        try:
            self.layout.set_export_running(is_running)
        except AttributeError:
            AppLogger.log_message(
                "debug",
                "Layout does not expose set_export_running; skipping UI toggle.",
            )

    def _handle_scan_progress(self, progress: ScanProgress) -> None:
        """Handle ScanProgress queue item."""
        self.layout.update_progress(progress)

        feed_visible = True
        if progress.data:
            feed_visible = bool(progress.data.get("_feed_visible", True))
            self.layout.update_dashboard(progress.data)

        if feed_visible:
            self.layout.append_live_feed(progress.to_display_text())

        if progress.phase in {"complete", "cancelled"}:
            self.is_scan_running = False
            self.layout.set_scan_running(False)

    def _handle_scan_result(self, result: ScanResult) -> None:
        """Handle ScanResult queue item."""
        self.layout.append_live_feed(result.to_display_text())

        if result.task_name == "Credentials":
            self.is_credential_prompt_running = False
            self._set_credential_button_state("normal")
            self.layout.set_credentials_cached(
                self.credential_manager.has_cached_credentials()
            )
            self.layout.update_status(
                phase="credentials",
                message=result.message,
            )
            return

        if result.scan_id:
            self.current_scan_id = result.scan_id
            self.layout.results_notebook.current_scan_id = result.scan_id

        if result.data:
            self.layout.update_dashboard(result.data)

        if result.task_name == "Coordinator" and result.status in {
            AppConfig.SCAN_STATUS_SUCCESS,
            AppConfig.SCAN_STATUS_WARNING,
            AppConfig.SCAN_STATUS_ERROR,
            AppConfig.SCAN_STATUS_CANCELLED,
        }:
            self.is_scan_running = False
            self.layout.set_scan_running(False)

            if self.current_scan_id:
                self._refresh_result_tables(self.current_scan_id, result.data)

    def _refresh_result_tables(
        self,
        scan_id: int,
        summary: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Refresh software and computer result tables after scan completion."""
        try:
            software_rows = self.database_manager.get_software_occurrences(scan_id)
            successful_count = 0

            if summary:
                successful_count = int(
                    summary.get("successful_inventory_count", 0) or 0
                )

            self.layout.results_notebook.populate_software_summary(
                software_rows,
                successful_count,
            )

            self.layout.results_notebook.populate_computers(
                self.database_manager.get_all_computers_with_counts(scan_id)
            )

        except sqlite3.Error as exc:
            AppLogger.log_message(
                "error",
                f"Failed to refresh result tables: {exc}",
            )

    def _set_credential_button_state(self, state: str) -> None:
        """
        Safely update credential button state.

        Args:
            state: Tkinter button state.
        """
        try:
            control_panel = self.layout.scan_control_panel
            if hasattr(control_panel, "credential_button"):
                control_panel.credential_button.configure(state=state)
        except tk.TclError as exc:
            AppLogger.log_message(
                "error",
                f"Failed to update credential button state: {exc}",
            )

    def _configure_styles(self) -> None:
        """Configure basic ttk styling."""
        style = ttk.Style(self)

        try:
            style.theme_use("clam")
        except tk.TclError:
            AppLogger.log_message(
                "warning",
                "Unable to apply clam theme; using default theme.",
            )

        style.configure("App.TFrame", background=AppConfig.DEFAULT_BACKGROUND)
        style.configure("TFrame", background=AppConfig.DEFAULT_BACKGROUND)
        style.configure("TLabel", background=AppConfig.DEFAULT_BACKGROUND)
        style.configure("TButton", padding=5)

# ------------------ Entry Point ------------------ #
def main() -> None:
    """
    Main application entry point.
    """
    try:
        app = MainApplication()
        app.mainloop()
    except Exception as exc:
        AppLogger.log_message(
            "critical",
            f"Unrecoverable application failure: {exc}",
        )


if __name__ == "__main__":
    main()
